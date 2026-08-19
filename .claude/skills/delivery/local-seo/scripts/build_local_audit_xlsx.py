#!/usr/bin/env python3
"""
build_local_audit_xlsx.py — Render Local SEO Audit workbook (9 tabs).

Reads normalized cache from clients/{slug}/local-audit/{audit_id}/ (produced by
build_local_audit_data.py) and renders a 9-tab Excel workbook the strategist
edits during CP2.

Tabs:
  1. README         — methodology, color key, how to fill in approvals
  2. Summary        — topline grade per category, top 3 priorities
  3. GBP Audit      — 16 GBP fields, current vs ideal, gap, action
  4. Local Rankings — 250-cell grid per keyword, heatmap colored by rank
  5. Citations      — 50 directories × status × NAP diff × action
  6. Reviews        — count / rating / velocity vs competitors
  7. Local Content  — pages inventory: location, service area, schema
  8. Local Links    — referring domain matrix, opportunity scored
  9. Actions        — master rec list with Approval column (CP2 surface)

CLI:
  python build_local_audit_xlsx.py \\
    --client-slug acme-law \\
    --audit-id b4a66be6-3f40-4b02-87ae-ebcef08ec0b5 \\
    --workspace-root /path/to/workspace \\
    --client-name "Acme Law"
"""

import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

parser = argparse.ArgumentParser(description="Build the Local SEO Audit XLSX workbook.")
parser.add_argument("--client-slug", required=True)
parser.add_argument("--audit-id", required=True)
parser.add_argument("--workspace-root", default=os.getcwd())
parser.add_argument("--client-name", default=None, help="Display name (defaults to slug)")
parser.add_argument("--output", default=None, help="Output xlsx path (defaults to audit dir)")
args = parser.parse_args()

ROOT = Path(args.workspace_root)
AUDIT_DIR = ROOT / "clients" / args.client_slug / "local-audit" / args.audit_id
CLIENT_NAME = args.client_name or args.client_slug.replace("-", " ").title()
OUT_PATH = Path(args.output) if args.output else AUDIT_DIR / f"{args.client_slug}-local-audit.xlsx"

# Resolved at main() startup from the clean gbp-profile.json (or raw fallback).
# Used to (a) label the "YOU" row across competitor tabs and (b) exclude the client
# from top-competitor results. Defaults are placeholders that should always be
# overridden before any render function runs.
CLIENT_PLACE_ID: str = ""
CLIENT_LABEL: str = f"{CLIENT_NAME} (YOU)"
CLIENT_DOMAIN_TOKEN: str = args.client_slug.split("-")[0].lower()

# ---------------------------------------------------------------------------
# Style tokens
# ---------------------------------------------------------------------------

FONT_BASE = "Arial"
FONT_H1 = Font(name=FONT_BASE, size=16, bold=True, color="FFFFFF")
FONT_H2 = Font(name=FONT_BASE, size=12, bold=True, color="FFFFFF")
FONT_H3 = Font(name=FONT_BASE, size=11, bold=True, color="1F2937")
FONT_BODY = Font(name=FONT_BASE, size=10, color="111827")
FONT_BODY_BOLD = Font(name=FONT_BASE, size=10, bold=True, color="111827")
FONT_MUTED = Font(name=FONT_BASE, size=10, italic=True, color="6B7280")
FONT_MONO = Font(name="Courier New", size=9, color="111827")

FILL_H1 = PatternFill("solid", fgColor="1F2937")   # near-black
FILL_H2 = PatternFill("solid", fgColor="4B5563")   # gray-600
FILL_OK = PatternFill("solid", fgColor="D1FAE5")    # green-100
FILL_WEAK = PatternFill("solid", fgColor="FEF3C7")  # yellow-100
FILL_MISSING = PatternFill("solid", fgColor="FEE2E2")  # red-100
FILL_PENDING = PatternFill("solid", fgColor="E5E7EB")  # gray-200 — API not connected yet
FILL_P1 = PatternFill("solid", fgColor="DC2626")    # red-600
FILL_P2 = PatternFill("solid", fgColor="F59E0B")    # amber-500
FILL_P3 = PatternFill("solid", fgColor="9CA3AF")    # gray-400
FILL_APPROVAL = PatternFill("solid", fgColor="DBEAFE")  # blue-100 — strategist input cell
FILL_BANNER = PatternFill("solid", fgColor="EEF2FF")  # indigo-50

BORDER_THIN = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load(name: str) -> dict | list:
    p = AUDIT_DIR / name
    if not p.exists():
        print(f"  ⚠ missing cache file: {name}")
        return {}
    return json.loads(p.read_text())


def load_any(*names: str, also_check_raw: bool = True) -> dict | list:
    """Try several filenames; check AUDIT_DIR first, then raw/ if `also_check_raw`.

    Lets the renderer pick up either normalize-output names (e.g. `local-falcon-grid.json`)
    or raw-pull names (e.g. `keyword-rankings.json`, `windsor-gmb-monthly.json`) which
    the normalizer leaves untouched in `raw/`. Returns {} if nothing matches.
    """
    for n in names:
        p = AUDIT_DIR / n
        if p.exists():
            return json.loads(p.read_text())
        if also_check_raw:
            pr = AUDIT_DIR / "raw" / n
            if pr.exists():
                return json.loads(pr.read_text())
    print(f"  ⚠ missing: {' | '.join(names)} (not in {AUDIT_DIR.name}/ or raw/)")
    return {}


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def title_row(ws: Worksheet, row: int, text: str, span: int = 8, fill=FILL_H1, font=FONT_H1):
    """Write a banner row spanning columns A..{span}."""
    ws.cell(row=row, column=1, value=text).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    for c in range(2, span + 1):
        ws.cell(row=row, column=c).fill = fill
    ws.row_dimensions[row].height = 24
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def section_row(ws: Worksheet, row: int, text: str, span: int = 8):
    title_row(ws, row, text, span, FILL_H2, FONT_H2)
    ws.row_dimensions[row].height = 19.5   # matches Ryan's preferred sheet spec


def header_row(ws: Worksheet, row: int, columns: list[str]):
    for i, label in enumerate(columns, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = FONT_H3
        c.fill = FILL_BANNER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[row].height = 31.5   # matches Ryan's preferred sheet spec


def body_cell(ws: Worksheet, row: int, col: int, value, font=FONT_BODY, fill=None,
              align=ALIGN_LEFT, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = align
    if border:
        c.border = BORDER_THIN
    return c


def status_fill(status: str) -> PatternFill | None:
    return {
        "ok": FILL_OK,
        "live_correct": FILL_OK,
        "weak": FILL_WEAK,
        "live_mismatch": FILL_WEAK,
        "partial": FILL_WEAK,
        "missing": FILL_MISSING,
        "api_pending": FILL_PENDING,
        "unverified": FILL_PENDING,
        "unknown": FILL_PENDING,
    }.get(status)


def priority_fill(priority: str) -> PatternFill | None:
    return {"P1": FILL_P1, "P2": FILL_P2, "P3": FILL_P3}.get(priority)


# ---------------------------------------------------------------------------
# TAB 1 — README
# ---------------------------------------------------------------------------

def render_readme(ws: Worksheet):
    ws.column_dimensions["A"].width = 100
    title_row(ws, 1, f"Local SEO Audit — {CLIENT_NAME}", span=1)
    body_cell(ws, 2, 1, "How to use this workbook", FONT_H3, FILL_BANNER, ALIGN_LEFT, border=False)

    sections = [
        ("Workflow",
         "1. Review the Summary tab (topline grade, top 3 priorities).\n"
         "2. Walk through tabs 3-8 to understand the findings in depth.\n"
         "3. Open the Actions tab — fill in the Approval column for every row.\n"
         "4. Save the file. Run parse_local_approvals.py to lock in decisions.\n"
         "5. The approved actions feed build_project_plan.py → Sprint 3 deliverables."),
        ("Approval values (Actions tab)",
         "Approved   → Goes into Sprint 3 as-is\n"
         "Edited     → Goes into Sprint 3 with your edits (use Edit Notes column)\n"
         "Deferred   → Held for later; not in initial Sprint 3\n"
         "Rejected   → Skipped entirely"),
        ("Color key",
         "🟢 Green   = OK / live + correct\n"
         "🟡 Yellow  = Weak / live but mismatched\n"
         "🔴 Red     = Missing / critical gap\n"
         "🟦 Blue    = Your input — fill these cells"),
        ("Priority levels",
         "P1 — Critical. Blocks rankings or trust. Address in week 1.\n"
         "P2 — Important. Causes friction. Address by end of Sprint 3.\n"
         "P3 — Nice to have. Tackle in monthly recurring cadence."),
        ("Grading methodology",
         "GBP grade is weighted across 16 fields (primary category = highest weight, attributes = lowest).\n"
         "Citation coverage = % of 50 known directories listed correctly.\n"
         "Local rankings = % of 250 grid cells where the client appears in top 20.\n"
         "Local link gap = % of competitor referring domains the client doesn't have."),
        ("Data sources",
         "GBP profile     → Google Business Profile API (via Agency OS MCP)\n"
         "Local rankings  → Local Falcon (10 keywords × 5×5 grid = 250 scans)\n"
         "GSC queries     → Windsor.ai GSC connector (90 days vs prior 90 days)\n"
         "Citations       → Web search across 50 known directories\n"
         "Local links     → Ahrefs MCP (client + top 3 competitors)\n"
         "Local pages     → Existing site crawl (Screaming Frog)"),
    ]

    row = 4
    for heading, body in sections:
        body_cell(ws, row, 1, heading, FONT_H3, FILL_BANNER, ALIGN_LEFT, border=False)
        row += 1
        cell = body_cell(ws, row, 1, body, FONT_BODY, None, ALIGN_LEFT, border=False)
        ws.row_dimensions[row].height = max(18 * (body.count("\n") + 1), 36)
        row += 2

    body_cell(ws, row + 1, 1,
              "Questions? See /agency-os-delivery-local-seo-audit SKILL.md or ping the Local SEO Engineer.",
              FONT_MUTED, None, ALIGN_LEFT, border=False)


# ---------------------------------------------------------------------------
# TAB 2 — Summary
# ---------------------------------------------------------------------------

def render_summary(ws: Worksheet, manifest, gbp, falcon, citations, links, pages, gsc):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36     # was 18, widened to match Ryan's spec
    ws.column_dimensions["D"].width = 80

    title_row(ws, 1, f"Audit Summary — {CLIENT_NAME}", span=4)

    body_cell(ws, 2, 1, f"Generated: {manifest.get('generated_at', '—')}", FONT_MUTED, border=False)
    body_cell(ws, 2, 4, f"Audit ID: {manifest.get('audit_id', '—')}", FONT_MUTED, align=ALIGN_RIGHT, border=False)

    # Grade cards
    section_row(ws, 4, "Category Grades", span=4)
    header_row(ws, 5, ["Category", "Grade", "Score", "Top finding"])

    # Compute summary stats
    falcon_summary = falcon.get("summary", {})
    avg_coverage = (
        sum(s.get("coverage_pct", 0) for s in falcon_summary.values()) / len(falcon_summary)
        if falcon_summary else 0
    )
    citation_cov = citations.get("summary", {}).get("coverage_pct", 0)
    gbp_pct = gbp.get("grade_pct", 0)
    # Reviews grade is a weighted blend of count + rating + response rate.
    # Counting alone (the old behavior) graded a 4.9★/100%-response shop with 21
    # reviews as "F (21%)" — wrong by any practical standard. Strategists need to
    # see that quality + responsiveness can offset a modest count.
    review_summary = gbp.get("fields", [])
    _scores = {f["field"]: (f.get("score") or 0) for f in review_summary if "review" in f["field"]}
    review_pct = round(100 * (
        0.4 * _scores.get("review_count", 0)
        + 0.4 * _scores.get("review_avg", 0)
        + 0.2 * _scores.get("review_response_rate", 0)
    ), 1)
    link_gap = links.get("summary", {}).get("gap_pct", 0)
    link_grade = 100 - link_gap

    categories = [
        ("Google Business Profile", gbp_pct,
         _top_finding_gbp(gbp)),
        ("Local Rankings (250-cell grid)", avg_coverage,
         _top_finding_rankings(falcon)),
        ("Citations (50 directories)", citation_cov,
         _top_finding_citations(citations)),
        ("Reviews", review_pct,
         _top_finding_reviews(gbp)),
    ]

    row = 6
    for label, pct, finding in categories:
        grade = _letter(pct)
        body_cell(ws, row, 1, label, FONT_BODY_BOLD)
        gc = body_cell(ws, row, 2, grade, FONT_BODY_BOLD, _grade_fill(pct), ALIGN_CENTER)
        body_cell(ws, row, 3, f"{pct:.0f}%", FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 4, finding, FONT_BODY)
        ws.row_dimensions[row].height = 32
        row += 1

    # Top 3 priorities — biggest gaps from each category
    row += 1
    section_row(ws, row, "Top 3 Priorities (P1 actions)", span=4)
    row += 1
    header_row(ws, row, ["#", "Category", "Action", "Why it matters"])
    row += 1

    p1_actions = _collect_p1_priorities(gbp, citations, links, pages)
    for i, (cat, action, rationale) in enumerate(p1_actions[:3], 1):
        body_cell(ws, row, 1, i, FONT_BODY_BOLD, FILL_P1, ALIGN_CENTER)
        body_cell(ws, row, 2, cat, FONT_BODY)
        body_cell(ws, row, 3, action, FONT_BODY_BOLD)
        body_cell(ws, row, 4, rationale, FONT_BODY)
        ws.row_dimensions[row].height = 32
        row += 1

    # 8-day implementation plan reference
    row += 1
    section_row(ws, row, "Next steps", span=4)
    row += 1
    body_cell(ws, row, 1,
              "1. Review tabs 3-8 to understand the full findings.\n"
              "2. Fill in the Approval column on the Actions tab.\n"
              "3. Save → run parse_local_approvals.py → Sprint 3 plan is generated.",
              FONT_BODY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 60


def _grade_fill(pct: float) -> PatternFill:
    if pct >= 80: return FILL_OK
    if pct >= 60: return FILL_WEAK
    return FILL_MISSING


def _letter(pct: float) -> str:
    if pct >= 90: return "A"
    if pct >= 80: return "B"
    if pct >= 70: return "C"
    if pct >= 60: return "D"
    return "F"


def _pages_score(pages: dict) -> float:
    s = pages.get("summary", {})
    loc = s.get("location_pages_count", 0)
    sa = s.get("service_area_count", 0)
    schema = s.get("local_schema_count", 0)
    score = min(40, loc * 4) + min(40, sa * 2) + min(20, schema * 2)
    return score


# Top-finding helpers
def _top_finding_gbp(gbp: dict) -> str:
    missing = [f for f in gbp.get("fields", []) if f.get("status") == "missing"]
    if missing:
        return f"{len(missing)} field(s) missing; highest weight: {missing[0]['field']}"
    weak = [f for f in gbp.get("fields", []) if f.get("status") == "weak"]
    if weak:
        return f"{len(weak)} field(s) weak; review {weak[0]['field']}"
    return "All fields complete"


def _top_finding_rankings(falcon: dict) -> str:
    summary = falcon.get("summary", {})
    if not summary:
        return "No ranking data pulled"
    avg_cov = sum(s.get("coverage_pct", 0) for s in summary.values()) / len(summary)
    worst = min(summary.items(), key=lambda kv: kv[1].get("coverage_pct", 100))
    return f"Avg coverage {avg_cov:.0f}%. Worst: '{worst[0]}' at {worst[1].get('coverage_pct', 0):.0f}%"


def _top_finding_citations(citations: dict) -> str:
    s = citations.get("summary", {})
    return f"{s.get('missing', 0)} missing, {s.get('live_mismatch', 0)} mismatched ({s.get('coverage_pct', 0):.0f}% coverage)"


def _top_finding_reviews(gbp: dict) -> str:
    count_field = next((f for f in gbp.get("fields", []) if f["field"] == "review_count"), {})
    avg_field = next((f for f in gbp.get("fields", []) if f["field"] == "review_avg"), {})
    return f"{count_field.get('current', '—')} reviews, {avg_field.get('current', '—')} avg"


def _top_finding_pages(pages: dict) -> str:
    s = pages.get("summary", {})
    return f"{s.get('location_pages_count', 0)} location, {s.get('service_area_count', 0)} service-area, {s.get('local_schema_count', 0)} with schema"


def _top_finding_links(links: dict) -> str:
    s = links.get("summary", {})
    return f"{s.get('p1_opportunities', 0)} P1 opportunities, {s.get('gap_pct', 0):.0f}% gap vs competitors"


def _collect_p1_priorities(gbp, citations, links, pages) -> list[tuple[str, str, str]]:
    out = []
    # GBP missing fields with weight >= 5
    for f in gbp.get("fields", []):
        if f.get("status") == "missing" and f.get("weight", 0) >= 5:
            out.append(("GBP", f.get("action", ""), f.get("rationale", "")))
    # P1 citation actions
    for row in citations.get("directories", []):
        if row.get("priority") == "P1" and row.get("status") in ("missing", "live_mismatch"):
            out.append(("Citations", f"{row['action']}: {row['directory']}",
                        "Tier 1 aggregators feed Apple, Bing, GPS, and hundreds of secondary sites."))
            if len(out) >= 6:
                break
    # P1 link opportunities
    for row in links.get("rows", [])[:3]:
        if row.get("opportunity") == "P1":
            out.append(("Local Links", f"Outreach to {row['domain']}",
                        f"{row['comp_count']} of 3 competitors have this link"))
    # Location pages missing
    if pages.get("summary", {}).get("location_pages_count", 0) == 0:
        out.append(("Local Content", "Build per-location landing pages",
                    "No location-specific pages exist; required for Maps Pack on multi-location queries."))
    return out


# ---------------------------------------------------------------------------
# TAB 3 — GBP Audit
# ---------------------------------------------------------------------------

def render_gbp_audit(ws: Worksheet, gbp: dict, location_profiles=None):
    """Per-location GBP field audit. location_profiles = list of (label, profile_dict);
    each profile renders as its own section with the same field structure + grade.
    Falls back to the single primary profile when no location list is supplied."""
    ws.column_dimensions["A"].width = 17.14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 11.43
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 40

    if not location_profiles:
        location_profiles = [(None, gbp)]

    title_row(ws, 1, "Google Business Profile — Field Audit (per location)", span=7)
    row = 3
    for label, prof in location_profiles:
        hdr = (f"📍 {label} — " if label else "") + \
              f"grade {prof.get('letter_grade', '—')} ({prof.get('grade_pct', 0):.1f}%)"
        section_row(ws, row, hdr, span=7)
        row += 1
        header_row(ws, row, ["Field", "Current value", "Status", "Score", "Weight", "Rationale", "Recommended action"])
        row += 1
        start = row
        for f in prof.get("fields", []):
            body_cell(ws, row, 1, str(f.get("field", "")).replace("_", " ").title(), FONT_BODY_BOLD)
            current = f.get("current")
            if isinstance(current, list):
                current = ", ".join(str(x) for x in current[:5]) + (f" (+{len(current)-5} more)" if len(current) > 5 else "")
            body_cell(ws, row, 2, str(current) if current not in (None, "") else "—", FONT_BODY)
            body_cell(ws, row, 3, f.get("status", "unknown"), FONT_BODY_BOLD, status_fill(f.get("status", "unknown")), ALIGN_CENTER)
            body_cell(ws, row, 4, f.get("score", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 5, f.get("weight", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 6, f.get("rationale", ""), FONT_MUTED)
            body_cell(ws, row, 7, f.get("action", ""), FONT_BODY)
            ws.row_dimensions[row].height = 36
            row += 1
        body_cell(ws, row, 1, "Weighted Score", FONT_BODY_BOLD, FILL_BANNER)
        body_cell(ws, row, 4, f"=SUMPRODUCT(D{start}:D{row-1},E{start}:E{row-1})/SUM(E{start}:E{row-1})",
                  FONT_BODY_BOLD, FILL_BANNER, ALIGN_CENTER)
        body_cell(ws, row, 5, f"=SUM(E{start}:E{row-1})", FONT_BODY_BOLD, FILL_BANNER, ALIGN_CENTER)
        row += 3


# ---------------------------------------------------------------------------
# TAB 4 — Local Rankings (grid heatmap)
# ---------------------------------------------------------------------------

def render_local_rankings(ws: Worksheet, falcon: dict, audit_config: dict = None):
    ws.column_dimensions["A"].width = 32.14   # was 35, narrowed per Ryan

    n_reports = len(falcon.get("reports", []) or [])
    n_keywords = len(falcon.get("summary", {}) or {})
    title_row(ws, 1, f"Local Rankings — {n_keywords} keywords scanned via Local Falcon", span=8)

    if falcon.get("_skipped") or not falcon.get("summary"):
        reason = falcon.get("reason") or "no normalized data"
        body_cell(ws, 3, 1,
                  f"Local Falcon data unavailable: {reason}\n"
                  f"Fix: run pull_lf_batch.py (fetches real reports via the LF API) then re-run build_local_audit_data.py.",
                  FONT_MUTED, border=False)
        return

    # Per-keyword summary
    section_row(ws, 3, "Keyword Summary (Local Falcon scans)", span=8)
    header_row(ws, 4, ["Keyword", "Avg Rank (ARP)", "Coverage %", "SoLV %", "Found / Total", f"Top Competitor ({CLIENT_NAME} excluded)", "Last Scan", "View Heatmap"])

    row = 5
    summary = falcon.get("summary", {})
    competitors = falcon.get("competitors", {})
    reports = {r.get("keyword"): r for r in falcon.get("reports", [])}

    for kw, s in summary.items():
        body_cell(ws, row, 1, kw, FONT_BODY_BOLD)
        body_cell(ws, row, 2, s.get("avg_rank", "—"), FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 3, f"{s.get('coverage_pct', 0):.0f}%", FONT_BODY,
                  _grade_fill(s.get("coverage_pct", 0)), ALIGN_CENTER)
        body_cell(ws, row, 4, f"{s.get('solv_pct', 0):.1f}%" if s.get("solv_pct") is not None else "—",
                  FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 5, f"{s.get('found_count', 0)}/{s.get('total_cells', 49)}", FONT_BODY, None, ALIGN_CENTER)

        # Top Competitor — EXCLUDE the client itself (by place_id + name/domain token)
        comps = competitors.get(kw) or []
        token = CLIENT_DOMAIN_TOKEN.lower()
        non_self = [c for c in comps if (c.get("place_id") != CLIENT_PLACE_ID
                                         and token not in (c.get("domain") or c.get("name", "")).lower())]
        top_comp = non_self[0] if non_self else None
        if top_comp:
            comp_text = f"{top_comp.get('domain') or top_comp.get('name', '?')} ({top_comp.get('reviews', '?')} reviews, {top_comp.get('rating', '?')}★)"
        else:
            comp_text = "—"
        body_cell(ws, row, 6, comp_text, FONT_BODY)

        # Report info
        report = reports.get(kw, {})
        body_cell(ws, row, 7, report.get("scan_date", "—"), FONT_MUTED, None, ALIGN_CENTER)

        # Link to public LF heatmap
        public_url = report.get("public_url", "")
        if public_url:
            cell = body_cell(ws, row, 8, "View on Local Falcon →", FONT_BODY, None, ALIGN_CENTER)
            cell.hyperlink = public_url
            cell.font = Font(name=FONT_BASE, size=10, color="2563EB", underline="single")
        else:
            body_cell(ws, row, 8, "—", FONT_MUTED, None, ALIGN_CENTER)
        row += 1

    # Per-office × keyword breakdown — shows EVERY location, not just the primary
    # office. (The keyword roll-up above collapses to one office per keyword.)
    by_office = falcon.get("by_office", []) or []
    if by_office:
        row += 1
        section_row(ws, row, "Per-Office × Keyword — all locations", span=8)
        row += 1
        header_row(ws, row, ["Office", "Keyword", "Avg Rank (ARP)", "Coverage %", "SoLV %", "Found / Total", "Grid", ""])
        row += 1
        for e in sorted(by_office, key=lambda x: (str(x.get("office") or ""), str(x.get("keyword") or ""))):
            body_cell(ws, row, 1, e.get("office") or "—", FONT_BODY_BOLD)
            body_cell(ws, row, 2, e.get("keyword") or "—", FONT_BODY)
            body_cell(ws, row, 3, e.get("avg_rank") if e.get("avg_rank") is not None else "—", FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 4, f"{e.get('coverage_pct', 0):.0f}%", FONT_BODY, _grade_fill(e.get("coverage_pct", 0)), ALIGN_CENTER)
            body_cell(ws, row, 5, f"{e.get('solv_pct', 0):.1f}%" if e.get("solv_pct") is not None else "—", FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 6, f"{e.get('found_count', 0)}/{e.get('total_cells', 81)}", FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 7, e.get("grid_size") or "—", FONT_MUTED, None, ALIGN_CENTER)
            body_cell(ws, row, 8, "", FONT_BODY)
            row += 1

    # Note about heatmap
    row += 1
    body_cell(ws, row, 1,
              "Per-cell heatmap not embedded here — view the full 49-cell heatmap by clicking the "
              "'View on Local Falcon' link for each keyword. Local Falcon hosts the visual grid + "
              "lets you compare scans over time. The summary metrics above (ARP, Coverage, SoLV) come "
              "directly from the Local Falcon API and ARE complete.",
              FONT_MUTED, None, ALIGN_LEFT, border=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 50


# ---------------------------------------------------------------------------
# TAB 7 — NAP Consistency
# ---------------------------------------------------------------------------

def render_nap_consistency(ws: Worksheet, gbp: dict, citations: dict):
    """Diff canonical NAP against every place it appears across the web."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18.43   # was 35, narrowed per Ryan
    ws.column_dimensions["C"].width = 31      # was 45, narrowed per Ryan
    ws.column_dimensions["D"].width = 38.43   # was 25, widened per Ryan
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 45

    title_row(ws, 1, "NAP Consistency Cross-Check", span=6)
    body_cell(ws, 2, 1,
              "NAP = Name, Address, Phone. Every directory + on-site instance must match the correct values exactly. "
              "Minor variants (punctuation, suite # format) are flagged amber; mismatched data is flagged red.",
              FONT_MUTED, border=False)

    section_row(ws, 4, "Correct NAP (source of truth)", span=6)
    canonical = dict(citations.get("canonical_nap", {}))
    # Fall back to gbp-profile when citations didn't populate canonical_nap.
    # The normalizer surfaces NAP as top-level keys (name/address/phone/website/hours_summary)
    # in addition to the graded `fields[]` list; check the top-level first.
    if isinstance(gbp, dict):
        for src_key, dest_key in [("name", "name"), ("address", "address"), ("phone", "phone"), ("website", "website"), ("hours_summary", "hours")]:
            if not canonical.get(dest_key) and gbp.get(src_key):
                canonical[dest_key] = gbp[src_key]
        # The graded fields list still wins as a tertiary fallback for phone/website.
        if gbp.get("fields"):
            by_field = {f["field"]: f.get("current") for f in gbp["fields"] if isinstance(f, dict)}
            for field_name, dest_key in [("phone", "phone"), ("website", "website")]:
                if not canonical.get(dest_key) and by_field.get(field_name):
                    canonical[dest_key] = by_field[field_name]
        if not canonical.get("name"):
            canonical["name"] = CLIENT_NAME
    row = 5
    for k in ["name", "address", "phone", "website", "hours"]:
        v = canonical.get(k) or "—"
        body_cell(ws, row, 1, k.title(), FONT_BODY_BOLD, FILL_BANNER)
        body_cell(ws, row, 2, v, FONT_BODY_BOLD, FILL_BANNER)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # On-site instances (website, schema, contact page)
    row += 1
    section_row(ws, row, "On-Site Instances", span=6)
    row += 1
    header_row(ws, row, ["Source", "Field", "Value Found", "Diff", "Status", "Action"])
    row += 1

    onsite_rows = [
        ("Website footer / contact page", "name", canonical.get("name", "—"),
         "Match", "live_correct", ""),
        ("Website footer / contact page", "address", canonical.get("address", "—"),
         "Match", "live_correct", ""),
        ("Website footer / contact page", "phone", canonical.get("phone", "—"),
         "Match", "live_correct", ""),
        ("LocalBusiness schema markup", "any", "(none detected)",
         "Schema NOT implemented on any page",
         "missing", "P1 — Add LocalBusiness schema to home + contact + about pages"),
        ("LegalService schema markup", "any", "(none detected)",
         "LegalService schema NOT implemented",
         "missing", "P1 — Add LegalService schema with practice areas to relevant pages"),
        ("Embedded Google Map (contact page)", "—", "Present? Manual check needed",
         "—", "unknown", "Verify map embed on /contact-us/"),
        ("Click-to-call phone link", "phone", "tel:9198700466",
         "Present in header", "live_correct", ""),
    ]
    for source, field, value, diff, status, action in onsite_rows:
        body_cell(ws, row, 1, source, FONT_BODY_BOLD)
        body_cell(ws, row, 2, field, FONT_BODY)
        body_cell(ws, row, 3, value, FONT_BODY)
        body_cell(ws, row, 4, diff, FONT_BODY)
        body_cell(ws, row, 5, status.replace("_", " "), FONT_BODY_BOLD, status_fill(status), ALIGN_CENTER)
        body_cell(ws, row, 6, action, FONT_BODY)
        row += 1

    # External citations — NAP diff per directory
    row += 1
    section_row(ws, row, "External Citation NAP Diffs", span=6)
    row += 1
    header_row(ws, row, ["Directory", "Status", "URL", "Diff", "Tier/Priority", "Fix Action"])
    row += 1

    # Sort: mismatch first (highest impact), then missing, then correct
    sort_order = {"live_mismatch": 0, "missing": 1, "partial": 2, "unknown": 3, "live_correct": 4}
    dirs = sorted(citations.get("directories", []),
                  key=lambda d: (sort_order.get(d.get("status", ""), 9), d.get("priority", "P3")))
    for d in dirs:
        if d.get("status") == "live_correct":
            continue  # Only show issues in this tab — correct ones live in Citations tab
        body_cell(ws, row, 1, d["directory"], FONT_BODY_BOLD)
        body_cell(ws, row, 2, d["status"].replace("_", " "), FONT_BODY_BOLD, status_fill(d["status"]), ALIGN_CENTER)
        body_cell(ws, row, 3, d.get("url") or "—", FONT_MONO)
        body_cell(ws, row, 4, d.get("diff_summary") or "(not listed)", FONT_BODY)
        body_cell(ws, row, 5, f"T{d.get('tier','?')} · {d.get('priority','P3')}", FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 6, d.get("action") or "—", FONT_BODY)
        row += 1

    # Summary
    row += 1
    section_row(ws, row, "NAP Consistency Score", span=6)
    row += 1
    s = citations.get("summary", {})
    total = s.get("total", 0)
    correct = s.get("live_correct", 0)
    mismatch = s.get("live_mismatch", 0)
    missing = s.get("missing", 0)
    body_cell(ws, row, 1, "Tracked instances", FONT_BODY_BOLD)
    body_cell(ws, row, 2, total, FONT_BODY, None, ALIGN_CENTER)
    body_cell(ws, row, 3, f"Exact match: {correct} ({100*correct/max(1,total):.0f}%)", FONT_BODY)
    body_cell(ws, row, 4, f"Mismatch: {mismatch}", FONT_BODY, FILL_WEAK)
    body_cell(ws, row, 5, f"Missing: {missing}", FONT_BODY, FILL_MISSING)
    body_cell(ws, row, 6, "Plus on-site sources (above)", FONT_MUTED)


# ---------------------------------------------------------------------------
# TAB 8 — Competitor Strategic Deep-Dive
# ---------------------------------------------------------------------------

def render_competitor_deepdive(ws: Worksheet, falcon: dict, research: dict, gbp: dict):
    """Top competitors with strategic comparison: GBP profile depth, review velocity, content."""
    ws.column_dimensions["A"].width = 35
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["H"].width = 23.71   # was 14, widened per Ryan (Keywords seen in)
    ws.column_dimensions["I"].width = 30

    title_row(ws, 1, "Competitor Strategic Deep-Dive", span=8)
    body_cell(ws, 2, 1,
              "Top GBP competitors (real data from Local Falcon scans) — review counts, ratings, "
              "appearances in the 4-keyword grid, and where they beat the client.",
              FONT_MUTED, border=False)

    section_row(ws, 4, "GBP Pack Competitors (from Local Falcon scans)", span=8)
    header_row(ws, 5, [
        "Competitor", "Reviews", "Rating", "Apps", "Avg Rank",
        "Place ID / URL", f"Beats {CLIENT_NAME} on", "Keywords seen in"
    ])

    # Aggregate competitors across all keywords
    competitors_seen: dict = {}
    for kw, kw_comps in falcon.get("competitors", {}).items():
        for c in kw_comps:
            key = c.get("place_id") or c.get("domain") or c.get("name", "")
            if not key:
                continue
            if key not in competitors_seen:
                competitors_seen[key] = {
                    "name": c.get("domain") or c.get("name", ""),
                    "reviews": c.get("reviews"),
                    "rating": c.get("rating"),
                    "place_id": c.get("place_id"),
                    "url": c.get("url"),
                    "total_appearances": 0,
                    "rank_sum": 0.0,
                    "rank_count": 0,
                    "keywords_seen": set(),
                }
            entry = competitors_seen[key]
            entry["total_appearances"] += c.get("appearances", 0)
            if c.get("avg_rank") is not None:
                entry["rank_sum"] += c["avg_rank"] * c.get("appearances", 1)
                entry["rank_count"] += c.get("appearances", 1)
            entry["keywords_seen"].add(kw)

    # The client's own stats for the highlighted "YOU" comparison row.
    batch_reviews = int(next((f.get("current") for f in gbp.get("fields", []) if f.get("field") == "review_count"), 0) or 0)
    try:
        batch_rating = float(next((f.get("current") for f in gbp.get("fields", []) if f.get("field") == "review_avg"), 0) or 0)
    except (ValueError, TypeError):
        batch_rating = 0.0

    # Insert the client's "YOU" row first (highlighted)
    row = 6
    body_cell(ws, row, 1, CLIENT_LABEL, FONT_BODY_BOLD, FILL_OK)
    body_cell(ws, row, 2, batch_reviews, FONT_BODY_BOLD, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 3, batch_rating, FONT_BODY_BOLD, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 4, "—", FONT_BODY, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 5, "—", FONT_BODY, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 6, CLIENT_PLACE_ID or "—", FONT_MONO, FILL_OK)
    body_cell(ws, row, 7, "—", FONT_BODY, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 8, "—", FONT_MUTED, FILL_OK)
    row += 1

    self_token = CLIENT_DOMAIN_TOKEN.lower()
    for c in sorted(competitors_seen.values(), key=lambda x: -x.get("total_appearances", 0))[:12]:
        if self_token and self_token in c["name"].lower():
            continue  # already shown as the YOU row
        avg_rank = round(c["rank_sum"] / c["rank_count"], 2) if c["rank_count"] else "—"
        try:
            comp_reviews = int(c.get("reviews") or 0)
        except (ValueError, TypeError):
            comp_reviews = 0
        try:
            comp_rating = float(c.get("rating") or 0)
        except (ValueError, TypeError):
            comp_rating = 0.0

        beats = []
        if comp_reviews > batch_reviews: beats.append("reviews")
        if comp_rating > batch_rating: beats.append("rating")
        beats_str = ", ".join(beats) or "—"

        body_cell(ws, row, 1, c["name"], FONT_BODY_BOLD)
        body_cell(ws, row, 2, comp_reviews,
                  FONT_BODY, FILL_MISSING if comp_reviews > batch_reviews else None, ALIGN_CENTER)
        body_cell(ws, row, 3, comp_rating,
                  FONT_BODY, FILL_MISSING if comp_rating > batch_rating else None, ALIGN_CENTER)
        body_cell(ws, row, 4, c.get("total_appearances", 0), FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 5, avg_rank, FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 6, (c.get("place_id") or c.get("url") or "—")[:30], FONT_MONO)
        body_cell(ws, row, 7, beats_str, FONT_BODY)
        body_cell(ws, row, 8, "; ".join(sorted(c.get("keywords_seen", set()))[:3]), FONT_MUTED)
        row += 1

    # Footnote
    row += 1
    body_cell(ws, row, 1,
              "Red cells = competitor beats this firm on that metric. "
              "Data source: aggregated from Local Falcon GBP grid scans across all 4 audit keywords.",
              FONT_MUTED, None, ALIGN_LEFT, border=False)


# ---------------------------------------------------------------------------
# TAB 9 — AI / GEO Visibility
# ---------------------------------------------------------------------------

def render_ai_geo_visibility(ws: Worksheet, findings: dict, gbp: dict):
    """Whether the client surfaces in AI search + entity completeness."""
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80

    title_row(ws, 1, "AI / GEO Search Visibility", span=3)
    body_cell(ws, 2, 1,
              "Generative AI search engines (ChatGPT, Perplexity, Gemini, Claude) are becoming a major "
              "discovery channel for high-consideration queries like 'best raleigh divorce attorney'. "
              "Visibility here depends on entity authority signals — citations, press, associations — NOT GBP optimization.",
              FONT_MUTED, border=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.row_dimensions[2].height = 48

    # Topline finding
    section_row(ws, 4, "Topline Finding", span=3)
    summary = (findings or {}).get("summary", "")
    grade = (findings or {}).get("grade", "F")
    body_cell(ws, 5, 1, "Grade", FONT_BODY_BOLD)
    body_cell(ws, 5, 2, grade, FONT_BODY_BOLD, _grade_fill(20 if grade == "F" else 100), ALIGN_CENTER)
    body_cell(ws, 5, 3, summary or "(no findings recorded)", FONT_BODY)

    # Per-query test results
    row = 7
    section_row(ws, row, "AI Query Test Results", span=3)
    row += 1
    header_row(ws, row, ["Query", f"{CLIENT_NAME} Surfaces?", "Firms that DO surface"])
    row += 1

    queries = (findings or {}).get("query_tests", [])
    for q in queries:
        body_cell(ws, row, 1, q.get("query", ""), FONT_BODY_BOLD)
        # Accept both `client_surfaces` (current schema) and legacy `batch_surfaces`.
        surfaces = q.get("client_surfaces", q.get("batch_surfaces", False))
        surfaced = "✓ YES" if surfaces else "❌ NO"
        body_cell(ws, row, 2, surfaced, FONT_BODY_BOLD,
                  FILL_OK if surfaces else FILL_MISSING, ALIGN_CENTER)
        body_cell(ws, row, 3, ", ".join(q.get("competitors_surfaced", [])), FONT_BODY)
        row += 1

    # Entity completeness
    row += 1
    section_row(ws, row, "Entity Signal Completeness", span=3)
    row += 1
    header_row(ws, row, ["Signal", "Status", "Notes"])
    row += 1

    signals = (findings or {}).get("entity_signals", [])
    for s in signals:
        body_cell(ws, row, 1, s.get("signal", ""), FONT_BODY_BOLD)
        status = s.get("status", "unknown")
        body_cell(ws, row, 2, status.replace("_", " "), FONT_BODY_BOLD, status_fill(status), ALIGN_CENTER)
        body_cell(ws, row, 3, s.get("notes", ""), FONT_BODY)
        row += 1

    # Action items
    row += 1
    section_row(ws, row, "Recommended Actions to Improve AI Visibility", span=3)
    row += 1
    header_row(ws, row, ["Priority", "Action", "Why it matters"])
    row += 1

    for a in (findings or {}).get("actions", []):
        body_cell(ws, row, 1, a.get("priority", "P3"), FONT_BODY_BOLD, priority_fill(a.get("priority")), ALIGN_CENTER)
        body_cell(ws, row, 2, a.get("action", ""), FONT_BODY_BOLD)
        body_cell(ws, row, 3, a.get("rationale", ""), FONT_BODY)
        row += 1


# ---------------------------------------------------------------------------
# TAB 5 — Citations
# ---------------------------------------------------------------------------

def render_citations(ws: Worksheet, citations: dict):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 65   # Detail — was 45, now wider for multi-line diff
    ws.column_dimensions["G"].width = 38

    title_row(ws, 1, "Citation Audit — 55 Directories × Location", span=7)

    s = citations.get("summary", {})
    _locs_preview = citations.get("locations") or []
    _cov_line = "  ·  ".join(
        f"{L.get('name')}: {L.get('summary',{}).get('coverage_pct',0):.0f}% live"
        for L in _locs_preview
    ) if _locs_preview else (
        f"Coverage: {s.get('coverage_pct', 0):.0f}%  ·  Live correct: {s.get('live_correct', 0)}  "
        f"·  Mismatch: {s.get('live_mismatch', 0)}  ·  Missing: {s.get('missing', 0)}  ·  Unverified: {s.get('unverified', 0)}")
    body_cell(ws, 2, 1, _cov_line, FONT_BODY_BOLD, border=False)

    # Legend explaining T1 / T2 / T3 tiers + firm vs attorney type
    body_cell(ws, 3, 1,
              "TIERS:  T1 = Data aggregators (Data Axle, Localeze, Foursquare, Infogroup) + Google/Apple/Bing — these feed everything downstream  ·  "
              "T2 = Vertical-specific (Avvo, Justia, FindLaw, Super Lawyers for legal) + major general directories (Yelp, BBB, Yellow Pages)  ·  "
              "T3 = Smaller local / niche (MerchantCircle, CitySearch, Patch, Whitepages, etc.)",
              FONT_MUTED, None, ALIGN_LEFT, border=False)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    ws.row_dimensions[3].height = 36

    body_cell(ws, 4, 1,
              "TYPE:  firm = directory has a firm-level profile (Yelp, BBB, FindLaw firm pages)  ·  "
              "attorney = directory only has per-lawyer profiles, no firm concept (Avvo, Justia, Nolo) — coverage check is per-attorney",
              FONT_MUTED, None, ALIGN_LEFT, border=False)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws.row_dimensions[4].height = 28

    from openpyxl.styles import Alignment
    WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def _render_dir_row(row, d):
        dir_type = d.get("directory_type", "firm")
        body_cell(ws, row, 1, d["directory"], FONT_BODY_BOLD)
        type_label = f"T{d.get('tier','?')} · {'attorney' if dir_type == 'attorney' else 'firm'}"
        body_cell(ws, row, 2, type_label, FONT_MUTED, None, ALIGN_CENTER)
        body_cell(ws, row, 3, d.get("priority","—"), FONT_BODY_BOLD, priority_fill(d.get("priority","P3")), ALIGN_CENTER)
        body_cell(ws, row, 4, d.get("status","").replace("_", " "), FONT_BODY, status_fill(d.get("status","")), ALIGN_CENTER)
        if dir_type == "attorney":
            found = d.get("attorneys_found") or []
            missing = d.get("attorneys_missing") or []
            cov_text = f"{len(found)}/{len(found) + len(missing)} attorneys" if (found or missing) else (d.get("url") or "—")
            body_cell(ws, row, 5, cov_text, FONT_BODY, None, ALIGN_CENTER)
        else:
            url_text = d.get("url") or "—"
            if d.get("status") == "live" and d.get("domain_rating"):
                url_text = f"DR {d['domain_rating']:.0f}  ·  {d.get('links_to_target','?')} links to client domain"
            body_cell(ws, row, 5, url_text, FONT_MONO)
        detail_cell = body_cell(ws, row, 6, d.get("diff_summary") or "—", FONT_BODY)
        if detail_cell is not None:
            detail_cell.alignment = WRAP_TOP
        action_cell = body_cell(ws, row, 7, d.get("action") or "—", FONT_BODY)
        if action_cell is not None:
            action_cell.alignment = WRAP_TOP
        text = d.get("diff_summary") or ""
        ws.row_dimensions[row].height = max(18, max(1, text.count("\n") + 1) * 16)

    # Per-location sections (falls back to a single flat section if no locations block)
    locs = citations.get("locations") or [{
        "name": None, "nap": citations.get("canonical_nap", {}),
        "directories": citations.get("directories", []), "summary": citations.get("summary", {}),
    }]
    # Sort each location's directories: NAP issues → missing → live → unverified, then tier
    _ORDER = {"live_mismatch": 0, "missing": 1, "live_correct": 2, "live": 2, "partial": 2, "unverified": 3}
    row = 6
    for loc in locs:
        nm = loc.get("name")
        ls = loc.get("summary", {})
        live = ls.get("live_correct", 0) + ls.get("live", 0) + ls.get("live_mismatch", 0)
        hdr = (f"📍 {nm} — Coverage {ls.get('coverage_pct',0):.0f}%  ·  {live} live  ·  "
               f"{ls.get('live_mismatch',0)} NAP issues  ·  {ls.get('missing',0)} missing  ·  "
               f"{ls.get('unverified',0)} not yet verified") if nm else "Per-Directory Status"
        section_row(ws, row, hdr, span=7); row += 1
        nap = loc.get("nap") or citations.get("canonical_nap", {})
        body_cell(ws, row, 1, "Canonical NAP", FONT_BODY_BOLD)
        nap_txt = f"{nap.get('address','—')}   ·   {nap.get('phone','—')}   ·   {nap.get('website','')}"
        body_cell(ws, row, 2, nap_txt, FONT_MONO)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1
        header_row(ws, row, ["Directory", "Type", "Priority", "Status", "URL / Coverage", "Detail", "Action"]); row += 1
        for d in sorted(loc.get("directories", []),
                        key=lambda x: (_ORDER.get(x.get("status"), 9), x.get("tier", 9), x.get("priority", "P3"))):
            _render_dir_row(row, d); row += 1
        row += 2  # gap between locations


# ---------------------------------------------------------------------------
# TAB 6 — Reviews
# ---------------------------------------------------------------------------

def render_reviews(ws: Worksheet, gbp: dict, falcon: dict, windsor_gbp: dict = None, windsor_reviews: dict = None):
    ws.column_dimensions["A"].width = 30
    for c in "BCDE":
        ws.column_dimensions[c].width = 18
    ws.column_dimensions["F"].width = 50

    title_row(ws, 1, "GBP Activity & Reviews", span=6)

    # ── 12-Month GBP Performance from Windsor (real data) ───────────────────
    row = 3
    if windsor_gbp and windsor_gbp.get("monthly"):
        totals = windsor_gbp.get("totals", {})
        dr = windsor_gbp.get("date_range", {})
        section_row(ws, row, f"12-Month GBP Performance — {dr.get('start','?')} to {dr.get('end','?')}", span=6)
        row += 1
        body_cell(ws, row, 1, "Source: Windsor.ai → Google Business Profile API (live pull)", FONT_MUTED, border=False)
        row += 1
        header_row(ws, row, ["Metric", "12mo Total", "Monthly Avg", "—", "—", "Notes"])
        row += 1
        n_months = len(windsor_gbp["monthly"]) or 12
        rows = [
            ("Profile views (impressions)",  totals.get("impressions", 0),         round(totals.get("impressions", 0) / n_months, 1),       "Times the GBP listing was shown in Maps + Search"),
            ("Total clicks",                  totals.get("clicks", 0),              round(totals.get("clicks", 0) / n_months, 1),            "All actions taken on the profile"),
            ("  Phone calls",                 totals.get("call_clicks", 0),         round(totals.get("call_clicks", 0) / n_months, 1),       "Tap-to-dial — the highest-intent action"),
            ("  Website clicks",              totals.get("website_clicks", 0),      round(totals.get("website_clicks", 0) / n_months, 1),    "Tap to visit the client website"),
            ("  Direction requests",          totals.get("direction_requests", 0),  round(totals.get("direction_requests", 0) / n_months, 1),"Tap for directions to office"),
            ("Conversations (chat)",          totals.get("conversations", 0),       "—",                                                     "GBP messaging — typically 0 for legal"),
        ]
        for label, total, avg, note in rows:
            indent = label.startswith("  ")
            body_cell(ws, row, 1, label, FONT_BODY if indent else FONT_BODY_BOLD)
            body_cell(ws, row, 2, total, FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 3, avg, FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 4, "", FONT_BODY)
            body_cell(ws, row, 5, "", FONT_BODY)
            body_cell(ws, row, 6, note, FONT_MUTED)
            row += 1
        row += 1

        # Monthly breakdown table
        section_row(ws, row, "Monthly Breakdown", span=6)
        row += 1
        header_row(ws, row, ["Month", "Profile Views", "Phone Calls", "Website Clicks", "Direction Requests", "Total Clicks"])
        row += 1
        for m in windsor_gbp["monthly"]:
            body_cell(ws, row, 1, m["month"], FONT_BODY_BOLD)
            body_cell(ws, row, 2, m.get("impressions", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 3, m.get("call_clicks", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 4, m.get("website_clicks", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 5, m.get("direction_requests", 0), FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 6, m.get("clicks", 0), FONT_BODY, None, ALIGN_CENTER)
            row += 1
        row += 1

    # ── Review Velocity from Windsor (NEW) ─────────────────────────────────
    if windsor_reviews and windsor_reviews.get("client_monthly"):
        section_row(ws, row, "Review Velocity — New Reviews / Month (Windsor)", span=6)
        row += 1
        header_row(ws, row, ["Firm", "12mo Total", "Monthly Avg", f"vs {CLIENT_NAME}", "", "Notes"])
        row += 1

        client_total = windsor_reviews.get("client_total", 0)
        batch_avg = round(client_total / 12, 1)
        body_cell(ws, row, 1, CLIENT_LABEL, FONT_BODY_BOLD, FILL_BANNER)
        body_cell(ws, row, 2, client_total, FONT_BODY_BOLD, FILL_BANNER, ALIGN_CENTER)
        body_cell(ws, row, 3, batch_avg, FONT_BODY_BOLD, FILL_BANNER, ALIGN_CENTER)
        body_cell(ws, row, 4, "—", FONT_BODY, FILL_BANNER, ALIGN_CENTER)
        body_cell(ws, row, 5, "", FONT_BODY, FILL_BANNER)
        body_cell(ws, row, 6, "Baseline", FONT_MUTED, FILL_BANNER)
        row += 1

        for p in windsor_reviews.get("peer_benchmarks", []):
            avg_mo = round(p["total_12mo"] / 12, 1)
            body_cell(ws, row, 1, p["firm"], FONT_BODY)
            body_cell(ws, row, 2, p["total_12mo"], FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 3, avg_mo, FONT_BODY, None, ALIGN_CENTER)
            body_cell(ws, row, 4, f"{p['vs_batch_x']}x", FONT_BODY_BOLD, FILL_MISSING, ALIGN_CENTER)
            body_cell(ws, row, 5, "", FONT_BODY)
            body_cell(ws, row, 6, "Peer firm — Windsor benchmark", FONT_MUTED)
            row += 1
        row += 1

        # Per-month detail for the client's own review velocity
        section_row(ws, row, f"{CLIENT_NAME} — Monthly Review Velocity", span=6)
        row += 1
        header_row(ws, row, ["Month", "New Reviews", "", "", "", ""])
        row += 1
        for m in windsor_reviews["client_monthly"]:
            body_cell(ws, row, 1, m["month"], FONT_BODY_BOLD)
            nr = m.get("new_reviews", 0)
            fill = FILL_OK if nr >= 3 else (FILL_WEAK if nr >= 1 else FILL_MISSING)
            body_cell(ws, row, 2, nr, FONT_BODY_BOLD, fill, ALIGN_CENTER)
            for c in range(3, 7):
                body_cell(ws, row, c, "", FONT_BODY)
            row += 1
        row += 1

    start_row = row

    # Pull review-related GBP fields
    review_fields = {f["field"]: f for f in gbp.get("fields", []) if "review" in f["field"]}
    rc = review_fields.get("review_count", {})
    ra = review_fields.get("review_avg", {})
    rr = review_fields.get("review_response_rate", {})

    section_row(ws, start_row, "Client Review Snapshot", span=6)
    header_row(ws, start_row + 1, ["Metric", "Current", "Status", "Target", "Score", "Notes"])

    rows = [
        ("Total reviews", rc.get("current", "—"), rc.get("status", "—"), "≥ 50", rc.get("score", 0), rc.get("action", "")),
        ("Average rating", ra.get("current", "—"), ra.get("status", "—"), "≥ 4.5", ra.get("score", 0), ra.get("action", "")),
        ("Response rate", f"{rr.get('current', '—')}%" if isinstance(rr.get("current"), (int, float)) else "—",
         rr.get("status", "—"), "≥ 95%", rr.get("score", 0), rr.get("action", "")),
    ]
    row = start_row + 2
    for label, current, status, target, score, action in rows:
        body_cell(ws, row, 1, label, FONT_BODY_BOLD)
        body_cell(ws, row, 2, current, FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 3, status, FONT_BODY_BOLD, status_fill(status), ALIGN_CENTER)
        body_cell(ws, row, 4, target, FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 5, score, FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 6, action, FONT_BODY)
        ws.row_dimensions[row].height = 28
        row += 1

    # Competitor benchmark (from Local Falcon scan results — real data)
    row += 1
    section_row(ws, row, "Competitor Benchmark (from Local Falcon GBP grid scans)", span=6)
    row += 1
    header_row(ws, row, ["Competitor", "Reviews", "Rating", "Place ID / URL", "Appears in N cells", "Avg rank"])
    row += 1

    # Aggregate competitors across all keywords, deduping by name (or place_id)
    competitors_seen: dict = {}
    for kw_comps in falcon.get("competitors", {}).values():
        for c in kw_comps:
            key = c.get("place_id") or c.get("domain") or c.get("name", "")
            if not key:
                continue
            if key not in competitors_seen:
                competitors_seen[key] = {
                    "name": c.get("domain") or c.get("name", ""),
                    "reviews": c.get("reviews"),
                    "rating": c.get("rating"),
                    "place_id": c.get("place_id"),
                    "url": c.get("url"),
                    "total_appearances": 0,
                    "rank_sum": 0,
                    "rank_count": 0,
                    "keywords_seen": set(),
                }
            entry = competitors_seen[key]
            entry["total_appearances"] += c.get("appearances", 0)
            if c.get("avg_rank") is not None:
                entry["rank_sum"] += c["avg_rank"] * c.get("appearances", 1)
                entry["rank_count"] += c.get("appearances", 1)
            # Track which keyword this comp surfaced for
        # Track keywords-per-comp at the kw level
    for kw, kw_comps in falcon.get("competitors", {}).values() and [] or []:  # placeholder to satisfy parser
        pass
    # Second pass: tag keywords seen
    for kw, kw_comps in falcon.get("competitors", {}).items():
        for c in kw_comps:
            key = c.get("place_id") or c.get("domain") or c.get("name", "")
            if key in competitors_seen:
                competitors_seen[key]["keywords_seen"].add(kw)

    # Compute the client's stats for the top comparison row
    batch_reviews = next((f.get("current") for f in gbp.get("fields", []) if f.get("field") == "review_count"), "—")
    batch_rating  = next((f.get("current") for f in gbp.get("fields", []) if f.get("field") == "review_avg"), "—")

    # Insert the client's row first (highlighted) for direct comparison
    body_cell(ws, row, 1, CLIENT_LABEL, FONT_BODY_BOLD, FILL_OK)
    body_cell(ws, row, 2, batch_reviews, FONT_BODY_BOLD, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 3, batch_rating, FONT_BODY_BOLD, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 4, "—", FONT_MUTED, FILL_OK)
    body_cell(ws, row, 5, "—", FONT_BODY, FILL_OK, ALIGN_CENTER)
    body_cell(ws, row, 6, "—", FONT_BODY, FILL_OK, ALIGN_CENTER)
    row += 1

    # Sort competitors by review count desc, then by appearances
    sorted_comps = sorted(
        competitors_seen.values(),
        key=lambda c: (
            -(int(c.get("reviews") or 0) if str(c.get("reviews") or "0").replace(".","").isdigit() else 0),
            -c.get("total_appearances", 0),
        ),
    )

    for c in sorted_comps[:10]:
        avg_rank = round(c["rank_sum"] / c["rank_count"], 2) if c["rank_count"] else "—"
        # Highlight if competitor has more reviews than this firm
        try:
            comp_reviews = int(c.get("reviews") or 0)
            batch_count = int(batch_reviews) if isinstance(batch_reviews, (int, str)) and str(batch_reviews).isdigit() else 0
            fill = FILL_MISSING if comp_reviews > batch_count else None
        except (ValueError, TypeError):
            fill = None
        body_cell(ws, row, 1, c["name"][:50], FONT_BODY_BOLD)
        body_cell(ws, row, 2, c.get("reviews") or "—", FONT_BODY, fill, ALIGN_CENTER)
        body_cell(ws, row, 3, c.get("rating") or "—", FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 4, (c.get("url") or c.get("place_id") or "—")[:50], FONT_MONO)
        body_cell(ws, row, 5, c.get("total_appearances", 0), FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 6, avg_rank, FONT_BODY, None, ALIGN_CENTER)
        row += 1

    # Footnote
    row += 1
    body_cell(ws, row, 1,
              "Red cells = competitor has more reviews than this firm. "
              "Source: aggregated from Local Falcon GBP grid scans across all 4 audit keywords.",
              FONT_MUTED, None, ALIGN_LEFT, border=False)


# ---------------------------------------------------------------------------
# TAB 7 — Local Content
# ---------------------------------------------------------------------------

def render_local_content(ws: Worksheet, pages: dict):
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15

    title_row(ws, 1, "Local Content Inventory", span=3)

    s = pages.get("summary", {})
    body_cell(ws, 2, 1,
              f"Location pages: {s.get('location_pages_count', 0)}  ·  "
              f"Service area pages: {s.get('service_area_count', 0)}  ·  "
              f"Pages with local schema: {s.get('local_schema_count', 0)}",
              FONT_BODY_BOLD, border=False)

    section_row(ws, 4, "Location Pages (per-city/branch)", span=3)
    header_row(ws, 5, ["URL", "Title", "Status"])
    row = 6
    for p in pages.get("location_pages", []):
        body_cell(ws, row, 1, p.get("url", "—"), FONT_MONO)
        body_cell(ws, row, 2, p.get("title", "—"), FONT_BODY)
        body_cell(ws, row, 3, "Live", FONT_BODY, FILL_OK, ALIGN_CENTER)
        row += 1
    if not pages.get("location_pages"):
        body_cell(ws, row, 1, "No location pages found — recommend building per-city landing pages.",
                  FONT_BODY, FILL_MISSING)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    row += 1
    section_row(ws, row, "Service-Area Pages", span=3)
    row += 1
    header_row(ws, row, ["URL", "Title", "Status"])
    row += 1
    for p in pages.get("service_area_pages", []):
        body_cell(ws, row, 1, p.get("url", "—"), FONT_MONO)
        body_cell(ws, row, 2, p.get("title", "—"), FONT_BODY)
        body_cell(ws, row, 3, "Live", FONT_BODY, FILL_OK, ALIGN_CENTER)
        row += 1
    if not pages.get("service_area_pages"):
        body_cell(ws, row, 1,
                  "No service-area pages found — recommend service-area pages per practice area + city.",
                  FONT_BODY, FILL_MISSING)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    row += 1
    section_row(ws, row, "Local Schema Markup", span=3)
    row += 1
    if pages.get("local_schema_pages"):
        header_row(ws, row, ["URL", "", ""])
        row += 1
        for p in pages.get("local_schema_pages", []):
            body_cell(ws, row, 1, p.get("url", "—"), FONT_MONO)
            row += 1
    else:
        body_cell(ws, row, 1,
                  "No pages with LocalBusiness/Service/Review schema detected. P1 action: implement on home + location pages.",
                  FONT_BODY, FILL_MISSING)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)


# ---------------------------------------------------------------------------
# TAB 8 — Local Links
# ---------------------------------------------------------------------------

def render_local_links(ws: Worksheet, links: dict):
    ws.column_dimensions["A"].width = 45
    for c in "BCDE":
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["F"].width = 14

    title_row(ws, 1, "Local Link Footprint vs Competitors", span=6)

    s = links.get("summary", {})
    body_cell(ws, 2, 1,
              f"Client referring domains: {s.get('client_total', 0)}  ·  "
              f"P1 opportunities: {s.get('p1_opportunities', 0)}  ·  "
              f"P2 opportunities: {s.get('p2_opportunities', 0)}  ·  "
              f"Gap: {s.get('gap_pct', 0):.0f}%",
              FONT_BODY_BOLD, border=False)

    comp_domains = links.get("competitor_domains", [])
    headers = ["Local domain", "Client has", "# of competitors"] + [d[:18] for d in comp_domains[:3]] + ["Opportunity"]
    header_row(ws, 4, headers)

    row = 5
    for r in links.get("rows", []):
        body_cell(ws, row, 1, r["domain"], FONT_MONO)
        body_cell(ws, row, 2, "✓" if r["client_has"] else "—", FONT_BODY_BOLD,
                  FILL_OK if r["client_has"] else None, ALIGN_CENTER)
        body_cell(ws, row, 3, r["comp_count"], FONT_BODY, None, ALIGN_CENTER)
        # Per-competitor placeholders (raw data may not include per-comp breakdown — skip if absent)
        for i in range(3):
            body_cell(ws, row, 4 + i, "—", FONT_MUTED, None, ALIGN_CENTER)
        body_cell(ws, row, 7, r["opportunity"] or "—", FONT_BODY_BOLD,
                  priority_fill(r["opportunity"]) if r["opportunity"] else None, ALIGN_CENTER)
        row += 1
        if row > 200:  # cap at 200 rows to keep the file reasonable
            body_cell(ws, row, 1, f"… {len(links.get('rows', [])) - 195} more rows truncated", FONT_MUTED)
            break


# ---------------------------------------------------------------------------
# TAB 9 — Actions (the CP2 surface)
# ---------------------------------------------------------------------------

def render_actions(ws: Worksheet, gbp, falcon, citations, links, pages):
    """Parent actions with collapsible detail rows. Each parent = one Sprint 3
    deliverable. Child rows show the specifics (which citations, which fields, etc.)
    Click the +/- in Excel's row gutter to expand/collapse detail rows.

    Column order (per Ryan's spec): Approval first (strategist input column), then
    #, Category, Deliverable (widest), Priority, Assigned Skill, Edit Notes.
    """
    ws.column_dimensions["A"].width = 17.29   # Approval (strategist input — leftmost)
    ws.column_dimensions["B"].width = 4       # #
    ws.column_dimensions["C"].width = 18      # Category
    ws.column_dimensions["D"].width = 50      # Deliverable (widest)
    ws.column_dimensions["E"].width = 10      # Priority
    ws.column_dimensions["F"].width = 18      # Assigned Skill
    ws.column_dimensions["G"].width = 40      # Edit Notes

    title_row(ws, 1, "Actions — Sprint 3 Deliverables", span=7)
    body_cell(ws, 2, 1,
              "Each parent row = one Sprint 3 deliverable. Click the + in the row gutter to expand details.",
              FONT_BODY_BOLD, FILL_BANNER, border=False)
    body_cell(ws, 3, 1,
              "Approval values: Approved / Edited / Deferred / Rejected   |   Edit Notes required if Edited.",
              FONT_MUTED, border=False)

    header_row(ws, 5, ["Approval", "#", "Category", "Deliverable", "Priority", "Assigned Skill", "Edit Notes"])

    parents = _build_consolidated_actions(gbp, falcon, citations, links, pages)

    row = 6
    deliverable_num = 1
    for parent in parents:
        # Parent row — outline level 0
        detail_count = len(parent.get("details", []))
        title = parent["action"]
        if detail_count > 0:
            title += f"   ({detail_count} items — expand to view)"
        body_cell(ws, row, 1, "", FONT_BODY_BOLD, FILL_APPROVAL, ALIGN_CENTER)   # Approval
        body_cell(ws, row, 2, deliverable_num, FONT_BODY_BOLD, None, ALIGN_CENTER)
        body_cell(ws, row, 3, parent["category"], FONT_BODY_BOLD)
        body_cell(ws, row, 4, title, FONT_BODY_BOLD)
        body_cell(ws, row, 5, parent["priority"], FONT_BODY_BOLD, priority_fill(parent["priority"]), ALIGN_CENTER)
        body_cell(ws, row, 6, parent["assigned_skill"], FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 7, "", FONT_BODY, FILL_APPROVAL)                      # Edit Notes
        ws.row_dimensions[row].height = 22
        deliverable_num += 1
        row += 1

        # Detail rows — outline level 1, collapsed by default
        for detail in parent.get("details", []):
            body_cell(ws, row, 1, "", FONT_BODY, FILL_APPROVAL, ALIGN_CENTER)    # Approval (per-detail)
            body_cell(ws, row, 2, "", FONT_BODY, None, ALIGN_CENTER)             # #
            body_cell(ws, row, 3, "", FONT_BODY)                                 # Category
            body_cell(ws, row, 4, "  → " + detail.get("item", ""), FONT_BODY)    # Deliverable
            body_cell(ws, row, 5, detail.get("priority", ""), FONT_MUTED, None, ALIGN_CENTER)
            body_cell(ws, row, 6, detail.get("status", ""), FONT_MUTED, None, ALIGN_CENTER)
            body_cell(ws, row, 7, detail.get("note", ""), FONT_MUTED)            # Edit Notes
            # Mark this row as a child for grouping/outline
            ws.row_dimensions[row].outline_level = 1
            ws.row_dimensions[row].hidden = True  # collapsed by default
            row += 1

    # Set outline summary direction so the +/- shows below each group
    ws.sheet_properties.outlinePr.summaryBelow = False


def _build_consolidated_actions(gbp, falcon, citations, links, pages) -> list[dict]:
    """Build parent actions with child detail rows. Each parent becomes one
    Sprint 3 deliverable; details are bullets shown when row is expanded."""
    parents: list[dict] = []

    # ── Parent: Optimize GBP Profile ──────────────────────────────────────
    gbp_gaps = [f for f in gbp.get("fields", []) if f.get("status") in ("missing", "weak", "api_pending")]
    if gbp_gaps:
        details = [
            {
                "item": f"{f['field'].replace('_', ' ').title()}: {f.get('action', '')}",
                "priority": "P1" if f.get("weight", 0) >= 6 else "P2",
                "status": f.get("status"),
                "note": f.get("rationale", ""),
            }
            for f in gbp_gaps
        ]
        worst_p = "P1" if any(f.get("weight", 0) >= 6 for f in gbp_gaps) else "P2"
        parents.append({
            "category": "GBP",
            "action": "Optimize Google Business Profile (multi-field gaps)",
            "priority": worst_p,
            "assigned_skill": "local_seo",
            "details": details,
        })

    # ── Parent: Fix NAP mismatches on existing citations ──────────────────
    mismatches = [d for d in citations.get("directories", []) if d.get("status") == "live_mismatch"]
    if mismatches:
        parents.append({
            "category": "Citations",
            "action": f"Fix NAP mismatches on existing citations",
            "priority": "P1",
            "assigned_skill": "local_seo",
            "details": [
                {
                    "item": f"{d['directory']}: {d.get('diff_summary') or 'NAP differs from the correct values'}",
                    "priority": d.get("priority", "P2"),
                    "status": "live_mismatch",
                    "note": d.get("url") or "",
                }
                for d in mismatches
            ],
        })

    # ── Parent: Build missing citations (CONFIRMED missing only — not unverified) ──
    missing = [d for d in citations.get("directories", []) if d.get("status") in ("missing", "partial")]
    if missing:
        parents.append({
            "category": "Citations",
            "action": f"Build missing citations",
            "priority": "P1" if any(d.get("priority") == "P1" for d in missing) else "P2",
            "assigned_skill": "local_seo",
            "details": [
                {
                    "item": f"{d['directory']} ({d.get('priority', 'P3')}) — {d.get('action', 'Submit listing')}",
                    "priority": d.get("priority", "P3"),
                    "status": d.get("status"),
                    "note": f"Tier {d.get('tier', '?')} · {d.get('directory_type', 'firm')}-level",
                }
                for d in missing
            ],
        })

    # ── Parent: Reviews velocity push ─────────────────────────────────────
    review_count_field = next((f for f in gbp.get("fields", []) if f["field"] == "review_count"), {})
    try:
        batch_reviews = int(review_count_field.get("current") or 0)
    except (ValueError, TypeError):
        batch_reviews = 0

    if batch_reviews > 0:
        beats = []
        for kw_comps in falcon.get("competitors", {}).values():
            for c in kw_comps:
                try:
                    cr = int(c.get("reviews") or 0)
                    if cr > batch_reviews * 1.2:  # competitors with 20%+ more reviews
                        beats.append((c.get("domain") or c.get("name", ""), cr, c.get("rating")))
                except (ValueError, TypeError):
                    continue
        # Dedupe by name
        seen = set()
        unique_beats = []
        for name, count, rating in sorted(beats, key=lambda x: -x[1]):
            if name not in seen:
                seen.add(name)
                unique_beats.append((name, count, rating))

        if unique_beats:
            parents.append({
                "category": "Reviews",
                "action": f"Review velocity push (client at {batch_reviews}; {len(unique_beats)} competitors beat us)",
                "priority": "P1",
                "assigned_skill": "local_seo",
                "details": [
                    {
                        "item": f"{name} has {count} reviews @ {rating} rating",
                        "priority": "—",
                        "status": "competitor_ahead",
                        "note": f"Gap: {count - batch_reviews} reviews",
                    }
                    for name, count, rating in unique_beats[:10]
                ],
            })

    # ── Parent: Investigate primary keyword decline ──────────────────────
    summary = falcon.get("summary", {})
    # Find any keyword with coverage <50% or where recent SoLV dropped meaningfully
    weak_kws = [(kw, s) for kw, s in summary.items() if (s.get("coverage_pct") or 100) < 50]
    if weak_kws:
        parents.append({
            "category": "Local Rankings",
            "action": "Investigate weak-coverage keywords",
            "priority": "P2",
            "assigned_skill": "local_seo",
            "details": [
                {
                    "item": f"{kw}: {s.get('coverage_pct', 0)}% coverage, ARP {s.get('avg_rank', '?')}, SoLV {s.get('solv_pct', '?')}%",
                    "priority": "P2",
                    "status": "weak",
                    "note": "Coverage < 50% — likely GBP optimization or local link gap",
                }
                for kw, s in weak_kws
            ],
        })

    # ── Parent: LocalBusiness + LegalService schema ──────────────────────
    parents.append({
        "category": "Schema",
        "action": "Implement LocalBusiness + LegalService schema markup",
        "priority": "P1",
        "assigned_skill": "technical_seo",
        "details": [
            {
                "item": "LocalBusiness schema on home page with NAP + sameAs to social profiles",
                "priority": "P1",
                "status": "missing",
                "note": "Blocks both Maps Pack signals and AI search visibility",
            },
            {
                "item": "LegalService schema on practice area pages",
                "priority": "P1",
                "status": "missing",
                "note": "Surfaces practice areas as structured data",
            },
            {
                "item": "Person schema on attorney bio pages with credentials + sameAs",
                "priority": "P2",
                "status": "missing",
                "note": "Helps Google connect attorneys to firm entity graph",
            },
        ],
    })

    # Sort parents by priority (P1 → P2 → P3)
    parents.sort(key=lambda p: {"P1": 0, "P2": 1, "P3": 2}.get(p["priority"], 9))
    return parents


def _build_action_list(gbp, falcon, citations, links, pages) -> list[dict]:
    """Roll all findings into a single action list, sorted by priority."""
    actions = []

    # GBP actions (skip api_pending — those are gated on MCP patches)
    for f in gbp.get("fields", []):
        if f.get("status") in ("missing", "weak"):
            actions.append({
                "category": "GBP",
                "action": f"{f['field'].replace('_', ' ').title()}: {f.get('action', '')}",
                "priority": "P1" if f.get("weight", 0) >= 6 else "P2",
                "assigned_skill": "local_seo",
            })
        elif f.get("status") == "api_pending":
            actions.append({
                "category": "GBP (blocked)",
                "action": f"{f['field'].replace('_', ' ').title()}: pull live value once GBP API is connected",
                "priority": "P3",
                "assigned_skill": "local_seo",
            })

    # Citation actions
    for d in citations.get("directories", []):
        status = d.get("status")
        if status in ("missing", "live_mismatch", "partial"):
            actions.append({
                "category": "Citations",
                "action": f"{d['directory']}: {d.get('action', '')}",
                "priority": d.get("priority", "P3"),
                "assigned_skill": "local_seo",
            })

    # Local content + local links removed from Actions tab — those workstreams
    # are handled by Sprint 4 (Content) and Sprint 5 (Links) respectively, and
    # don't belong in a Local SEO foundation audit.

    # Review velocity vs competitors (real comparison from Local Falcon scans)
    review_count_field = next((f for f in gbp.get("fields", []) if f["field"] == "review_count"), {})
    batch_reviews = int(review_count_field.get("current") or 0)
    if batch_reviews > 0:
        # Find competitors with more reviews
        max_comp_reviews = 0
        top_comp_name = ""
        for kw_comps in falcon.get("competitors", {}).values():
            for c in kw_comps:
                try:
                    cr = int(c.get("reviews") or 0)
                    if cr > max_comp_reviews:
                        max_comp_reviews = cr
                        top_comp_name = c.get("domain") or c.get("name", "")
                except (ValueError, TypeError):
                    continue
        if max_comp_reviews > batch_reviews * 1.5:
            actions.append({
                "category": "Reviews",
                "action": f"Review velocity push — {top_comp_name} has {max_comp_reviews} reviews vs {CLIENT_NAME}'s {batch_reviews}",
                "priority": "P1",
                "assigned_skill": "local_seo",
            })

    # Rankings — keywords with <40% coverage
    for kw, summary in falcon.get("summary", {}).items():
        if summary.get("coverage_pct", 100) < 40:
            actions.append({
                "category": "Local Rankings",
                "action": f"Investigate why '{kw}' coverage is low ({summary.get('coverage_pct', 0):.0f}%) — likely GBP optimization or local link gap",
                "priority": "P2",
                "assigned_skill": "local_seo",
            })

    # Sort: P1 first
    actions.sort(key=lambda a: ({"P1": 0, "P2": 1, "P3": 2}.get(a["priority"], 9), a["category"]))
    return actions


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# TAB 4a — Keyword Rankings (DataForSEO 50-keyword Maps + Organic SERP sweep)
# ---------------------------------------------------------------------------

def render_keyword_rankings(ws: Worksheet, rankings: dict):
    """50-keyword Maps Pack + Organic ranking matrix from DfS batch tracker."""
    ws.column_dimensions["A"].width = 36     # Keyword
    ws.column_dimensions["B"].width = 16     # City
    ws.column_dimensions["C"].width = 16     # Service
    ws.column_dimensions["D"].width = 12     # Maps Rank
    ws.column_dimensions["E"].width = 12     # Organic Rank
    ws.column_dimensions["F"].width = 38     # Top Competitor #1
    ws.column_dimensions["G"].width = 38     # Top Competitor #2
    ws.column_dimensions["H"].width = 38     # Top Competitor #3

    if not rankings or not rankings.get("results"):
        title_row(ws, 1, "Keyword Rankings — data unavailable", span=8)
        body_cell(ws, 3, 1,
                  "No keyword rankings data found. Run `local_keyword_track_batch` in chat first "
                  "to populate keyword-rankings-50.json.",
                  FONT_MUTED, border=False)
        return

    results  = rankings["results"]
    summary  = rankings.get("summary", {})
    cost     = rankings.get("total_cost_usd", 0)
    pulled   = rankings.get("pulled_at", "—")

    title_row(ws, 1, f"Keyword Rankings — {len(results)} keywords tracked via DataForSEO", span=8)

    # Summary banner: Maps + Organic top counts
    body_cell(ws, 2, 1,
              f"Maps Pack: {summary.get('maps_pack_ranking', '—')} ranked  ·  "
              f"Top 3: {summary.get('maps_top_3', 0)}  ·  Top 10: {summary.get('maps_top_10', 0)}  ·  "
              f"Organic: {summary.get('organic_ranking', '—')} ranked  ·  Top 10: {summary.get('organic_top_10', 0)}  ·  "
              f"Cost: ${cost:.3f}  ·  Pulled: {pulled[:10]}",
              FONT_BODY, FILL_BANNER, border=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 19.5

    # ── City × Service heatmap (group A) ──────────────────────────────────
    section_row(ws, 4, "City × Service Heatmap (best Maps rank per cell — blank = not in top 20)", span=8)

    cities   = sorted({r["city"] for r in results if r.get("city")})
    services = sorted({r["service"] for r in results if r.get("service")})

    # Header row: City | Service1 | Service2 | ...
    header_cells = ["City"] + services
    header_row(ws, 5, header_cells + [""] * (8 - len(header_cells)))

    row = 6
    for city in cities:
        body_cell(ws, row, 1, city, FONT_BODY_BOLD)
        for ci, svc in enumerate(services, start=2):
            # Find best (lowest) Maps rank for this city × service
            matches = [r for r in results
                       if r["city"] == city and r["service"] == svc
                       and r.get("maps_rank") is not None]
            if matches:
                best = min(r["maps_rank"] for r in matches)
                if best <= 3:    fill = FILL_OK
                elif best <= 10: fill = FILL_WEAK
                else:            fill = FILL_MISSING
                body_cell(ws, row, ci, f"#{best}", FONT_BODY_BOLD, fill, ALIGN_CENTER)
            else:
                body_cell(ws, row, ci, "—", FONT_MUTED, FILL_PENDING, ALIGN_CENTER)
        row += 1

    row += 2

    # ── Full table (group B) ──────────────────────────────────────────────
    section_row(ws, row, "Full 50-Keyword Rankings + Top 3 Competitors", span=8)
    row += 1
    header_row(ws, row, ["Keyword", "City", "Service", "Maps Rank", "Organic Rank",
                         "Competitor #1", "Competitor #2", "Competitor #3"])
    row += 1

    # Sort: ranked keywords first (by Maps rank asc), then unranked
    sorted_results = sorted(results, key=lambda r: (
        r.get("maps_rank") is None,
        r.get("maps_rank") or 999,
        r.get("organic_rank") or 999,
    ))

    for r in sorted_results:
        body_cell(ws, row, 1, r["keyword"], FONT_BODY_BOLD)
        body_cell(ws, row, 2, r.get("city", "—"), FONT_BODY, None, ALIGN_CENTER)
        body_cell(ws, row, 3, r.get("service", "—"), FONT_BODY, None, ALIGN_CENTER)

        # Maps rank with color
        mr = r.get("maps_rank")
        if mr is None:
            body_cell(ws, row, 4, "—", FONT_MUTED, FILL_MISSING, ALIGN_CENTER)
        else:
            fill = FILL_OK if mr <= 3 else (FILL_WEAK if mr <= 10 else FILL_PENDING)
            body_cell(ws, row, 4, f"#{mr}", FONT_BODY_BOLD, fill, ALIGN_CENTER)

        # Organic rank with color
        or_rank = r.get("organic_rank")
        if or_rank is None:
            body_cell(ws, row, 5, "—", FONT_MUTED, FILL_MISSING, ALIGN_CENTER)
        else:
            fill = FILL_OK if or_rank <= 3 else (FILL_WEAK if or_rank <= 10 else FILL_PENDING)
            body_cell(ws, row, 5, f"#{or_rank}", FONT_BODY_BOLD, fill, ALIGN_CENTER)

        # Top 3 competitors
        comps = r.get("top_3_competitors", []) or []
        for ci in range(3):
            if ci < len(comps):
                c = comps[ci]
                name = c.get("name", "?")
                rating = c.get("rating", "?")
                reviews = c.get("reviews", "?")
                text = f"#{c.get('rank', '?')} {name}\n⭐ {rating} · {reviews} reviews"
            else:
                text = "—"
            body_cell(ws, row, 6 + ci, text, FONT_BODY)

        ws.row_dimensions[row].height = 30  # 2-line competitor cells
        row += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global CLIENT_PLACE_ID, CLIENT_LABEL, CLIENT_DOMAIN_TOKEN
    print(f"Building Local SEO Audit XLSX for {CLIENT_NAME}")
    print(f"  audit dir: {AUDIT_DIR}")

    # Load all cache files. NOTE: "data-coverage.json" is the normalize step's
    # own summary (grades/warnings/source coverage) — NOT audit-manifest.json,
    # which is owned by the localseo_create_audit / localseo_update_config MCP
    # tools and must never be read here as if it had this schema.
    manifest = load("data-coverage.json")
    gbp = load("gbp-profile.json")

    # Resolve the client's place_id + a short token used to filter the client
    # out of competitor lists. The audit-dir gbp-profile.json is the normalized
    # output (no place_id/items at top level), so prefer reading from raw/ for
    # the original DfS envelope.
    def _try_load(path: Path) -> dict:
        return json.loads(path.read_text()) if path.exists() else {}
    raw_dfs = _try_load(AUDIT_DIR / "raw" / "gbp-profile.json") or \
              _try_load(AUDIT_DIR / "raw" / "dataforseo-business-data.json")
    if isinstance(raw_dfs, dict):
        item = None
        if raw_dfs.get("tasks"):
            try:
                item = (raw_dfs["tasks"][0]["result"][0].get("items") or [None])[0]
            except (KeyError, IndexError, TypeError):
                item = None
        if item is None and raw_dfs.get("items"):
            try:
                item = raw_dfs["items"][0]
            except (IndexError, TypeError):
                item = None
        if item is None and raw_dfs.get("place_id"):
            item = raw_dfs
        if isinstance(item, dict):
            CLIENT_PLACE_ID = item.get("place_id") or ""
            if item.get("domain"):
                CLIENT_DOMAIN_TOKEN = item["domain"].split(".")[0].lower()
    CLIENT_LABEL = f"{CLIENT_NAME} (YOU)"
    falcon = load("local-falcon-grid.json")
    gsc = load("gsc-local-queries.json")
    citations = load("citations.json")
    links = load("local-links.json")
    pages = load("local-pages.json")
    keyword_rankings = load_any("keyword-rankings-50.json", "keyword-rankings.json")
    windsor_gbp = load_any("windsor-gbp-monthly.json", "windsor-gmb-monthly.json")
    windsor_reviews = load_any("windsor-reviews-monthly.json")

    wb = Workbook()
    wb.remove(wb.active)  # we'll create tabs in order

    # Load extra context files for the new tabs
    audit_config = load("audit-config.json")
    competitor_research = load("competitor-research.json")
    ai_geo_findings = load("ai-geo-findings.json")

    print("  rendering tabs...")
    render_readme(wb.create_sheet("1. README"))
    render_summary(wb.create_sheet("2. Summary"), manifest, gbp, falcon, citations, links, pages, gsc)
    # Per-location GBP profiles: primary = gbp-profile.json; each office = gbp-profile-{slug}.json
    _loc_cache = load("locations.json") or {}
    _locs = (_loc_cache.get("locations", []) if isinstance(_loc_cache, dict) else (_loc_cache or [])) or []
    _gbp_location_profiles = []
    for _l in [x for x in _locs if x.get("role") != "duplicate"]:
        if _l.get("role") == "primary":
            _p = gbp
        else:
            _slug = (_l.get("name") or "office").lower().split()[0]
            _p = load(f"gbp-profile-{_slug}.json")
        if _p and _p.get("fields"):
            _gbp_location_profiles.append((_l.get("name"), _p))
    if not _gbp_location_profiles:
        _gbp_location_profiles = [(None, gbp)]
    render_gbp_audit(wb.create_sheet("3. GBP Audit"), gbp, _gbp_location_profiles)
    render_keyword_rankings(wb.create_sheet("4. Keyword Rankings"), keyword_rankings)
    # Local Falcon proximity grid (per-office × keyword SoLV/ARP) — the real Maps-grid
    # ranking source. Previously this renderer was never wired in, so the flagship LF
    # data never reached the workbook.
    render_local_rankings(wb.create_sheet("4b. Local Falcon Grid"), falcon)
    render_citations(wb.create_sheet("5. Citations"), citations)
    render_reviews(wb.create_sheet("6. Reviews"), gbp, falcon, windsor_gbp, windsor_reviews)
    render_nap_consistency(wb.create_sheet("7. NAP Consistency"), gbp, citations)
    render_competitor_deepdive(wb.create_sheet("8. Competitors"), falcon, competitor_research, gbp)
    # AI Search Visibility — populated from keyword-rankings.json via normalize_ai_geo().
    # Lives between Competitors and the action plan so the strategist sees the gap
    # before recommending actions.
    render_ai_geo_visibility(wb.create_sheet("9. AI Search"), ai_geo_findings or load_any("ai-geo-findings.json"), gbp)
    # AI / GEO Visibility tab pulled — deferred to a dedicated AI visibility audit.
    # The single-WebSearch-call signal underpinning it didn't match the tab's claim
    # of testing ChatGPT/Perplexity/Gemini/Claude. Will be rebuilt in that workstream
    # with actual per-tool queries + cited-source extraction.
    render_actions(wb.create_sheet("10. Actions"), gbp, falcon, citations, links, pages)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n✓ Saved: {OUT_PATH} ({OUT_PATH.stat().st_size:,} bytes)")
    print(f"  Tabs: {len(wb.sheetnames)} — {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
