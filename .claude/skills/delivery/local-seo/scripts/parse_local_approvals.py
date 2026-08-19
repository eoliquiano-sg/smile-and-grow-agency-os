#!/usr/bin/env python3
"""parse_local_approvals.py — read the strategist-edited Local SEO audit xlsx
and emit a structured approvals JSON.

Workflow:
  1. build_local_audit_xlsx.py writes {slug}-local-audit.xlsx with an Actions tab.
  2. Strategist fills the Approval column on each row: Approved / Edited / Deferred / Rejected.
  3. They optionally edit the Action / Specific Next Step cells inline.
  4. This script parses the workbook and emits {slug}-local-approvals.json with shape:

{
  "audit_id":       "...",
  "totals":         {"Approved": n, "Edited": n, "Rejected": n, "Deferred": n, "Unmarked": n},
  "buckets": {
    "Approved":     [<row>, ...],
    "Edited":       [<row>, ...],
    "Rejected":     [<row>, ...],
    "Deferred":     [<row>, ...],
    "Unmarked":     [<row>, ...]
  }
}

Each <row>: row, approval, action, category, priority, sprint_step, assigned_skill,
            specific_next_step, source_finding.

build_local_project_plan.py consumes the Approved + Edited buckets to emit Sprint 3
deliverables.
"""

import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook


parser = argparse.ArgumentParser()
parser.add_argument("--client-slug", required=True)
parser.add_argument("--audit-id", help="Audit folder name under clients/{slug}/local-audit/")
parser.add_argument("--audit-dir", help="Explicit audit folder path (overrides --audit-id)")
parser.add_argument("--workspace-root", default=os.environ.get("AGENCY_OS_ROOT", "."))
args = parser.parse_args()

SLUG = args.client_slug

if args.audit_dir:
    AUDIT_DIR = Path(args.audit_dir)
elif args.audit_id:
    AUDIT_DIR = Path(args.workspace_root) / "clients" / SLUG / "local-audit" / args.audit_id
else:
    sys.exit("Provide --audit-id or --audit-dir.")

XLSX = AUDIT_DIR / f"{SLUG}-local-audit.xlsx"
OUT  = AUDIT_DIR / f"{SLUG}-local-approvals.json"

if not XLSX.exists():
    sys.exit(f"xlsx not found: {XLSX}")

wb = load_workbook(XLSX, data_only=True)
# Tab name is "Actions" in the legacy layout and "9. Actions" in the numbered layout
# produced by build_local_audit_xlsx.py — accept either.
actions_tab = next(
    (name for name in wb.sheetnames if name == "Actions" or name.lower().endswith(" actions")),
    None,
)
if not actions_tab:
    sys.exit("Actions tab missing from workbook. Re-run build_local_audit_xlsx.py.")
ws = wb[actions_tab]

# Locate columns by header (case-insensitive, trimmed). The numbered build wraps
# the table in banner rows, so the header row isn't always row 1 — scan the
# first 8 rows for a row that contains both 'approval' and 'deliverable'/'action'.
def _find_header_row(ws) -> int:
    for r in range(1, min(8, ws.max_row) + 1):
        labels = [str(c.value or "").strip().lower() for c in ws[r]]
        if "approval" in labels and ("action" in labels or "deliverable" in labels):
            return r
    return 1
header_row = _find_header_row(ws)
headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[str(cell.value).strip().lower()] = cell.column
# 'deliverable' is the column name in the numbered build; aliased to 'action'.
if "action" not in headers and "deliverable" in headers:
    headers["action"] = headers["deliverable"]

def col(name: str) -> int | None:
    return headers.get(name.lower())

REQUIRED = ["approval", "action"]
missing = [c for c in REQUIRED if not col(c)]
if missing:
    sys.exit(f"Actions tab missing required columns: {missing}")

approval_col = col("approval")
action_col = col("action")
category_col = col("category")
priority_col = col("priority")
sprint_step_col = col("sprint step") or col("sprint_step") or col("step")
skill_col = col("assigned skill") or col("assigned_skill") or col("skill")
next_step_col = col("specific next step") or col("specific_next_step") or col("next step")
finding_col = col("source finding") or col("source_finding") or col("finding")

VALID = {"approved", "edited", "deferred", "rejected"}
buckets: dict[str, list] = {"Approved": [], "Edited": [], "Deferred": [], "Rejected": [], "Unmarked": []}

def get(row_idx: int, c: int | None) -> str:
    if not c:
        return ""
    v = ws.cell(row=row_idx, column=c).value
    return "" if v is None else str(v).strip()

for r in range(header_row + 1, ws.max_row + 1):
    action = get(r, action_col)
    if not action:
        continue  # skip blank rows
    approval_raw = get(r, approval_col).lower()
    bucket = "Unmarked"
    for valid in VALID:
        if approval_raw == valid:
            bucket = valid.capitalize()
            break
    record = {
        "row": r,
        "approval": bucket,
        "action": action,
        "category": get(r, category_col),
        "priority": get(r, priority_col),
        "sprint_step": get(r, sprint_step_col),
        "assigned_skill": get(r, skill_col),
        "specific_next_step": get(r, next_step_col),
        "source_finding": get(r, finding_col),
    }
    buckets[bucket].append(record)

totals = {k: len(v) for k, v in buckets.items()}

output = {
    "audit_id": args.audit_id or AUDIT_DIR.name,
    "xlsx_path": str(XLSX),
    "totals": totals,
    "buckets": buckets,
}

OUT.write_text(json.dumps(output, indent=2))
print(f"Wrote {OUT}")
print(f"  Approved: {totals['Approved']}  Edited: {totals['Edited']}  "
      f"Deferred: {totals['Deferred']}  Rejected: {totals['Rejected']}  "
      f"Unmarked: {totals['Unmarked']}")
if totals["Unmarked"] > 0:
    print(f"  ⚠ {totals['Unmarked']} rows are unmarked — strategist should review before proceeding.")
