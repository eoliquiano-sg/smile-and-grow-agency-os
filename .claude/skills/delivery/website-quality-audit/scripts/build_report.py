#!/usr/bin/env python3
"""Build the WQA visual report — v2 with 5 sections + title slides.

See SKILL.md for the workflow. Pre-builds all HTML chunks before the main
template to avoid f-string nesting issues.
"""

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import date
from openpyxl import load_workbook


# ============ CLI ============
parser = argparse.ArgumentParser()
parser.add_argument("--client-slug", required=True)
parser.add_argument("--audit-id")
parser.add_argument("--audit-dir")
parser.add_argument("--workspace-root", default=os.environ.get("AGENCY_OS_ROOT", "."))
parser.add_argument("--root-domain", required=True)
parser.add_argument("--client-name", default=None)
parser.add_argument("--primary-color", default="#2563EB")
args = parser.parse_args()

SLUG = args.client_slug
ROOT = args.root_domain.rstrip("/")
PRIMARY = args.primary_color
CLIENT_NAME = args.client_name or SLUG.replace("-", " ").title()
A_BASE = args.audit_dir or os.path.join(args.workspace_root, "clients", SLUG, "wqa", "audits", args.audit_id or "")

XLSX        = os.path.join(A_BASE, f"{SLUG}-wqa-data.xlsx")
APPROVALS   = os.path.join(A_BASE, f"{SLUG}-approvals.json")
KEYWORDS    = os.path.join(A_BASE, f"{SLUG}-ahrefs-keywords.json")
GA4_MONTHLY = os.path.join(A_BASE, f"{SLUG}-ga4-monthly.json")
GSC_MONTHLY = os.path.join(A_BASE, f"{SLUG}-gsc-monthly.json")
GA4_90D     = os.path.join(A_BASE, f"{SLUG}-ga4-90d-compare.json")
GSC_CURRENT = os.path.join(A_BASE, f"{SLUG}-gsc-90d-current.json")
GSC_PRIOR   = os.path.join(A_BASE, f"{SLUG}-gsc-90d-prior.json")
TARGETS     = os.path.join(A_BASE, f"{SLUG}-target-pages.json")
PSI_TARGETS = os.path.join(A_BASE, f"{SLUG}-psi-targets.json")
OUT         = os.path.join(A_BASE, f"{SLUG}-wqa-report.html")


def safe_load(path, default):
    if not os.path.exists(path): return default
    return json.load(open(path))

approvals    = safe_load(APPROVALS, {"totals": {}, "buckets": {"Approved": [], "Edit": [], "Rejected": [], "Deferred": [], "Unmarked": []}, "edits_detected": []})
keywords     = safe_load(KEYWORDS, [])
ga4_monthly  = safe_load(GA4_MONTHLY, {"organic": [], "total": []})
gsc_monthly  = safe_load(GSC_MONTHLY, [])
ga4_90d      = safe_load(GA4_90D, {"organic": {"current": {}, "prior": {}}, "total": {"current": {}, "prior": {}}, "periods": {}})
gsc_current  = safe_load(GSC_CURRENT, {"by_page": {}, "site": {}, "period": {}})
gsc_prior    = safe_load(GSC_PRIOR, {"by_page": {}, "site": {}, "period": {}})
psi_targets  = safe_load(PSI_TARGETS, [])  # currently unused — PSI removed from report

# Guard: the report is worthless without performance data. If the Phase 1 pulls
# were skipped, warn LOUDLY rather than silently emit a report with blank
# KPIs/charts. Populating this data is a REQUIRED step, not optional.
_perf_missing = []
if not gsc_monthly:
    _perf_missing.append("gsc-monthly")
if not (ga4_monthly.get("organic") or ga4_monthly.get("total")):
    _perf_missing.append("ga4-monthly")
if not gsc_current.get("by_page"):
    _perf_missing.append("gsc-90d-current")
if _perf_missing:
    print(
        "\n*** WARNING: performance data missing/empty -> "
        + ", ".join(_perf_missing)
        + ".\n    The report's Performance Metrics KPIs and trend chart will be BLANK.\n"
        "    Run the Phase 1 Windsor pulls (scoped to the client's domain) and\n"
        "    generate the *-gsc-monthly / *-ga4-monthly / *-gsc-90d-current files\n"
        "    BEFORE build_report.py. See WQA SKILL.md §1.2a.\n"
    )


def norm(u):
    if not u: return ""
    return u.split("?")[0].split("#")[0].rstrip("/").lower()


# ============ Read Aggregator ============
wb = load_workbook(XLSX, data_only=True)
ws = wb["Aggregator"]
hdrs = {c.value: c.column for c in ws[1] if c.value}

def cv(r, name):
    if name not in hdrs: return None
    return r[hdrs[name] - 1].value

def cv_any(r, names):
    """Read the first column name that exists in the headers."""
    for n in names:
        if n in hdrs:
            return r[hdrs[n] - 1].value
    return None

# Build per-URL GSC position lookup from the 90d-current file (col was dropped from Aggregator)
gsc_pos_by_url = {}
for u_norm, stats in (gsc_current.get("by_page") or {}).items():
    gsc_pos_by_url[u_norm] = stats.get("position", 0)

rows = []
for r in ws.iter_rows(min_row=2, values_only=False):
    if not cv(r, "Address"): continue
    addr_cell = r[hdrs["Address"] - 1]
    full_url = addr_cell.hyperlink.target if addr_cell.hyperlink else None
    u_norm = norm(full_url) if full_url else ""
    rows.append({
        "url_action": cv_any(r, ["URL Action", "Technical Action"]) or "",
        "address": cv(r, "Address") or "",
        "full_url": full_url,
        "status": cv(r, "Status"),
        "status_type": cv(r, "Status Type") or "",
        "page_type": cv(r, "Page Type") or "Other",
        "funnel": cv(r, "Funnel Stage") or "Other",
        "title": cv(r, "Title") or "",
        "word_count": cv(r, "Word Count") or 0,
        "ga4_sessions":    cv_any(r, ["GA4 Sessions (90d)", "GA4 Sessions"]) or 0,
        "ga4_conversions": cv_any(r, ["GA4 Conv (90d)", "GA4 Conv"]) or 0,
        "gsc_clicks":      cv_any(r, ["GSC Clicks (90d)", "GSC Clicks"]) or 0,
        "gsc_impressions": cv_any(r, ["GSC Imp (90d)", "GSC Imp"]) or 0,
        "gsc_ctr":         cv_any(r, ["GSC CTR (90d)", "GSC CTR"]) or 0,
        "gsc_position":    gsc_pos_by_url.get(u_norm, 0),   # read from 90d-current file
        "ref_domains": cv(r, "Ref Domains") or 0,
        "top_keyword": cv(r, "Top Keyword") or "",
    })


# ============ Helpers ============
def num_or_zero(x):
    return x if isinstance(x, (int, float)) else 0

def fmt(n):
    if n is None or n == "": return "—"
    if isinstance(n, float):
        if abs(n - round(n)) < 0.001: return f"{int(n):,}"
        return f"{n:,.1f}"
    return f"{int(n):,}"

def fmt_pct(n):
    if n is None or n == 0: return "—"
    return f"{n*100:.1f}%"

def pct_change(cur, prev):
    if not prev or prev == 0: return None
    return (cur - prev) / prev

def trend_arrow(delta, inverse=False):
    if delta is None: return '<span class="trend neutral">—</span>'
    pct = delta * 100
    up = pct > 0
    good = up if not inverse else (not up)
    arrow = "▲" if up else ("▼" if pct < 0 else "—")
    cls = "trend-up-good" if (up and good) else ("trend-up-bad" if (up and not good) else
          "trend-down-bad" if (not up and not good) else "trend-down-good")
    return f'<span class="trend {cls}">{arrow} {abs(pct):.1f}%</span>'

def esc(s):
    if s is None: return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

def addr_link(addr, url=None):
    if not addr: return ""
    if url is None:
        url = ROOT + (addr if addr.startswith("/") else "/" + addr) if not addr.startswith("http") else addr
    return f'<a href="{esc(url)}" target="_blank">{esc(addr)}</a>'


# ============ Compute trends ============
ga4_org_cur = ga4_90d.get("organic", {}).get("current", {})
ga4_org_pri = ga4_90d.get("organic", {}).get("prior", {})
delta_sessions = pct_change(ga4_org_cur.get("sessions", 0), ga4_org_pri.get("sessions", 0))
delta_conv = pct_change(ga4_org_cur.get("conversions", 0), ga4_org_pri.get("conversions", 0))

gsc_cur_site = gsc_current.get("site", {})
gsc_pri_site = gsc_prior.get("site", {})
delta_clicks = pct_change(gsc_cur_site.get("clicks", 0), gsc_pri_site.get("clicks", 0))
delta_imp = pct_change(gsc_cur_site.get("impressions", 0), gsc_pri_site.get("impressions", 0))
delta_ctr = pct_change(gsc_cur_site.get("ctr", 0), gsc_pri_site.get("ctr", 0))
delta_pos = pct_change(gsc_cur_site.get("position", 0), gsc_pri_site.get("position", 0))

cur_period = gsc_current.get("period", {})
prior_period = gsc_prior.get("period", {})
period_label = f"{cur_period.get('date_from', '—')} to {cur_period.get('date_to', '—')}"
prior_period_label = f"{prior_period.get('date_from', '—')} to {prior_period.get('date_to', '—')}"


# ============ Aggregations ============
live_rows = [r for r in rows if r["status_type"] == "live"]

clicks_by_pt = defaultdict(lambda: {"count": 0, "clicks": 0, "impressions": 0, "sessions": 0, "conversions": 0, "rows": []})
for r in live_rows:
    pt = r["page_type"]
    clicks_by_pt[pt]["count"] += 1
    clicks_by_pt[pt]["clicks"] += num_or_zero(r["gsc_clicks"])
    clicks_by_pt[pt]["impressions"] += num_or_zero(r["gsc_impressions"])
    clicks_by_pt[pt]["sessions"] += num_or_zero(r["ga4_sessions"])
    clicks_by_pt[pt]["conversions"] += num_or_zero(r["ga4_conversions"])
    clicks_by_pt[pt]["rows"].append(r)

clicks_by_fn = defaultdict(lambda: {"count": 0, "clicks": 0, "rows": []})
for r in live_rows:
    fn = r["funnel"]
    clicks_by_fn[fn]["count"] += 1
    clicks_by_fn[fn]["clicks"] += num_or_zero(r["gsc_clicks"])
    clicks_by_fn[fn]["rows"].append(r)

total_clicks_90d = gsc_cur_site.get("clicks", 0)
total_imp_90d = gsc_cur_site.get("impressions", 0)
total_ctr_90d = gsc_cur_site.get("ctr", 0)
total_pos_90d = gsc_cur_site.get("position", 0)
total_sessions_90d = ga4_org_cur.get("sessions", 0)
total_conv_90d = ga4_org_cur.get("conversions", 0)

status_mix = Counter(r["status_type"] for r in rows)
broken_with_signal = sorted([r for r in rows if r["status_type"] == "broken"],
                            key=lambda r: -(num_or_zero(r["gsc_impressions"]) + num_or_zero(r["ref_domains"]) * 100))[:15]
thin_count = sum(1 for r in live_rows if 0 < num_or_zero(r["word_count"]) < 618)

def kw_bucket(p):
    if p is None: return "?"
    if p <= 3: return "Top 3"
    if p <= 10: return "4-10"
    if p <= 20: return "11-20"
    return "21+"
kw_buckets = Counter(kw_bucket(k.get("best_position")) for k in keywords)
kw_intent = Counter()
for k in keywords:
    flags = []
    if k.get("is_branded"): flags.append("Branded")
    if k.get("is_commercial"): flags.append("Commercial")
    if k.get("is_informational"): flags.append("Informational")
    if k.get("is_transactional"): flags.append("Transactional")
    if k.get("is_navigational"): flags.append("Navigational")
    if not flags: flags.append("Other")
    for f in flags: kw_intent[f] += 1

top_keywords = sorted(keywords, key=lambda k: -(k.get("sum_traffic") or 0))[:15]

# ============ Technical aggregations ============

# Indexability breakdown across all known URLs (not just live)
indexability_counts = Counter()
for r in rows:
    label = "Indexable" if (r.get("title") or "") and r["status_type"] == "live" else \
            ("Broken/Redirect" if r["status_type"] in ("broken", "redirect") else "Non-indexable")
    indexability_counts[label] += 1

# Pages missing critical elements (live HTML pages only)
missing_title = sum(1 for r in live_rows if not (r.get("title") or "").strip())
missing_meta = 0  # We don't reliably have meta in the aggregator slice — flag later
missing_h1 = 0   # Same. Compute if we can.
# We can read these directly from the xlsx
for r in ws.iter_rows(min_row=2, values_only=False):
    if not cv(r, "Address"): continue
    if (cv(r, "Status Type") or "") != "live": continue
    if not (cv(r, "Meta Description") or "").strip(): missing_meta += 1
    if not (cv(r, "H1") or "").strip(): missing_h1 += 1

# Crawl depth distribution (live pages) — read directly from xlsx since rows[] doesn't carry crawl_depth
crawl_depth_buckets = defaultdict(int)
for r in ws.iter_rows(min_row=2, values_only=False):
    if not cv(r, "Address"): continue
    if (cv(r, "Status Type") or "") != "live": continue
    cd = cv(r, "Crawl Depth")
    if cd is None: continue
    try:
        cd_i = int(cd)
    except (TypeError, ValueError):
        continue
    bucket = "0 (Home)" if cd_i == 0 else \
             "1 (Top-level)" if cd_i == 1 else \
             "2" if cd_i == 2 else \
             "3" if cd_i == 3 else \
             "4+" if cd_i >= 4 else "?"
    crawl_depth_buckets[bucket] += 1

# Redirects → home (high-priority: lose topical relevance)
redirects_to_home = []
for r in rows:
    if r["status_type"] != "redirect": continue
    # Need the redirect_target column
for r in ws.iter_rows(min_row=2, values_only=False):
    if not cv(r, "Address"): continue
    if (cv(r, "Status Type") or "") != "redirect": continue
    target = (cv(r, "Redirect Target") or "").rstrip("/")
    if target == ROOT:
        redirects_to_home.append({
            "address": cv(r, "Address"),
            "url": r[hdrs["Address"] - 1].hyperlink.target if r[hdrs["Address"] - 1].hyperlink else None,
            "gsc_impressions": cv(r, "GSC Imp") or 0,
            "ref_domains": cv(r, "Ref Domains") or 0,
        })
redirects_to_home.sort(key=lambda x: -num_or_zero(x["gsc_impressions"]))
redirects_to_home = redirects_to_home[:10]

# Action-type counts moved below — after tech_recs/content_recs are classified


# ============ Content aggregations ============

# Word count distribution (live pages only)
WC_BUCKETS = [(0, 300, "<300 (very thin)"), (300, 617, "300-617 (thin)"),
              (618, 1500, "618-1,500 (medium)"), (1500, 3000, "1,500-3,000 (good)"),
              (3000, 10**9, "3,000+ (long-form)")]
wc_buckets = Counter()
for r in live_rows:
    wc = num_or_zero(r["word_count"])
    if wc == 0: continue
    for lo, hi, label in WC_BUCKETS:
        if lo <= wc <= hi:
            wc_buckets[label] += 1
            break

# Avg word count by page type (live only)
wc_by_pt = defaultdict(lambda: {"sum": 0, "count": 0, "thin": 0})
for r in live_rows:
    pt = r["page_type"]
    wc = num_or_zero(r["word_count"])
    if wc == 0: continue
    wc_by_pt[pt]["sum"] += wc
    wc_by_pt[pt]["count"] += 1
    if wc < 618: wc_by_pt[pt]["thin"] += 1
for pt, d in wc_by_pt.items():
    d["avg"] = int(d["sum"] / d["count"]) if d["count"] else 0

# Action-type counts moved below — after rec classification

# Overall content KPIs
all_wcs = [num_or_zero(r["word_count"]) for r in live_rows if num_or_zero(r["word_count"]) > 0]
median_wc = sorted(all_wcs)[len(all_wcs)//2] if all_wcs else 0
avg_wc = int(sum(all_wcs) / len(all_wcs)) if all_wcs else 0

# Striking distance + trends
gsc_cur_by_url = gsc_current.get("by_page", {})
gsc_pri_by_url = gsc_prior.get("by_page", {})
striking = []
for r in rows:
    pos = num_or_zero(r["gsc_position"])
    imp = num_or_zero(r["gsc_impressions"])
    if not (4 <= pos <= 15 and imp >= 200): continue
    u_norm = norm(r["full_url"]) if r["full_url"] else None
    cur = gsc_cur_by_url.get(u_norm) if u_norm else None
    pri = gsc_pri_by_url.get(u_norm) if u_norm else None
    if cur and pri:
        d_clicks = pct_change(cur.get("clicks", 0), pri.get("clicks", 0))
        d_imp = pct_change(cur.get("impressions", 0), pri.get("impressions", 0))
        d_ctr = pct_change(cur.get("ctr", 0), pri.get("ctr", 0))
        d_pos = pct_change(cur.get("position", 0), pri.get("position", 0))
    else:
        d_clicks = d_imp = d_ctr = d_pos = None
    striking.append({
        "address": r["address"], "url": r["full_url"],
        "top_keyword": r.get("top_keyword", ""),
        "position": pos, "impressions": imp,
        "clicks": num_or_zero(r["gsc_clicks"]),
        "ctr": num_or_zero(r["gsc_ctr"]),
        "d_clicks": d_clicks, "d_imp": d_imp, "d_ctr": d_ctr, "d_pos": d_pos,
    })
striking.sort(key=lambda x: -x["impressions"])
striking = striking[:20]


# Recommendations
active_recs = approvals.get("buckets", {}).get("Approved", []) + approvals.get("buckets", {}).get("Edit", [])
deferred_recs = approvals.get("buckets", {}).get("Deferred", [])

def classify_rec(rec):
    """The parser now writes Category directly from the Recommendations tab.
    Fallback heuristic kept for legacy approvals files."""
    cat = (rec.get("category") or "").lower()
    if cat == "technical": return "technical"
    if cat == "content":   return "content"
    a = rec.get("action_type", "")
    if a in {"Fix 404", "301 redirect", "Evaluate redirect", "Add canonical",
             "Add schema", "Noindex", "Update robots", "Indexability fix",
             "Sitemap fix", "Monitor"}: return "technical"
    return "content"

tech_recs    = [r for r in active_recs if classify_rec(r) == "technical"]
content_recs = [r for r in active_recs if classify_rec(r) == "content"]
deferred_tech    = [r for r in deferred_recs if classify_rec(r) == "technical"]
deferred_content = [r for r in deferred_recs if classify_rec(r) == "content"]

# Action-type counts now that classifications exist
tech_action_counts = Counter(r.get("action_type") for r in tech_recs)
content_action_counts = Counter(r.get("action_type") for r in content_recs)


# ============ Target pages (read from approvals — strategist-confirmed) ============
# Approved + Edit only. Unmarked rows are excluded until strategist approves.

PT_VALUE = {"Money Page": 6.0, "Authority Page": 1.8, "Content Hub": 1.3, "Blog Post": 1.0}

tp_data = approvals.get("target_pages", {})
tp_buckets = tp_data.get("buckets", {}) if isinstance(tp_data, dict) else {}
approved_targets = (tp_buckets.get("Approved") or []) + (tp_buckets.get("Edit") or [])
unmarked_targets = tp_buckets.get("Unmarked") or []
rejected_targets = tp_buckets.get("Rejected") or []
deferred_targets = tp_buckets.get("Deferred") or []

# Final roster for the report = approved/edited rows in their original rank order
target_pages = sorted(approved_targets, key=lambda t: t.get("rank") or 999)
print(f"Target pages — Approved+Edit: {len(approved_targets)} · Unmarked: {len(unmarked_targets)} · Rejected: {len(rejected_targets)} · Deferred: {len(deferred_targets)}")


# ============ Styling ============
PT_COLORS = {"Homepage": "#7c3aed", "Money Page": "#9333ea", "Authority Page": "#0ea5e9",
             "Content Hub": "#2563eb", "Blog Post": "#6366f1", "Utility Page": "#64748b"}
FN_COLORS = {"Brand": "#7c3aed", "Bottom Funnel": "#dc2626", "Mid Funnel": "#ea580c",
             "Top Funnel": "#0ea5e9", "Utility": "#64748b"}
STATUS_COLORS = {"live": "#16a34a", "redirect": "#eab308", "broken": "#dc2626",
                 "rate_limited": "#6366f1", "unknown": "#94a3b8"}


def render_kpi(label, value, sub=None, delta=None, delta_inverse=False):
    arrow = trend_arrow(delta, inverse=delta_inverse) if delta is not None else ""
    sub_html = '<div class="sub">' + esc(sub) + '</div>' if sub else ""
    return ('<div class="kpi">'
            '<div class="label">' + esc(label) + '</div>'
            '<div class="value">' + str(value) + '</div>'
            + arrow + sub_html + '</div>')


def render_rec_card(rec, idx=None):
    prio = rec.get("priority") or ""
    action = rec.get("action_type") or ""
    addr = rec.get("page_address") or ""
    url = rec.get("url") or (ROOT + addr if addr.startswith("/") else None)
    next_step = rec.get("specific_next_step") or "—"
    current = rec.get("current_state") or "—"
    target = rec.get("target_state") or "—"
    why = rec.get("why_it_matters") or "—"
    hours = rec.get("_legacy_hours") or "—"
    sprint = rec.get("sprint") or "—"
    edited = rec.get("edited")
    edit_badge = '<span class="edit-badge">edited</span>' if edited else ''
    num_chip = ('<span class="rec-num">' + str(idx) + '</span>') if idx else ''
    return ('<div class="rec-card">'
        '<div class="rec-head">'
          '<div class="rec-meta">' + num_chip +
            '<span class="badge prio-' + prio + '">' + esc(prio) + '</span>'
            '<span class="pill">' + esc(action) + '</span>' + edit_badge +
          '</div>'
          '<div class="rec-addr">' + addr_link(addr, url=url) + '</div>'
        '</div>'
        '<div class="rec-body">'
          '<div class="rec-step">' + esc(next_step) + '</div>'
          '<div class="rec-grid">'
            '<div><div class="rec-label">Current</div><div class="rec-value">' + esc(current) + '</div></div>'
            '<div><div class="rec-label">Target</div><div class="rec-value">' + esc(target) + '</div></div>'
            '<div><div class="rec-label">Why it matters</div><div class="rec-value">' + esc(why) + '</div></div>'
            '<div><div class="rec-label">Sprint · Hours</div><div class="rec-value">' + esc(sprint) + ' · ' + esc(str(hours)) + 'h</div></div>'
          '</div>'
        '</div>'
      '</div>')


# ============ Pre-build HTML chunks ============

# 1a KPI strip
kpi_strip = ''.join([
    render_kpi("Organic sessions", fmt(total_sessions_90d), "GA4 · medium = organic", delta=delta_sessions),
    render_kpi("Organic conversions", fmt(total_conv_90d), "GA4", delta=delta_conv),
    render_kpi("GSC clicks", fmt(total_clicks_90d), "Google search results", delta=delta_clicks),
    render_kpi("GSC impressions", fmt(total_imp_90d), None, delta=delta_imp),
    render_kpi("Avg CTR", fmt_pct(total_ctr_90d), None, delta=delta_ctr),
    render_kpi("Avg position", fmt(total_pos_90d), "impression-weighted", delta=delta_pos, delta_inverse=True),
    render_kpi("Indexable URLs", fmt(sum(1 for r in rows if (r.get('title') or ''))), f"of {len(rows)} total"),
    render_kpi("Ranking keywords", fmt(len(keywords)), "Ahrefs · US"),
])

# 1c By-page-type summary table
def _share(c, total): return fmt_pct(c/total) if total else "—"
pt_rows = []
for pt, d in sorted(clicks_by_pt.items(), key=lambda kv: -kv[1]["clicks"]):
    pt_rows.append(
        '<tr><td><span class="badge" style="background:' + PT_COLORS.get(pt, "#94a3b8") + '">' + esc(pt) + '</span></td>'
        '<td class="num">' + fmt(d["count"]) + '</td>'
        '<td class="num">' + fmt(d["clicks"]) + '</td>'
        '<td class="num">' + _share(d["clicks"], total_clicks_90d) + '</td></tr>')
pt_summary_tbody = ''.join(pt_rows)

# 1c Accordions per page type
pt_accordions = []
for pt, d in sorted(clicks_by_pt.items(), key=lambda kv: -kv[1]["clicks"]):
    page_rows_html = []
    sorted_pages = sorted(d["rows"], key=lambda r: -num_or_zero(r["gsc_clicks"]))[:50]
    for r in sorted_pages:
        page_rows_html.append(
            '<tr><td>' + addr_link(r["address"], url=r["full_url"]) + '</td>'
            '<td class="num">' + fmt(r["ga4_sessions"]) + '</td>'
            '<td class="num">' + fmt(r["gsc_clicks"]) + '</td>'
            '<td class="num">' + fmt(r["gsc_impressions"]) + '</td>'
            '<td class="num">' + fmt(r["gsc_position"]) + '</td>'
            '<td class="num">' + fmt(r["ga4_conversions"]) + '</td></tr>')
    extra_note = ''
    if d["count"] > 50:
        extra_note = '<div style="font-size:11px; color:var(--muted); margin-top:6px;">Showing top 50 of ' + str(d["count"]) + ' URLs in this category.</div>'
    pt_accordions.append(
        '<details class="accordion">'
          '<summary>' + esc(pt) + ' — ' + str(d["count"]) + ' URLs · ' + fmt(d["clicks"]) + ' clicks</summary>'
          '<div class="acc-body">'
            '<table class="compact"><thead><tr>'
              '<th>Address</th><th class="num">Sessions</th><th class="num">Clicks</th>'
              '<th class="num">Impressions</th><th class="num">Avg Pos</th><th class="num">Conversions</th>'
            '</tr></thead><tbody>' + ''.join(page_rows_html) + '</tbody></table>'
            + extra_note +
          '</div>'
        '</details>')
pt_accordions_html = ''.join(pt_accordions)

# 1d Funnel summary
fn_rows = []
for fn, d in sorted(clicks_by_fn.items(), key=lambda kv: -kv[1]["clicks"]):
    fn_rows.append(
        '<tr><td><span class="badge" style="background:' + FN_COLORS.get(fn, "#94a3b8") + '">' + esc(fn) + '</span></td>'
        '<td class="num">' + fmt(d["count"]) + '</td>'
        '<td class="num">' + fmt(d["clicks"]) + '</td>'
        '<td class="num">' + _share(d["clicks"], total_clicks_90d) + '</td></tr>')
fn_summary_tbody = ''.join(fn_rows)

# 1e Keywords
def _intent_str(k):
    flags = []
    for n, key in [("Branded", "is_branded"), ("Commercial", "is_commercial"),
                   ("Informational", "is_informational"), ("Transactional", "is_transactional"),
                   ("Navigational", "is_navigational")]:
        if k.get(key): flags.append(n)
    return ", ".join(flags) or "—"

kw_rows = []
for k in top_keywords:
    kw_rows.append(
        '<tr><td>' + esc(k.get("keyword", "")) + '</td>'
        '<td>' + _intent_str(k) + '</td>'
        '<td class="num">' + fmt(k.get("best_position")) + '</td>'
        '<td class="num">' + fmt(k.get("volume")) + '</td></tr>')
kw_tbody = ''.join(kw_rows)

# 1f Striking distance
def _td_with_trend(value_html, trend_html):
    return '<td class="num">' + value_html + '<br><span style="font-size:10px;">' + trend_html + '</span></td>'

striking_rows = []
for s in striking:
    striking_rows.append(
        '<tr><td>' + (esc(s["top_keyword"]) or "—") + '</td>'
        '<td>' + addr_link(s["address"], url=s["url"]) + '</td>'
        + _td_with_trend(fmt(s["position"]), trend_arrow(s["d_pos"], inverse=True))
        + _td_with_trend(fmt(s["impressions"]), trend_arrow(s["d_imp"]))
        + _td_with_trend(fmt(s["clicks"]), trend_arrow(s["d_clicks"]))
        + _td_with_trend(fmt_pct(s["ctr"]), trend_arrow(s["d_ctr"]))
        + '</tr>')
striking_tbody = ''.join(striking_rows)
striking_section = ('<table class="compact"><thead><tr>'
    '<th>Top Keyword</th><th>URL</th>'
    '<th class="num">Position</th><th class="num">Impressions</th>'
    '<th class="num">Clicks</th><th class="num">CTR</th>'
    '</tr></thead><tbody>' + striking_tbody + '</tbody></table>') if striking_rows else '<div class="empty-state">No striking-distance opportunities found.</div>'

# 2a Site health
status_rows = []
for s, n in status_mix.most_common():
    status_rows.append(
        '<tr><td><span class="badge" style="background:' + STATUS_COLORS.get(s, "#94a3b8") + '">' + esc(s) + '</span></td>'
        '<td class="num">' + fmt(n) + '</td>'
        '<td class="num">' + fmt_pct(n/len(rows)) + '</td></tr>')
status_tbody = ''.join(status_rows)

broken_rows = []
for r in broken_with_signal:
    broken_rows.append(
        '<tr><td>' + addr_link(r["address"], url=r["full_url"]) + '</td>'
        '<td class="num">' + esc(str(r["status"])) + '</td>'
        '<td class="num">' + fmt(r["gsc_impressions"]) + '</td>'
        '<td class="num">' + fmt(r["ref_domains"]) + '</td></tr>')
broken_section = ''
if broken_rows:
    broken_section = ('<h4 style="margin:20px 0 8px; font-size:14px;">Top broken URLs by signal</h4>'
        '<table class="compact"><thead><tr>'
        '<th>Address</th><th class="num">Status</th><th class="num">GSC Imp</th><th class="num">Backlinks</th>'
        '</tr></thead><tbody>' + ''.join(broken_rows) + '</tbody></table>')

# 2b Indexability breakdown
indexability_rows = []
for label in ["Indexable", "Non-indexable", "Broken/Redirect"]:
    n = indexability_counts.get(label, 0)
    pct_s = fmt_pct(n / len(rows)) if rows else "—"
    color = {"Indexable": "#16a34a", "Non-indexable": "#94a3b8", "Broken/Redirect": "#dc2626"}[label]
    indexability_rows.append(
        '<tr><td><span class="badge" style="background:' + color + '">' + esc(label) + '</span></td>'
        '<td class="num">' + fmt(n) + '</td>'
        '<td class="num">' + pct_s + '</td></tr>')
indexability_tbody = ''.join(indexability_rows)

# 2b Missing-element KPIs
missing_kpi_strip = ''.join([
    render_kpi("Missing title", fmt(missing_title), f"of {len(live_rows)} live"),
    render_kpi("Missing meta", fmt(missing_meta), f"of {len(live_rows)} live"),
    render_kpi("Missing H1", fmt(missing_h1), f"of {len(live_rows)} live"),
    render_kpi("Thin (<618w)", fmt(thin_count), f"of {len(live_rows)} live"),
])

# 2b Crawl depth distribution
DEPTH_ORDER = ["0 (Home)", "1 (Top-level)", "2", "3", "4+"]
crawl_depth_rows = []
total_with_depth = sum(crawl_depth_buckets.values())
for label in DEPTH_ORDER:
    n = crawl_depth_buckets.get(label, 0)
    pct_s = fmt_pct(n / total_with_depth) if total_with_depth else "—"
    color = {"0 (Home)": "#7c3aed", "1 (Top-level)": "#0ea5e9",
             "2": "#16a34a", "3": "#eab308", "4+": "#dc2626"}.get(label, "#94a3b8")
    crawl_depth_rows.append(
        '<tr><td><span class="badge" style="background:' + color + '">' + esc(label) + '</span></td>'
        '<td class="num">' + fmt(n) + '</td>'
        '<td class="num">' + pct_s + '</td></tr>')
crawl_depth_tbody = ''.join(crawl_depth_rows)

# 2b Redirects to home
r2h_rows = []
for r in redirects_to_home:
    r2h_rows.append(
        '<tr><td>' + addr_link(r["address"], url=r.get("url")) + '</td>'
        '<td class="num">' + fmt(r["gsc_impressions"]) + '</td>'
        '<td class="num">' + fmt(r["ref_domains"]) + '</td></tr>')
r2h_section = ''
if r2h_rows:
    r2h_section = ('<h4>Redirects pointing to homepage — relevance leak</h4>'
        '<div class="subtitle" style="margin-bottom:8px;">URLs 301-ing to / instead of a topical match. Backlinks and impressions lose relevance signal.</div>'
        '<table class="compact"><thead><tr><th>Source URL</th><th class="num">GSC Impressions</th><th class="num">Backlinks</th></tr></thead>'
        '<tbody>' + ''.join(r2h_rows) + '</tbody></table>')

# Technical summary removed per request — "Technical work this cycle" panel is no longer rendered.
tech_summary_html = ''

# Deferred technical
deferred_tech_section = ''
if deferred_tech:
    rows_html = []
    for r in deferred_tech:
        rows_html.append(
            '<tr><td><span class="pill">' + esc(r.get("action_type", "")) + '</span></td>'
            '<td>' + addr_link(r.get("page_address", ""), url=r.get("url")) + '</td>'
            '<td>' + esc((r.get("specific_next_step") or "")[:200]) + '</td></tr>')
    deferred_tech_section = ('<section class="panel"><h3>Deferred technical work (' + str(len(deferred_tech)) + ')</h3>'
        '<div class="subtitle">Technical recommendations the strategist deprioritized for next quarter.</div>'
        '<table class="compact"><thead><tr><th>Action</th><th>Page</th><th>Next Step</th></tr></thead><tbody>'
        + ''.join(rows_html) + '</tbody></table></section>')

# 3 Content overview KPIs
content_kpi_strip = ''.join([
    render_kpi("Live pages w/ content", fmt(len(all_wcs))),
    render_kpi("Average word count", fmt(avg_wc)),
    render_kpi("Median word count", fmt(median_wc)),
    render_kpi("Thin pages (<618w)", fmt(thin_count), f"{fmt_pct(thin_count/len(live_rows)) if live_rows else '—'} of live"),
])

# 3 Word count distribution
WC_LABELS_ORDER = [l for _, _, l in WC_BUCKETS]
wc_dist_rows = []
total_wc_pages = sum(wc_buckets.values())
for label in WC_LABELS_ORDER:
    n = wc_buckets.get(label, 0)
    pct_s = fmt_pct(n / total_wc_pages) if total_wc_pages else "—"
    color = "#dc2626" if "thin" in label or "<300" in label else \
            "#eab308" if "618" in label else \
            "#16a34a" if "1,500-3,000" in label else \
            "#7c3aed" if "3,000+" in label else "#94a3b8"
    wc_dist_rows.append(
        '<tr><td><span class="badge" style="background:' + color + '">' + esc(label) + '</span></td>'
        '<td class="num">' + fmt(n) + '</td>'
        '<td class="num">' + pct_s + '</td></tr>')
wc_dist_tbody = ''.join(wc_dist_rows)

# 3 Avg word count by page type
wc_by_pt_rows = []
for pt, d in sorted(wc_by_pt.items(), key=lambda kv: -kv[1]["avg"]):
    wc_by_pt_rows.append(
        '<tr><td><span class="badge" style="background:' + PT_COLORS.get(pt, "#94a3b8") + '">' + esc(pt) + '</span></td>'
        '<td class="num">' + fmt(d["count"]) + '</td>'
        '<td class="num">' + fmt(d["avg"]) + '</td>'
        '<td class="num">' + fmt(d["thin"]) + '</td></tr>')
wc_by_pt_tbody = ''.join(wc_by_pt_rows)

# 3 Content work summary — counts only, no hours
content_action_rows = []
for action, count in content_action_counts.most_common():
    if not action: continue
    content_action_rows.append(
        '<tr><td><span class="pill">' + esc(action) + '</span></td>'
        '<td class="num">' + fmt(count) + '</td></tr>')
content_summary_html = (
    '<div class="kpi-grid" style="grid-template-columns: repeat(2, 1fr);">'
    + render_kpi("Approved content actions", fmt(len(content_recs)))
    + render_kpi("Distinct action types", fmt(len(content_action_counts)))
    + '</div>'
    + '<table class="compact" style="margin-top:18px;">'
      '<thead><tr><th>Action Type</th><th class="num">Pages</th></tr></thead>'
      '<tbody>' + ''.join(content_action_rows) + '</tbody></table>'
    + '<div class="callout" style="margin-top:14px;">Full detail for each content action lives in the audit spreadsheet\'s <strong>Recommendations</strong> tab — the rewrite outline, current state, target state, and rationale.</div>'
) if content_recs else '<div class="empty-state">No approved content recommendations.</div>'

# Deferred content
deferred_content_section = ''
if deferred_content:
    rows_html = []
    for r in deferred_content:
        rows_html.append(
            '<tr><td><span class="pill">' + esc(r.get("action_type", "")) + '</span></td>'
            '<td>' + addr_link(r.get("page_address", ""), url=r.get("url")) + '</td>'
            '<td>' + esc((r.get("specific_next_step") or "")[:200]) + '</td></tr>')
    deferred_content_section = ('<section class="panel"><h3>Deferred content work (' + str(len(deferred_content)) + ')</h3>'
        '<table class="compact"><thead><tr><th>Action</th><th>Page</th><th>Next Step</th></tr></thead><tbody>'
        + ''.join(rows_html) + '</tbody></table></section>')

# Target pages (approved/edited from spreadsheet)
target_rows = []
for t in target_pages:
    target_rows.append(
        '<tr>'
        '<td class="num">' + str(t.get("rank") or "") + '</td>'
        '<td>' + addr_link(t.get("page_address") or "", url=t.get("url")) + '</td>'
        '<td><span class="badge" style="background:' + PT_COLORS.get(t.get("page_type") or "", "#94a3b8") + '">' + esc(t.get("page_type") or "") + '</span></td>'
        '<td>' + esc(t.get("target_keyword") or "—") + '</td>'
        '<td class="num">' + fmt(t.get("kw_volume")) + '</td>'
        '<td class="num">' + fmt(t.get("current_position")) + '</td>'
        '<td class="num">' + fmt(t.get("ref_domains")) + '</td>'
        '<td class="num"><strong>' + fmt(t.get("score")) + '</strong></td>'
        '</tr>')
target_tbody = ''.join(target_rows) or ''

# Unmarked target pages callout
unmarked_target_callout = ''
if unmarked_targets:
    unmarked_target_callout = (
        '<div class="callout callout-warn"><strong>' + str(len(unmarked_targets)) + ' target pages still unmarked.</strong> '
        'Open the Target Pages tab in the spreadsheet and set Approval = Approved / Edit / Rejected / Deferred for each row before this section is considered final.</div>'
    )

# ============ 6-MONTH IMPLEMENTATION PLAN ============
# Sprint 1 = Technical (all in Month 1, overflow to Month 2 if > 30 items)
# Sprint 2 = Content — 8,000 words/month capacity, avg 2,000 words/piece = 4 pieces/month
#   Content title/meta tasks are separate "copywriting" budget, ~10/month
# Sprint 3 = Links — 6 links/month × 6 months = 36 link slots distributed across approved target pages

MONTHS = 6
CONTENT_WORD_BUDGET = 8000
AVG_PIECE_WORDS = 2000
LINKS_PER_MONTH = 6
COPYWRITING_PER_MONTH = 10  # rewrite title/meta pace

def words_for(rec):
    """Word-count estimate for content scheduling. 0 = copywriting only (no body words)."""
    action = (rec.get("action_type") or "").strip()
    if action == "Rewrite":          return 2000
    if action == "Expand content":   return 1500
    if action == "Refresh content":  return 1000
    if action == "Update onpage":    return 500   # H1 add or minor onpage tweak
    if action == "Consolidate":      return 1500
    if action == "Evaluate":         return 0
    if action == "Rewrite title/meta": return 0  # copywriting bucket
    return 0


schedule = {m: {"technical": [], "content": [], "copywriting": [], "links": []}
            for m in range(1, MONTHS + 1)}

# Scheduling rule: Month 1 is reserved for Sprint 1 (Planning/WQA, already in
# progress) + Sprint 2 (Technical fixes only). Content body, copywriting, and
# link-building all start Month 2 — there's not enough time in the first month
# after the WQA finishes to also produce content or run outreach.
CONTENT_START_MONTH = 2
LINKS_START_MONTH = 2

# ---- Sprint 2: Technical ----
PRIO_ORDER_LOCAL = {"P1": 0, "P2": 1, "P3": 2, "": 9}
tech_sorted = sorted(tech_recs,
                     key=lambda r: (PRIO_ORDER_LOCAL.get(r.get("priority"), 9),
                                    r.get("page_address") or ""))
if len(tech_sorted) <= 30:
    schedule[1]["technical"] = tech_sorted
else:
    half = (len(tech_sorted) + 1) // 2
    schedule[1]["technical"] = tech_sorted[:half]
    schedule[2]["technical"] = tech_sorted[half:]

# ---- Sprint 4: Content ----
content_body = []
content_copy = []
for r in content_recs:
    w = words_for(r)
    if w == 0:
        content_copy.append(r)
    else:
        content_body.append((r, w))
content_body.sort(key=lambda x: (PRIO_ORDER_LOCAL.get(x[0].get("priority"), 9), -x[1]))

# Body content — fill months by word budget, starting Month 2
current_month = CONTENT_START_MONTH
used = 0
for rec, w in content_body:
    if used + w > CONTENT_WORD_BUDGET:
        if current_month < MONTHS:
            current_month += 1
            used = 0
        # If at last month, keep stuffing (overflow into the last month visually)
    rec["_scheduled_words"] = w
    schedule[current_month]["content"].append(rec)
    used += w

# Copywriting tasks — distribute across Months 2-6 (skipping Month 1)
copy_sorted = sorted(content_copy,
                     key=lambda r: (PRIO_ORDER_LOCAL.get(r.get("priority"), 9),
                                    r.get("page_address") or ""))
copy_iter = iter(copy_sorted)
for m in range(CONTENT_START_MONTH, MONTHS + 1):
    for _ in range(COPYWRITING_PER_MONTH):
        try:
            schedule[m]["copywriting"].append(next(copy_iter))
        except StopIteration:
            break
# Anything leftover goes to the last month
for rec in copy_iter:
    schedule[MONTHS]["copywriting"].append(rec)

# ---- Sprint 5: Link building ----
# Link slots distributed across Months 2-6 (skipping Month 1).
approved_tp = target_pages  # already filtered to Approved + Edit
link_months = MONTHS - LINKS_START_MONTH + 1  # 5 months
total_link_slots = LINKS_PER_MONTH * link_months
if approved_tp:
    n_pages = len(approved_tp)
    base = total_link_slots // n_pages
    extras = total_link_slots % n_pages
    # Higher-ranked target pages get the extras
    page_link_count = []
    for i, tp in enumerate(approved_tp):
        page_link_count.append((tp, base + (1 if i < extras else 0)))
    # Build flat slot list, interleaving higher-priority pages first
    flat_slots = []
    max_n = max(n for _, n in page_link_count) if page_link_count else 0
    for round_i in range(max_n):
        for tp, n in page_link_count:
            if round_i < n:
                flat_slots.append(tp)
    # Distribute LINKS_PER_MONTH per month starting Month 2
    for m_idx, m in enumerate(range(LINKS_START_MONTH, MONTHS + 1)):
        start = m_idx * LINKS_PER_MONTH
        schedule[m]["links"] = flat_slots[start:start + LINKS_PER_MONTH]

# Roll-ups for chart
month_stats = []
for m in range(1, MONTHS + 1):
    s = schedule[m]
    body_words = sum(words_for(r) for r in s["content"])
    month_stats.append({
        "month": m,
        "technical": len(s["technical"]),
        "content_pieces": len(s["content"]),
        "content_words": body_words,
        "copywriting": len(s["copywriting"]),
        "links": len(s["links"]),
    })

SPRINT_ORDER = {
    "Sprint 1 (Planning)": 1, "Sprint 1": 1,
    "Sprint 2 (Technical)": 2, "Sprint 2": 2,
    "Sprint 3 (Local)": 3, "Sprint 3": 3,
    "Sprint 4 (Content)": 4, "Sprint 4": 4,
    "Sprint 5 (Links)": 5, "Sprint 5": 5,
    "Backlog": 6, "Done": 7, "TBD": 9,
}

def render_item_accordion(idx, rec):
    """An individual approved item — collapsed by default, expands to the full rec detail."""
    prio = rec.get("priority") or ""
    action = rec.get("action_type") or ""
    category = rec.get("category") or ""
    addr = rec.get("page_address") or ""
    url = rec.get("url") or (ROOT + addr if addr.startswith("/") else None)
    edited = rec.get("edited")
    next_step = rec.get("specific_next_step") or "—"
    overrides = rec.get("aggregator_overrides_applied")
    edit_badge = '<span class="edit-badge">edited</span>' if edited else ''
    cat_pill = ('<span class="cat-pill cat-' + category.lower() + '">' + esc(category) + '</span>') if category else ''
    override_block = ''
    if overrides:
        override_block = '<div class="rec-grid-full" style="background:#dbeafe;"><div class="rec-label">Aggregator overrides applied</div><div class="rec-value">' + esc(' · '.join(overrides)) + '</div></div>'
    return (
        '<details class="accordion item-accordion">'
          '<summary>'
            '<div class="item-summary">'
              '<span class="item-num">#' + str(idx) + '</span>'
              '<span class="badge prio-' + prio + '">' + esc(prio) + '</span>'
              + cat_pill +
              '<span class="pill">' + esc(action) + '</span>'
              + edit_badge +
              '<span class="item-addr">' + addr_link(addr, url=url) + '</span>'
            '</div>'
          '</summary>'
          '<div class="acc-body item-detail">'
            '<div class="rec-step">' + esc(next_step) + '</div>'
            + override_block +
          '</div>'
        '</details>')

# ---- Build the 6-month plan UI ----

MONTH_LABELS = [f"Month {m}" for m in range(1, MONTHS + 1)]

# Sprint summary line
total_tech = len(tech_recs)
total_content_pieces = sum(len(schedule[m]["content"]) for m in range(1, MONTHS + 1))
total_copywriting = sum(len(schedule[m]["copywriting"]) for m in range(1, MONTHS + 1))
total_links = sum(len(schedule[m]["links"]) for m in range(1, MONTHS + 1))

plan_summary_html = (
    '<div class="kpi-grid" style="grid-template-columns: repeat(4, 1fr); margin-bottom: 22px;">'
    + render_kpi("Sprint 2 · Technical", fmt(total_tech), "Month 1, overflow M2")
    + render_kpi("Sprint 4 · Content", fmt(total_content_pieces) + " pieces", f"+ {total_copywriting} copywriting · M2-M6")
    + render_kpi("Sprint 5 · Links", fmt(total_links) + " links", f"{LINKS_PER_MONTH}/mo × {link_months} mo · {len(approved_tp)} pages")
    + render_kpi("Capacity assumed", "8,000w/mo", f"avg {AVG_PIECE_WORDS:,}w per piece")
    + '</div>')


# Render one item accordion per work item inside a month
def render_month_item(rec, badge_text, badge_class="cat-content"):
    """A single line item inside a month accordion — expandable to deep detail."""
    prio = rec.get("priority") or ""
    action = rec.get("action_type") or ""
    addr = rec.get("page_address") or ""
    url = rec.get("url") or (ROOT + addr if addr.startswith("/") else None)
    edited = rec.get("edited")
    next_step = rec.get("specific_next_step") or "—"
    current = rec.get("current_state") or "—"
    target = rec.get("target_state") or "—"
    why = rec.get("why_it_matters") or "—"
    notes = rec.get("approver_notes") or ""
    edit_badge = '<span class="edit-badge">edited</span>' if edited else ''
    notes_block = ('<div class="rec-grid-full"><div class="rec-label">Strategist notes</div><div class="rec-value">' + esc(notes) + '</div></div>') if notes else ''
    return (
        '<details class="accordion item-accordion">'
          '<summary>'
            '<div class="item-summary">'
              '<span class="cat-pill ' + badge_class + '">' + esc(badge_text) + '</span>'
              '<span class="badge prio-' + prio + '">' + esc(prio) + '</span>'
              '<span class="pill">' + esc(action) + '</span>'
              + edit_badge +
              '<span class="item-addr">' + addr_link(addr, url=url) + '</span>'
            '</div>'
          '</summary>'
          '<div class="acc-body item-detail">'
            '<div class="rec-step">' + esc(next_step) + '</div>'
            '<div class="rec-grid">'
              '<div><div class="rec-label">Current state</div><div class="rec-value">' + esc(current) + '</div></div>'
              '<div><div class="rec-label">Target state</div><div class="rec-value">' + esc(target) + '</div></div>'
              '<div><div class="rec-label">Why it matters</div><div class="rec-value">' + esc(why) + '</div></div>'
            '</div>'
            + notes_block +
          '</div>'
        '</details>')


def render_link_slot(idx, tp):
    """Compact accordion for a link-building slot pointing at a target page."""
    addr = tp.get("page_address") or ""
    url = tp.get("url")
    return (
        '<details class="accordion item-accordion">'
          '<summary>'
            '<div class="item-summary">'
              '<span class="cat-pill cat-link">LINK</span>'
              '<span class="pill">Acquire 1 link</span>'
              '<span class="item-addr">' + addr_link(addr, url=url) + '</span>'
              '<span class="link-meta">→ ' + esc(tp.get("target_keyword") or "—") + '</span>'
            '</div>'
          '</summary>'
          '<div class="acc-body item-detail">'
            '<div class="rec-step">Run a link-building campaign cycle for this target page. Tactics: digital PR pitch, niche-edit / link insertion outreach, podcast appearance, guest post — pick whichever lane is realistic for the keyword and vertical.</div>'
            '<div class="rec-grid">'
              '<div><div class="rec-label">Target keyword</div><div class="rec-value">' + esc(tp.get("target_keyword") or "—") + ' (vol ' + fmt(tp.get("kw_volume")) + ')</div></div>'
              '<div><div class="rec-label">Current position</div><div class="rec-value">' + fmt(tp.get("current_position")) + '</div></div>'
              '<div><div class="rec-label">Current ref domains</div><div class="rec-value">' + fmt(tp.get("ref_domains")) + '</div></div>'
              '<div><div class="rec-label">Score</div><div class="rec-value">' + fmt(tp.get("score")) + '</div></div>'
            '</div>'
          '</div>'
        '</details>')


# Build monthly accordions
month_accordions = []
for m in range(1, MONTHS + 1):
    s = schedule[m]
    stats = month_stats[m - 1]
    parts = []

    if s["technical"]:
        parts.append('<h5 class="month-sub">Technical (' + str(len(s["technical"])) + ')</h5>')
        for r in s["technical"]:
            parts.append(render_month_item(r, "TECHNICAL", badge_class="cat-technical"))

    if s["content"]:
        parts.append('<h5 class="month-sub">Content pieces (' + str(stats["content_words"]) + 'w · ' + str(len(s["content"])) + ' pieces)</h5>')
        for r in s["content"]:
            parts.append(render_month_item(r, "CONTENT", badge_class="cat-content"))

    if s["copywriting"]:
        parts.append('<h5 class="month-sub">Copywriting tasks — titles &amp; metas (' + str(len(s["copywriting"])) + ')</h5>')
        for r in s["copywriting"]:
            parts.append(render_month_item(r, "COPYWRITING", badge_class="cat-content"))

    if s["links"]:
        parts.append('<h5 class="month-sub">Link building (' + str(len(s["links"])) + ' links)</h5>')
        for i, tp in enumerate(s["links"], 1):
            parts.append(render_link_slot(i, tp))

    if not parts:
        parts.append('<div class="empty-state">No work scheduled for this month.</div>')

    meta_parts = []
    if stats["technical"]: meta_parts.append(f"{stats['technical']} technical")
    if stats["content_pieces"]: meta_parts.append(f"{stats['content_pieces']} content ({stats['content_words']:,}w)")
    if stats["copywriting"]: meta_parts.append(f"{stats['copywriting']} copywriting")
    if stats["links"]: meta_parts.append(f"{stats['links']} links")
    meta = " · ".join(meta_parts) or "Nothing scheduled"

    month_accordions.append(
        '<details class="accordion sprint-accordion" ' + ('open' if m == 1 else '') + '>'
          '<summary>'
            '<span class="sprint-name">Month ' + str(m) + '</span>'
            '<span class="sprint-meta">' + esc(meta) + '</span>'
          '</summary>'
          '<div class="acc-body">' + ''.join(parts) + '</div>'
        '</details>')

sprint_accordions_html = plan_summary_html + ''.join(month_accordions)


# Chart data for JS
chart_data = {
    "ga4_monthly_organic": ga4_monthly.get("organic", []),
    "gsc_monthly": gsc_monthly,
    "by_page_type_clicks": [{"label": pt, "value": d["clicks"]} for pt, d in clicks_by_pt.items() if d["clicks"] > 0],
    "by_funnel_clicks": [{"label": fn, "value": d["clicks"]} for fn, d in clicks_by_fn.items() if d["clicks"] > 0],
    "pt_colors": PT_COLORS,
    "fn_colors": FN_COLORS,
    "status_mix": [{"label": s, "value": n, "color": STATUS_COLORS.get(s, "#94a3b8")} for s, n in status_mix.most_common()],
}


# ============ Final template ============

# Inline Chart.js so the report renders in sandboxed/offline viewers. A CDN
# <script src> is blocked in many preview contexts (and when the client opens
# the file offline), which leaves every chart blank even though the data is
# embedded. Vendored copy lives in the skill's assets/; fall back to CDN only
# if it's somehow missing.
_chartjs_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "assets", "chart.umd.min.js"
)
try:
    with open(_chartjs_path, encoding="utf-8") as _cf:
        CHARTJS_TAG = "<script>\n" + _cf.read() + "\n</script>"
except OSError:
    CHARTJS_TAG = '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>'

html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{esc(CLIENT_NAME)} — Website Quality Audit</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Figtree:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
{CHARTJS_TAG}
<style>
  :root {{
    --primary: {PRIMARY};
    --primary-soft: {PRIMARY}1A;
    --text: #0a0e1a;
    --muted: #4b5563;
    --border: #e5e7eb;
    --bg: #ffffff;
    --soft: #f7f7f8;
    --ink: #0a0e1a;
    --display: 'Bebas Neue', 'Arial Narrow', sans-serif;
    --body: 'Figtree', -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
  }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: var(--body);
         margin: 0; color: var(--text); background: #ededee; line-height: 1.55; font-size: 14px;
         font-weight: 400; }}
  /* Bebas Neue ALL CAPS for every heading */
  h1, h2, h3, h4, h5, h6,
  .hero h1, .title-slide h2,
  section.panel h3, section.panel h4 {{
    font-family: var(--display);
    text-transform: uppercase;
    letter-spacing: 0.02em;
    font-weight: 400;  /* Bebas Neue only has one weight */
    line-height: 1.05;
  }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 32px 24px 80px; }}

  header.hero {{ background: var(--ink); color: white;
                padding: 40px 44px; margin-bottom: 28px;
                border-radius: 14px;
                border: 1px solid var(--ink);
                position: relative; overflow: hidden; }}
  header.hero::after {{ content: ""; position: absolute; right: -40px; top: -40px;
                        width: 200px; height: 200px; border-radius: 50%;
                        background: var(--primary); opacity: 0.15; }}
  header.hero h1 {{ margin: 0 0 8px; font-size: 56px; line-height: 0.95; font-family: var(--display);
                    text-transform: uppercase; font-weight: 400; letter-spacing: 0.01em; }}
  header.hero .meta {{ font-size: 13px; font-family: var(--body); opacity: 0.78;
                       text-transform: uppercase; letter-spacing: 0.08em; font-weight: 500; }}

  .title-slide {{ background: var(--ink); color: white; border-radius: 14px;
                  padding: 48px 44px; margin: 44px 0 20px; text-align: left;
                  position: relative; overflow: hidden;
                  border-left: 8px solid var(--primary); }}
  .title-slide .section-num {{ font-size: 12px; font-family: var(--body); font-weight: 700;
                                letter-spacing: 0.25em; text-transform: uppercase;
                                color: var(--primary); margin-bottom: 14px; }}
  .title-slide h2 {{ margin: 0 0 14px; font-size: 64px; line-height: 0.92;
                     font-family: var(--display); text-transform: uppercase;
                     font-weight: 400; letter-spacing: 0.01em; }}
  .title-slide .desc {{ font-size: 16px; font-family: var(--body); line-height: 1.55;
                        opacity: 0.85; max-width: 800px; font-weight: 400; }}

  section.panel {{ background: var(--bg); border: 1px solid var(--border); border-radius: 12px;
             padding: 28px 30px; margin-bottom: 18px; }}
  section.panel h3 {{ margin: 0 0 6px; font-size: 30px; font-family: var(--display);
                       text-transform: uppercase; font-weight: 400; letter-spacing: 0.01em;
                       line-height: 1.0; color: var(--ink); }}
  section.panel h4 {{ margin: 18px 0 8px; font-size: 20px; font-family: var(--display);
                       text-transform: uppercase; font-weight: 400; letter-spacing: 0.02em;
                       color: var(--ink); }}
  section.panel .subtitle {{ color: var(--muted); font-size: 13px; margin-bottom: 20px;
                              font-family: var(--body); line-height: 1.55; }}

  .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; }}
  .kpi {{ padding: 18px; background: var(--soft); border-radius: 10px;
          border: 1px solid var(--border); }}
  .kpi .label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em;
                color: var(--muted); font-weight: 700; font-family: var(--body); }}
  .kpi .value {{ font-size: 38px; font-family: var(--display); text-transform: uppercase;
                  letter-spacing: 0.01em; margin-top: 6px; color: var(--ink);
                  line-height: 1.0; font-weight: 400; }}
  .kpi .sub {{ font-size: 11px; color: var(--muted); margin-top: 4px; font-family: var(--body); }}
  .trend {{ display: inline-block; font-size: 11px; font-weight: 700; padding: 3px 10px;
            border-radius: 999px; margin-top: 8px; font-family: var(--body);
            letter-spacing: 0.04em; }}
  .trend-up-good {{ background: #dcfce7; color: #16a34a; }}
  .trend-up-bad {{ background: #fee2e2; color: #dc2626; }}
  .trend-down-good {{ background: #dcfce7; color: #16a34a; }}
  .trend-down-bad {{ background: #fee2e2; color: #dc2626; }}
  .trend.neutral {{ background: #f1f5f9; color: #64748b; }}

  .row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  .chart-wrap {{ height: 300px; position: relative; }}
  .chart-wrap-wide {{ height: 380px; position: relative; }}

  table {{ width: 100%; border-collapse: collapse; font-size: 13px; font-family: var(--body); }}
  table th {{ text-align: left; padding: 10px 12px; background: var(--ink); color: white;
              border-bottom: none; font-weight: 700; font-family: var(--body);
              font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; }}
  table td {{ padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  table tr:last-child td {{ border-bottom: none; }}
  table.compact td {{ padding: 7px 10px; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  th.num {{ text-align: right; }}
  a {{ color: var(--primary); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}

  .badge {{ display: inline-block; padding: 3px 10px; border-radius: 4px; font-size: 11px;
            font-weight: 700; color: white; font-family: var(--body);
            letter-spacing: 0.05em; text-transform: uppercase; }}
  .badge.prio-P1 {{ background: #dc2626; }}
  .badge.prio-P2 {{ background: #ea580c; }}
  .badge.prio-P3 {{ background: #ca8a04; }}
  .pill {{ display: inline-block; padding: 3px 10px; border-radius: 4px; font-size: 11px;
           font-weight: 700; background: var(--primary-soft); color: var(--primary);
           font-family: var(--body); letter-spacing: 0.04em; text-transform: uppercase; }}
  .edit-badge {{ display: inline-block; padding: 2px 7px; border-radius: 4px;
                background: #fef3c7; color: #92400e; font-size: 10px; font-weight: 700;
                text-transform: uppercase; letter-spacing: 0.05em; }}

  details.accordion {{ border: 1px solid var(--border); border-radius: 8px; margin: 8px 0; background: white; }}
  details.accordion[open] {{ background: var(--soft); }}
  details.accordion > summary {{ cursor: pointer; padding: 14px 18px; font-weight: 600;
                                 list-style: none; display: flex; align-items: center;
                                 justify-content: space-between; }}
  details.accordion > summary::-webkit-details-marker {{ display: none; }}
  details.accordion > summary::after {{ content: "▾"; color: var(--muted); margin-left: 10px; transition: transform 0.2s; }}
  details.accordion[open] > summary::after {{ transform: rotate(180deg); }}
  details.accordion .acc-body {{ padding: 0 18px 16px; }}

  .rec-card {{ background: white; border: 1px solid var(--border); border-radius: 10px;
               margin: 12px 0; overflow: hidden; }}
  .rec-head {{ padding: 14px 18px; background: var(--soft); border-bottom: 1px solid var(--border);
                display: flex; align-items: center; justify-content: space-between; gap: 12px;
                flex-wrap: wrap; }}
  .rec-meta {{ display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }}
  .rec-num {{ background: var(--primary); color: white; font-size: 11px; font-weight: 700;
              padding: 3px 8px; border-radius: 4px; }}
  .rec-addr {{ font-weight: 600; font-size: 14px; }}
  .rec-body {{ padding: 18px; }}
  .rec-step {{ font-size: 14px; color: var(--text); line-height: 1.6; margin-bottom: 14px;
               padding-bottom: 14px; border-bottom: 1px dashed var(--border); }}
  .rec-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
  .rec-grid > div {{ background: var(--soft); padding: 10px 12px; border-radius: 6px; }}
  .rec-label {{ font-size: 10px; text-transform: uppercase; letter-spacing: 0.06em;
                color: var(--muted); font-weight: 700; margin-bottom: 3px; }}
  .rec-value {{ font-size: 13px; color: var(--text); }}

  .callout {{ background: var(--primary-soft); border-left: 4px solid var(--primary);
              padding: 12px 18px; border-radius: 6px; font-size: 13px; color: var(--text); margin: 14px 0; }}
  .callout-warn {{ background: #fef3c7; border-left-color: #ca8a04; color: #78350f; }}
  .empty-state {{ padding: 28px; text-align: center; color: var(--muted); font-style: italic;
                  background: var(--soft); border-radius: 8px; }}

  /* Sprint accordion — keeps the open-by-default summary */
  details.sprint-accordion > summary {{
    background: var(--ink); color: white;
    padding: 16px 22px; border-radius: 8px; cursor: pointer;
    display: flex; align-items: center; justify-content: space-between;
    list-style: none; margin: 0;
    border-left: 6px solid var(--primary);
  }}
  details.sprint-accordion[open] > summary {{ border-radius: 8px 8px 0 0; }}
  details.sprint-accordion > summary::-webkit-details-marker {{ display: none; }}
  details.sprint-accordion > summary::after {{ content: "▾"; color: white; opacity: 0.6; }}
  details.sprint-accordion .sprint-name {{ font-size: 22px; font-family: var(--display);
                                            text-transform: uppercase; letter-spacing: 0.02em;
                                            font-weight: 400; }}
  details.sprint-accordion .sprint-meta {{ font-size: 11px; font-family: var(--body);
                                            text-transform: uppercase; letter-spacing: 0.08em;
                                            opacity: 0.7; margin-left: auto; margin-right: 14px;
                                            font-weight: 700; }}
  details.sprint-accordion > .acc-body {{ padding: 12px 14px; background: var(--soft);
                                           border: 1px solid var(--border); border-top: none;
                                           border-radius: 0 0 8px 8px; }}

  /* Individual item accordion */
  details.item-accordion {{ background: white; border: 1px solid var(--border); border-radius: 6px;
                            margin: 6px 0; }}
  details.item-accordion[open] {{ box-shadow: 0 2px 6px rgba(0,0,0,0.06); }}
  details.item-accordion > summary {{ padding: 10px 14px; cursor: pointer; list-style: none; }}
  details.item-accordion > summary::-webkit-details-marker {{ display: none; }}
  details.item-accordion .item-summary {{ display: flex; align-items: center; gap: 10px;
                                            flex-wrap: wrap; }}
  .item-num {{ font-size: 12px; font-weight: 700; color: var(--muted); min-width: 32px; }}
  .item-addr {{ font-weight: 600; flex: 1; min-width: 0; }}
  .cat-pill {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 10px;
               font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; }}
  .cat-pill.cat-technical {{ background: #DBEAFE; color: #1d4ed8; }}
  .cat-pill.cat-content   {{ background: #FCE7F3; color: #be185d; }}
  .item-detail {{ padding: 12px 16px 16px; border-top: 1px solid var(--border); background: var(--soft); }}
  .item-detail .rec-step {{ font-size: 14px; line-height: 1.6; margin-bottom: 14px;
                              padding-bottom: 14px; border-bottom: 1px dashed var(--border); color: var(--text); }}
  .item-detail .rec-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
  .item-detail .rec-grid > div {{ background: white; padding: 10px 12px; border-radius: 6px;
                                    border: 1px solid var(--border); }}
  .item-detail .rec-label {{ font-size: 10px; text-transform: uppercase; letter-spacing: 0.06em;
                              color: var(--muted); font-weight: 700; margin-bottom: 3px; }}
  .item-detail .rec-value {{ font-size: 13px; color: var(--text); }}
  .item-detail .rec-grid-full {{ margin-top: 10px; background: #fef3c7; padding: 10px 12px;
                                  border-radius: 6px; }}

  footer {{ margin-top: 32px; text-align: center; color: var(--muted); font-size: 12px; }}
</style>
</head>
<body>
<div class="container">

<header class="hero">
  <h1>{esc(CLIENT_NAME)} — Website Quality Audit</h1>
  <div class="meta">Trailing 90 days · {esc(period_label)} · vs prior 90 days ({esc(prior_period_label)}) · Generated {date.today().isoformat()}</div>
</header>

<div class="title-slide">
  <div class="section-num">Section 1 of 4</div>
  <h2>Performance Metrics</h2>
  <div class="desc">How organic traffic, keywords, and conversions are performing — sourced from Google Analytics 4 (organic medium) and Google Search Console. All trend percentages compare the trailing 90 days to the prior 90-day window.</div>
</div>

<section class="panel">
  <h3>At a glance</h3>
  <div class="subtitle">8 headline numbers · trailing 90 days · trend vs prior 90 days</div>
  <div class="kpi-grid">{kpi_strip}</div>
</section>

<section class="panel">
  <h3>Organic performance over time</h3>
  <div class="subtitle">Monthly organic sessions (GA4, organic medium only) + GSC clicks + GSC impressions · GA4 history: {len(ga4_monthly.get('organic', []))} months · GSC limited to ~16 months by the Search Console API</div>
  <div class="chart-wrap-wide"><canvas id="chart-timeseries"></canvas></div>
</section>

<section class="panel">
  <h3>Performance by page type</h3>
  <div class="subtitle">GSC clicks split across the page-type classifier · expand each type to see every page in that category</div>
  <div class="row">
    <div><div class="chart-wrap"><canvas id="chart-page-type"></canvas></div></div>
    <div>
      <table class="compact">
        <thead><tr><th>Type</th><th class="num">URLs</th><th class="num">Clicks</th><th class="num">Share</th></tr></thead>
        <tbody>{pt_summary_tbody}</tbody>
      </table>
    </div>
  </div>
  <div style="margin-top:18px;">{pt_accordions_html}</div>
</section>

<section class="panel">
  <h3>Performance by funnel stage</h3>
  <div class="subtitle">GSC clicks split across Brand / Bottom / Mid / Top / Utility funnel stages</div>
  <div class="row">
    <div><div class="chart-wrap"><canvas id="chart-funnel"></canvas></div></div>
    <div>
      <table class="compact">
        <thead><tr><th>Funnel</th><th class="num">URLs</th><th class="num">Clicks</th><th class="num">Share</th></tr></thead>
        <tbody>{fn_summary_tbody}</tbody>
      </table>
    </div>
  </div>
</section>

<section class="panel">
  <h3>Top keywords</h3>
  <div class="subtitle">Top 15 ranking keywords by estimated traffic · current snapshot from Ahrefs (US) · keyword-level period comparison requires Ahrefs keyword-history with date_compared (not yet wired in this environment)</div>
  <table class="compact">
    <thead><tr><th>Keyword</th><th>Intent</th><th class="num">Position</th><th class="num">Volume</th></tr></thead>
    <tbody>{kw_tbody}</tbody>
  </table>
</section>

<section class="panel">
  <h3>Striking distance opportunities</h3>
  <div class="subtitle">URLs ranking in positions 4–15 with ≥ 200 impressions in the trailing 90 days · trend arrows compare to the prior 90 days · sorted by impressions</div>
  {striking_section}
</section>

<div class="title-slide">
  <div class="section-num">Section 2 of 4</div>
  <h2>Technical Improvements</h2>
  <div class="desc">Site health, page-speed, and technical-SEO recommendations. Content and link work live in their own sections.</div>
</div>

<section class="panel">
  <h3>Site health snapshot</h3>
  <div class="subtitle">Status mix across all known URLs · top broken URLs sorted by signal (impressions + backlinks)</div>
  <div class="row">
    <div><div class="chart-wrap"><canvas id="chart-status"></canvas></div></div>
    <div>
      <table class="compact">
        <thead><tr><th>Status</th><th class="num">Count</th><th class="num">% of total</th></tr></thead>
        <tbody>{status_tbody}</tbody>
      </table>
      <div class="callout callout-warn">{fmt(thin_count)} live pages have fewer than 618 words (below the depth benchmark).</div>
    </div>
  </div>
  {broken_section}
</section>

<section class="panel">
  <h3>Indexability + critical elements</h3>
  <div class="subtitle">Whether Google can index each URL · how many live pages are missing fundamental on-page elements</div>
  <div class="kpi-grid" style="grid-template-columns: repeat(4, 1fr); margin-bottom: 18px;">{missing_kpi_strip}</div>
  <div class="row">
    <div>
      <h4>Indexability breakdown</h4>
      <table class="compact">
        <thead><tr><th>Status</th><th class="num">URLs</th><th class="num">Share</th></tr></thead>
        <tbody>{indexability_tbody}</tbody>
      </table>
    </div>
    <div>
      <h4>Crawl depth distribution</h4>
      <div class="subtitle" style="margin-bottom:8px;">How many clicks from the homepage. Deep pages (depth 4+) often signal weak internal linking.</div>
      <table class="compact">
        <thead><tr><th>Depth</th><th class="num">Pages</th><th class="num">Share</th></tr></thead>
        <tbody>{crawl_depth_tbody}</tbody>
      </table>
    </div>
  </div>
</section>

<section class="panel">
  <h3>Redirects analysis</h3>
  <div class="subtitle">Where 301/302 traffic is being routed — and whether equity is being preserved or thrown to the homepage</div>
  {r2h_section if r2h_section else '<div class="empty-state">No problematic redirect-to-home cases detected.</div>'}
</section>


{deferred_tech_section}

<div class="title-slide">
  <div class="section-num">Section 3 of 4</div>
  <h2>Content</h2>
  <div class="desc">Page-level content recommendations — rewrites, depth expansions, content-hub improvements, and E-E-A-T upgrades.</div>
</div>

<section class="panel">
  <h3>Content depth overview</h3>
  <div class="subtitle">How much content lives behind each indexable URL — the depth benchmark for ranking pages is ~618 words</div>
  <div class="kpi-grid" style="grid-template-columns: repeat(4, 1fr); margin-bottom: 18px;">{content_kpi_strip}</div>
  <div class="row">
    <div>
      <h4>Word count distribution</h4>
      <div class="subtitle" style="margin-bottom:8px;">Live pages bucketed by word count.</div>
      <table class="compact">
        <thead><tr><th>Bucket</th><th class="num">Pages</th><th class="num">Share</th></tr></thead>
        <tbody>{wc_dist_tbody}</tbody>
      </table>
    </div>
    <div>
      <h4>Depth by page type</h4>
      <div class="subtitle" style="margin-bottom:8px;">Avg words per page · count of thin pages within each type.</div>
      <table class="compact">
        <thead><tr><th>Page Type</th><th class="num">Pages</th><th class="num">Avg Words</th><th class="num">Thin</th></tr></thead>
        <tbody>{wc_by_pt_tbody}</tbody>
      </table>
    </div>
  </div>
</section>

<section class="panel">
  <h3>Content work this cycle</h3>
  <div class="subtitle">Headline counts only — full per-page detail (rewrite outline, current/target state, rationale) lives in the audit spreadsheet's Recommendations tab.</div>
  {content_summary_html}
</section>

{deferred_content_section}

<div class="title-slide">
  <div class="section-num">Section 4 of 4</div>
  <h2>Links</h2>
  <div class="desc">Target pages for link-building — P1 + P2 pages from the audit, with current backlink profile and (pending) competitive gap analysis.</div>
</div>

<section class="panel">
  <h3>Target pages — link-building roster (top 20)</h3>
  <div class="subtitle">Pages ranked by link-building value: keyword volume × current-rank position multiplier × business value (Money Page 3.0× · Authority 1.5× · Content Hub 1.3× · Blog 1.0×). Homepage, utility pages, 404s, and redirects excluded.</div>
  <table class="compact">
    <thead><tr>
      <th class="num">#</th><th>Page</th><th>Page Type</th><th>Target Keyword</th>
      <th class="num">KW Vol</th><th class="num">Current Pos</th>
      <th class="num">Ref Domains</th><th class="num">Score</th>
    </tr></thead>
    <tbody>{target_tbody}</tbody>
  </table>
  {unmarked_target_callout}
  <div class="callout callout-warn" style="margin-top:18px;">
    <strong>Backlink gap analysis — pending.</strong> For each target page above, the next iteration will pull the top 5 SERP results for the target keyword, compare their referring domains to the target page's, and surface the gap (e.g., "competitors have X linking domains you don't"). Requires Ahrefs SERP-overview + referring-domains calls per page.
  </div>
</section>

<footer>
  Generated from <code>{esc(os.path.basename(XLSX))}</code> · trailing 90 days = {esc(period_label)} · WQA v3
</footer>

</div>

<script>
const CD = {json.dumps(chart_data)};
const PRIMARY = "{PRIMARY}";

// Time series
(function() {{
  const ga4 = CD.ga4_monthly_organic;
  const gsc = CD.gsc_monthly;
  const months = Array.from(new Set([...ga4.map(r => r.month), ...gsc.map(r => r.month)])).sort();
  const labels = months.map(m => {{
    const [y, mm] = m.split('-');
    const monthName = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][parseInt(mm)-1];
    return monthName + " '" + y.slice(2);
  }});
  const ga4Map = Object.fromEntries(ga4.map(r => [r.month, r.sessions]));
  const gscClicksMap = Object.fromEntries(gsc.map(r => [r.month, r.clicks]));
  const gscImpMap = Object.fromEntries(gsc.map(r => [r.month, r.impressions]));

  new Chart(document.getElementById('chart-timeseries'), {{
    type: 'line',
    data: {{
      labels: labels,
      datasets: [
        {{ label: 'Organic sessions (GA4)', data: months.map(m => ga4Map[m] ?? null),
           borderColor: PRIMARY, backgroundColor: PRIMARY + '22', yAxisID: 'y',
           tension: 0.3, fill: true, spanGaps: true }},
        {{ label: 'GSC clicks', data: months.map(m => gscClicksMap[m] ?? null),
           borderColor: '#16a34a', backgroundColor: 'transparent', yAxisID: 'y',
           tension: 0.3, spanGaps: true }},
        {{ label: 'GSC impressions', data: months.map(m => gscImpMap[m] ?? null),
           borderColor: '#ea580c', backgroundColor: 'transparent', yAxisID: 'yImp',
           tension: 0.3, borderDash: [4, 4], spanGaps: true }},
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{ legend: {{ position: 'top', labels: {{ font: {{ size: 11 }} }} }} }},
      scales: {{
        y: {{ position: 'left', beginAtZero: true,
              title: {{ display: true, text: 'Sessions / Clicks', font: {{ size: 11 }} }},
              ticks: {{ font: {{ size: 11 }} }} }},
        yImp: {{ position: 'right', beginAtZero: true, grid: {{ drawOnChartArea: false }},
                 title: {{ display: true, text: 'Impressions', font: {{ size: 11 }} }},
                 ticks: {{ font: {{ size: 11 }},
                          callback: v => v >= 1e6 ? (v/1e6).toFixed(1)+'M' : v >= 1e3 ? (v/1e3).toFixed(0)+'K' : v }} }},
        x: {{ ticks: {{ font: {{ size: 11 }}, maxRotation: 50, minRotation: 0 }} }}
      }}
    }}
  }});
}})();

new Chart(document.getElementById('chart-page-type'), {{
  type: 'pie',
  data: {{
    labels: CD.by_page_type_clicks.map(d => d.label),
    datasets: [{{ data: CD.by_page_type_clicks.map(d => d.value),
      backgroundColor: CD.by_page_type_clicks.map(d => CD.pt_colors[d.label] || '#94a3b8'),
      borderWidth: 2, borderColor: '#fff' }}]
  }},
  options: {{ responsive: true, maintainAspectRatio: false,
             plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }}, padding: 12 }} }} }} }}
}});

new Chart(document.getElementById('chart-funnel'), {{
  type: 'pie',
  data: {{
    labels: CD.by_funnel_clicks.map(d => d.label),
    datasets: [{{ data: CD.by_funnel_clicks.map(d => d.value),
      backgroundColor: CD.by_funnel_clicks.map(d => CD.fn_colors[d.label] || '#94a3b8'),
      borderWidth: 2, borderColor: '#fff' }}]
  }},
  options: {{ responsive: true, maintainAspectRatio: false,
             plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }}, padding: 12 }} }} }} }}
}});

new Chart(document.getElementById('chart-status'), {{
  type: 'doughnut',
  data: {{
    labels: CD.status_mix.map(d => d.label),
    datasets: [{{ data: CD.status_mix.map(d => d.value),
      backgroundColor: CD.status_mix.map(d => d.color),
      borderWidth: 2, borderColor: '#fff' }}]
  }},
  options: {{ responsive: true, maintainAspectRatio: false,
             plugins: {{ legend: {{ position: 'bottom' }} }} }}
}});
</script>
</body>
</html>
"""

with open(OUT, "w") as f:
    f.write(html)
print(f"\nSaved: {OUT} ({len(html):,} bytes)")
