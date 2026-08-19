#!/usr/bin/env python3
"""Build the WQA audit spreadsheet — the v3 first output.

Tabs (final order):
1. Aggregator      — every URL with full data
2. Recommendations — project-plan matrix with Approval column
3. Keywords        — every ranking keyword
4. Notes           — starter template
5. Redirects       — 3xx with source → target
6. Errors          — 4xx with traffic/backlink priority
7. Action Plan     — grouped summary by action type

Snapshots the original "Specific Next Step" text per row to
{slug}-wqa-recommendations-original.json for later approval-diff.

Usage:
  python3 build_audit_xlsx.py --client-slug the-blueprint-training --audit-id <id>
  python3 build_audit_xlsx.py --audit-dir /path/to/audit-folder --client-slug the-blueprint-training
"""

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============ CLI ============

parser = argparse.ArgumentParser()
parser.add_argument("--client-slug", required=True, help="Client slug, e.g. the-blueprint-training")
parser.add_argument("--audit-id", help="Audit UUID (folder name under wqa/audits/)")
parser.add_argument("--audit-dir", help="Explicit audit folder path (overrides --audit-id)")
parser.add_argument("--workspace-root", default=os.environ.get("AGENCY_OS_ROOT", "."),
                    help="Workspace root (default: $AGENCY_OS_ROOT or cwd)")
parser.add_argument("--root-domain", required=True,
                    help="Client root domain, e.g. https://theblueprint.training")
args = parser.parse_args()

SLUG = args.client_slug
ROOT = args.root_domain.rstrip("/")

if args.audit_dir:
    A_BASE = args.audit_dir
    # Derive client folder from audit-dir: .../clients/<folder>/wqa/audits/<id>
    CLIENT_BASE = os.path.normpath(os.path.join(A_BASE, "..", "..", ".."))
elif args.audit_id:
    CLIENT_BASE = os.path.join(args.workspace_root, "clients", SLUG)
    A_BASE = os.path.join(CLIENT_BASE, "wqa", "audits", args.audit_id)
else:
    sys.exit("Provide --audit-id or --audit-dir.")

assert os.path.isdir(A_BASE), f"Audit folder not found: {A_BASE}"

# --- Hard prerequisite: SF crawl ---
# Screaming Frog is the source of truth for per-page crawl data (title, meta,
# H1, word count, inlinks, outlinks, crawl depth, indexability). Ahrefs is
# used for keywords + backlinks only — it does NOT substitute for SF.
SF_CRAWL_PRECHECK = os.path.join(args.workspace_root if not args.audit_dir else os.path.normpath(os.path.join(A_BASE, '..', '..', '..')), "crawls", "latest-crawl.json")
if not os.path.exists(SF_CRAWL_PRECHECK):
    sys.exit(
        "\nERROR: No Screaming Frog crawl found at:\n  " + SF_CRAWL_PRECHECK + "\n\n"
        "WQA requires an SF crawl as the source of truth for per-page on-page data\n"
        "(title, meta description, H1, word count, inlinks, outlinks, crawl depth,\n"
        "indexability). Ahrefs is used for keywords and backlinks only — it does\n"
        "NOT substitute for an SF crawl.\n\n"
        "To proceed:\n"
        "  1. Run Screaming Frog SEO Spider against the client's domain.\n"
        "  2. Export the Internal HTML report (Internal > HTML).\n"
        "  3. Use wqa_upload_crawl OR drop the CSV into clients/{slug}/crawls/.\n"
        "  4. Re-run this script.\n"
    )

# Validate the SF crawl has actual data
try:
    _sf_check = json.load(open(SF_CRAWL_PRECHECK))
    _sf_pages_check = _sf_check.get("pages", []) if isinstance(_sf_check, dict) else _sf_check
    if len(_sf_pages_check) < 10:
        sys.exit(
            f"\nERROR: SF crawl at {SF_CRAWL_PRECHECK} contains only {len(_sf_pages_check)} pages.\n"
            "Expected at least 10. Re-export from Screaming Frog with the full Internal HTML report."
        )
except (json.JSONDecodeError, KeyError) as e:
    sys.exit(f"\nERROR: SF crawl file at {SF_CRAWL_PRECHECK} is malformed: {e}")

# Inputs
CRAWLED_PAGES = os.path.join(A_BASE, f"{SLUG}-ahrefs-crawled-pages.json")
SITE_AUDIT    = os.path.join(A_BASE, f"{SLUG}-ahrefs-pages.json")
REDIR_TARGETS = os.path.join(A_BASE, f"{SLUG}-redirect-targets.json")
TOP_PAGES     = os.path.join(A_BASE, f"{SLUG}-ahrefs.json")
KEYWORDS      = os.path.join(A_BASE, f"{SLUG}-ahrefs-keywords.json")
GA4_PATH      = os.path.join(A_BASE, f"{SLUG}-ga4.json")
GSC_PATH      = os.path.join(A_BASE, f"{SLUG}-gsc.json")
GSC_PRIOR     = os.path.join(A_BASE, f"{SLUG}-gsc-90d-prior.json")  # per-URL prior 90d (for trend Δ)
GA4_PRIOR     = os.path.join(A_BASE, f"{SLUG}-ga4-90d-prior.json")  # per-URL prior 90d (optional)
SF_CRAWL      = os.path.join(CLIENT_BASE, "crawls", "latest-crawl.json")

# Outputs
OUT       = os.path.join(A_BASE, f"{SLUG}-wqa-data.xlsx")
SNAPSHOT  = os.path.join(A_BASE, f"{SLUG}-wqa-recommendations-original.json")


# ============ Helpers ============

def norm(u):
    if not u: return ""
    u = u.split("?")[0].split("#")[0].rstrip("/")
    return u.lower()

def strip_root(url):
    if not url: return ""
    if url.startswith(ROOT):
        path = url[len(ROOT):]
        return path if path else "/"
    return url

def load(path, default=None):
    if not os.path.exists(path):
        return default if default is not None else []
    with open(path) as f:
        return json.load(f)


# ============ Load data ============

crawled = load(CRAWLED_PAGES)
site_audit = load(SITE_AUDIT)
redirect_targets = load(REDIR_TARGETS)
top_pages = load(TOP_PAGES)
keywords_data = load(KEYWORDS)
ga4_rows = load(GA4_PATH)
gsc_rows = load(GSC_PATH)
sf_pages = []
if os.path.exists(SF_CRAWL):
    sf_pages = load(SF_CRAWL).get("pages", []) if isinstance(load(SF_CRAWL), dict) else load(SF_CRAWL)

crawled_by_url    = {norm(p["url"]): p for p in crawled}
site_audit_by_url = {norm(p["url"]): p for p in site_audit}
target_by_url     = {norm(r["source_url"]): r for r in redirect_targets}
top_by_url        = {norm(r["url"]): r for r in top_pages}
sf_by_url         = {norm(p["url"]): p for p in sf_pages}


# Aggregate GA4 + GSC by normalized URL — page_location and page often have
# query-string / fragment variants that collapse to the same canonical URL.
# Last-write-wins would lose 90%+ of the homepage sessions, so we sum.
def aggregate_ga4(rows):
    out = {}
    for r in rows:
        k = norm(r.get("page_location") or r.get("page"))
        if not k: continue
        if k not in out:
            out[k] = {"sessions": 0, "users": 0, "engagement_rate": 0.0,
                      "_eng_weight": 0.0, "conversions": 0}
        d = out[k]
        s = r.get("sessions") or 0
        d["sessions"] += s
        d["users"] += r.get("users") or 0
        d["conversions"] += r.get("conversions") or 0
        # session-weighted engagement_rate
        if s > 0 and r.get("engagement_rate") is not None:
            d["_eng_weight"] += s
            d["engagement_rate"] += (r["engagement_rate"]) * s
    for d in out.values():
        if d["_eng_weight"] > 0:
            d["engagement_rate"] = d["engagement_rate"] / d["_eng_weight"]
        del d["_eng_weight"]
    return out


def aggregate_gsc(rows):
    out = {}
    for r in rows:
        k = norm(r.get("page"))
        if not k: continue
        if k not in out:
            out[k] = {"clicks": 0, "impressions": 0, "_pos_weight": 0.0, "position": 0.0}
        d = out[k]
        d["clicks"] += r.get("clicks") or 0
        imp = r.get("impressions") or 0
        d["impressions"] += imp
        if imp > 0 and r.get("position") is not None:
            d["_pos_weight"] += imp
            d["position"] += r["position"] * imp
    for d in out.values():
        d["ctr"] = (d["clicks"] / d["impressions"]) if d["impressions"] else 0
        if d["_pos_weight"] > 0:
            d["position"] = d["position"] / d["_pos_weight"]
        del d["_pos_weight"]
    return out


ga4_by_url = aggregate_ga4(ga4_rows)
gsc_by_url = aggregate_gsc(gsc_rows)

# Prior 90d per-URL data for trend Δ. These files are produced during data pull
# (gsc-90d-prior.json) — if missing, trend columns show blank.
gsc_prior_by_url = {}
if os.path.exists(GSC_PRIOR):
    _prior_data = load(GSC_PRIOR, {})
    if isinstance(_prior_data, dict):
        gsc_prior_by_url = _prior_data.get("by_page", {}) or {}
ga4_prior_by_url = {}
if os.path.exists(GA4_PRIOR):
    _prior_ga4 = load(GA4_PRIOR, [])
    # Expected shape: list of {page_location, sessions, users, conversions, engagement_rate}
    if isinstance(_prior_ga4, list):
        ga4_prior_by_url = {norm(r.get("page_location") or r.get("page")): r for r in _prior_ga4 if r.get("page_location") or r.get("page")}

def pct_delta(cur, prev):
    """Return Δ as a fraction (0.15 = +15%). None if no prior data or div-by-zero."""
    if prev is None or prev == 0 or cur is None: return None
    return (cur - prev) / prev

kw_count_by_url = Counter(norm(k["best_position_url"]) for k in keywords_data if k.get("best_position_url"))
kw_traffic_by_url = defaultdict(int)
for k in keywords_data:
    if k.get("best_position_url"):
        kw_traffic_by_url[norm(k["best_position_url"])] += k.get("sum_traffic") or 0


# ============ Universe ============

domain_host = ROOT.replace("https://", "").replace("http://", "")
universe = set()
for u in crawled_by_url:    universe.add(u)
for u in site_audit_by_url: universe.add(u)
for u in sf_by_url:         universe.add(u)
for u in gsc_by_url:        universe.add(u)
for u in top_by_url:        universe.add(u)
for u in ga4_by_url:        universe.add(u)
universe = {u for u in universe if u and domain_host in u}

ASSET_EXT = (".png", ".jpg", ".jpeg", ".webp", ".svg", ".gif", ".ico", ".css", ".js",
             ".pdf", ".xml", ".woff", ".woff2", ".ttf")
universe = {u for u in universe if not u.endswith(ASSET_EXT)}
print(f"URL universe: {len(universe)}")


# ============ Classification ============

MONEY = ["/pricing", "/enroll", "/apply", "/join", "/membership", "/program",
         "/courses", "/course", "/coaching", "/1-on-1", "/workshops",
         "/book-a-call", "/strategy-call", "/agency-program", "/sop-library",
         "/community", "/sign-up", "/get-started", "/templates-and-sops",
         "/slack-channel", "/agency-training", "/bootcamps",
         "/seo-sprints-training", "/agents", "/blueprint-affiliate-program",
         "/services", "/service", "/products", "/product"]
AUTHORITY = ["/about", "/team", "/founder", "/testimonials", "/success-stories",
             "/case-studies", "/results", "/reviews", "/wins", "/author"]
UTILITY = ["/contact", "/support", "/help", "/faq", "/privacy", "/privacy-policy",
           "/terms", "/terms-and-conditions", "/refund-policy", "/login",
           "/dashboard", "/checkout", "/book-goodies", "/book", "/user",
           "/sign-up/embed", "/feed", "/category", "/tag", "/wp-content"]
CONTENT = ["/blog", "/resources", "/guides", "/podcast", "/videos",
           "/free-training", "/youtube"]

# ---- Per-client pattern overrides (clients/{slug}/wqa/patterns.json) ----
# These EXTEND the hardcoded defaults above so the classifier can learn a
# vertical's URL structure (e.g. law-firm practice-area / city service pages
# like /louisville-bankruptcy-attorney or /owensboro-bankruptcy-lawyers that
# would otherwise fall through to "Blog Post" and tank the link roster).
#
#   "*_prefixes"  -> matched against the full path (startswith)
#   "*_contains"  -> matched against the FIRST path segment (substring). Used
#                    for top-level service slugs that don't share a prefix.
MONEY_CONTAINS, AUTHORITY_CONTAINS, UTILITY_CONTAINS, CONTENT_CONTAINS = [], [], [], []
_patterns_path = os.path.join(CLIENT_BASE, "wqa", "patterns.json")
if os.path.exists(_patterns_path):
    try:
        _pat = json.load(open(_patterns_path))
        MONEY     += _pat.get("money_prefixes", [])
        AUTHORITY += _pat.get("authority_prefixes", [])
        UTILITY   += _pat.get("utility_prefixes", [])
        CONTENT   += _pat.get("content_prefixes", [])
        MONEY_CONTAINS     = _pat.get("money_contains", [])
        AUTHORITY_CONTAINS = _pat.get("authority_contains", [])
        UTILITY_CONTAINS   = _pat.get("utility_contains", [])
        CONTENT_CONTAINS   = _pat.get("content_contains", [])
        print(f"Loaded client patterns: +{len(_pat.get('money_prefixes',[]))} money prefixes, "
              f"{len(MONEY_CONTAINS)} money contains, +{len(_pat.get('authority_prefixes',[]))} authority prefixes")
    except Exception as e:
        print(f"WARN: could not parse patterns.json ({e}); using defaults only")


def classify_page_type(url):
    path = url.replace(ROOT, "").rstrip("/")
    if path == "": return "Homepage"
    seg1 = path.lstrip("/").split("/")[0].lower()
    def matches(p, patterns):
        return any(p == pat or p.startswith(pat + "/") for pat in patterns)
    def seg_has(needles):
        return any(n.strip("/").lower() in seg1 for n in needles)
    # Vertical "contains" rules run alongside the prefix defaults; Money first
    # so service/practice slugs win over a generic content fallback.
    if matches(path, MONEY)     or seg_has(MONEY_CONTAINS):     return "Money Page"
    if matches(path, AUTHORITY) or seg_has(AUTHORITY_CONTAINS): return "Authority Page"
    if matches(path, UTILITY)   or seg_has(UTILITY_CONTAINS):   return "Utility Page"
    if matches(path, CONTENT)   or seg_has(CONTENT_CONTAINS):   return "Content Hub"
    return "Blog Post"

def funnel_stage(page_type, url):
    if page_type == "Homepage":       return "Brand"
    if page_type == "Money Page":     return "Bottom Funnel"
    if page_type == "Authority Page": return "Mid Funnel"
    if page_type == "Utility Page":   return "Utility"
    if page_type == "Content Hub":    return "Top Funnel"
    return "Top Funnel"

def status_for(u):
    """SF crawl is the primary source of truth. Ahrefs supplements only for
    URLs SF didn't reach (e.g. legacy URLs still showing in GSC)."""
    s = sf_by_url.get(u)
    if s and s.get("status_code"): return s["status_code"]
    c = crawled_by_url.get(u)
    if c and c.get("http_code"): return c["http_code"]
    sa = site_audit_by_url.get(u)
    if sa:
        if "Page" in (sa.get("page_type") or []): return 200
        if sa.get("redirect_code"): return sa["redirect_code"]
    return None

def status_type(code):
    if code is None: return "unknown"
    if code == 200:  return "live"
    if 300 <= code < 400: return "redirect"
    if 400 <= code < 500: return "broken"
    if code == 429:  return "rate_limited"
    if 500 <= code < 600: return "error"
    return f"http_{code}"

def get_redirect_target(u, code):
    """Resolve where a 3xx URL redirects to. Prefer SF crawl's Redirect URL field
    (the most current/authoritative); fall back to HEAD-fetch results from
    discover_redirects.py if available."""
    if code is None or not (300 <= code < 400): return None
    s = sf_by_url.get(u)
    if s and s.get("redirect_url"):
        return s["redirect_url"]
    t = target_by_url.get(u)
    return t.get("target_url") if t else None


# ============ Build records ============

records = []
for u in sorted(universe):
    code = status_for(u)
    s_type = status_type(code)
    pt = classify_page_type(u)
    target = get_redirect_target(u, code)
    sf = sf_by_url.get(u, {})
    sa = site_audit_by_url.get(u, {})
    g4 = ga4_by_url.get(u, {})
    sc = gsc_by_url.get(u, {})
    tp = top_by_url.get(u, {})
    cr = crawled_by_url.get(u, {})

    title = sf.get("title") or (cr.get("title") if cr.get("title") else "") or ""
    meta_arr = sa.get("meta_description") or []
    meta = sf.get("meta_description") or (meta_arr[0] if meta_arr else "")
    h1 = sf.get("h1") or ""
    word_count = sf.get("word_count")
    inlinks = sf.get("inlinks") if sf else sa.get("incoming_links")
    outlinks = sf.get("outlinks") if sf else None
    crawl_depth = sf.get("crawl_depth") if sf else None
    indexable = (sf.get("indexability") == "Indexable") if sf.get("indexability") else (
        not sa.get("page_is_noindex") if sa else None
    )

    # has_crawl_data = at least one on-page field came back from SF/site-audit crawl
    has_crawl_data = bool(sf) or bool(sa)

    # Per-URL trend deltas (current 90d vs prior 90d)
    g4_p = ga4_prior_by_url.get(u, {})
    sc_p = gsc_prior_by_url.get(u, {})
    d_ga4_sessions = pct_delta(g4.get("sessions") or 0, g4_p.get("sessions"))
    d_ga4_users    = pct_delta(g4.get("users") or 0,    g4_p.get("users"))
    d_ga4_conv     = pct_delta(g4.get("conversions") or 0, g4_p.get("conversions"))
    d_gsc_clicks   = pct_delta(sc.get("clicks") or 0,   sc_p.get("clicks"))
    d_gsc_imp      = pct_delta(sc.get("impressions") or 0, sc_p.get("impressions"))
    d_gsc_ctr      = pct_delta(sc.get("ctr") or 0,      sc_p.get("ctr"))
    d_gsc_pos      = pct_delta(sc.get("position") or 0, sc_p.get("position"))

    records.append({
        "url": u, "address": strip_root(u),
        "status_code": code, "status_type": s_type,
        "redirect_target": target,
        "page_type": pt, "funnel": funnel_stage(pt, u),
        "indexable": indexable,
        "has_crawl_data": has_crawl_data,
        "title": title, "meta_description": meta, "h1": h1,
        "word_count": word_count, "crawl_depth": crawl_depth,
        "inlinks": inlinks, "outlinks": outlinks,
        "ga4_sessions": g4.get("sessions") or 0,
        "ga4_users": g4.get("users") or 0,
        "ga4_engagement_rate": g4.get("engagement_rate") or 0,
        "ga4_conversions": g4.get("conversions") or 0,
        "d_ga4_sessions": d_ga4_sessions,
        "d_ga4_users":    d_ga4_users,
        "d_ga4_conv":     d_ga4_conv,
        "gsc_clicks": sc.get("clicks") or 0,
        "gsc_impressions": sc.get("impressions") or 0,
        "gsc_ctr": sc.get("ctr") or 0,
        "gsc_position": sc.get("position") or 0,
        "d_gsc_clicks":   d_gsc_clicks,
        "d_gsc_imp":      d_gsc_imp,
        "d_gsc_ctr":      d_gsc_ctr,
        "d_gsc_pos":      d_gsc_pos,
        "ahrefs_traffic": tp.get("sum_traffic") or 0,
        "ahrefs_keywords": kw_count_by_url.get(u, tp.get("keywords") or 0),
        "top_keyword": tp.get("top_keyword") or "",
        "top_kw_volume": tp.get("top_keyword_volume") or 0,
        "top_kw_pos": tp.get("top_keyword_best_position"),
        "referring_domains": tp.get("referring_domains") or 0,
        "ur": tp.get("ur"),
        "last_crawled": cr.get("last_crawled", ""),
    })


# ============ Issue detection ============
# Each finding is a tuple: (category, action, priority, notes, sprint).
# A single page can generate multiple findings — each becomes a Recommendation row.
# The Aggregator shows the highest-priority finding in each (Technical, Content) column.

PRIO_RANK = {"P1": 1, "P2": 2, "P3": 3, "": 9}
GARBAGE_PATTERNS = ("/feed", "/wp-content/", "/wp-json/", "/wp-admin/",
                    "/tag/", "/category/", "/page/",
                    "/trainings_cat/", "/trainings_tag/")

def is_likely_garbage(address):
    """WordPress auto-generated taxonomy/feed pages — should be noindexed."""
    a = address.lower()
    return any(p in a for p in GARBAGE_PATTERNS)


def detect_issues(rec):
    """Return list of findings. Empty list = no action needed."""
    findings = []
    code = rec["status_code"]
    imp = rec["gsc_impressions"]
    bl = rec["referring_domains"]
    ses = rec["ga4_sessions"]
    pos = rec["gsc_position"]
    ctr = rec["gsc_ctr"]
    wc = rec.get("word_count") or 0
    pt = rec["page_type"]
    title = (rec.get("title") or "").strip()
    meta = (rec.get("meta_description") or "").strip()
    h1 = (rec.get("h1") or "").strip()
    top_vol = rec.get("top_kw_volume") or 0
    top_kw = rec.get("top_keyword") or ""
    address = rec["address"]
    has_crawl = rec.get("has_crawl_data", False)

    # Sprint mapping per category — Sprint 1 is reserved for the WQA itself
    # (planning + onboarding + audit). Auto-detected actions land in:
    #   Sprint 2 = Technical · Sprint 3 = Local · Sprint 4 = Content · Sprint 5 = Links
    # The string values MUST match the Aggregator + Recommendations Sprint dropdowns exactly.
    S_TECH    = "Sprint 2 (Technical)"
    S_CONTENT = "Sprint 4 (Content)"
    S_BACKLOG = "Backlog"

    # ------- 4xx broken — short-circuit, ignore content fields -------
    if code and 400 <= code < 500:
        if imp > 100 or bl > 5 or ses > 10:
            findings.append(("Technical", "Fix 404", "P1",
                f"HIGH-SIGNAL 4xx ({code}) — {imp} imp, {bl} RDs, {ses} sessions/90d. Restore or 301 to closest live equivalent.",
                S_TECH))
        elif imp > 0 or bl > 0:
            findings.append(("Technical", "Fix 404", "P2",
                f"4xx ({code}) with some signal — {imp} imp, {bl} RDs. 301 to relevant live URL.",
                S_TECH))
        else:
            findings.append(("Technical", "Monitor", "P3",
                f"4xx ({code}) with no signal. Likely safe to ignore.", S_BACKLOG))
        return findings

    # ------- 3xx redirects -------
    if code and 300 <= code < 400:
        tgt = rec.get("redirect_target") or "unknown"
        if (imp > 100 or bl > 5) and tgt and tgt.rstrip("/") == ROOT:
            findings.append(("Technical", "Evaluate redirect", "P2",
                f"301 to homepage with {imp} imp / {bl} RDs. Redirect to topical match to preserve relevance.",
                S_TECH))
        else:
            findings.append(("Technical", "Leave as is", "P3",
                f"301 → {tgt}. Working as intended.", "Done"))
        return findings

    # ------- 429 etc -------
    if code == 429:
        findings.append(("Technical", "Monitor", "P3",
            "Ahrefs got rate-limited during crawl. Re-check status next audit.", S_BACKLOG))
        return findings

    # ===== Status = live (or unknown) — issues can stack =====

    # Garbage taxonomy/feed pages indexed
    if is_likely_garbage(address):
        if wc == 0 and not title:
            findings.append(("Technical", "Noindex", "P2",
                f"WP auto-generated archive/feed page with no meaningful content. Add noindex to remove from crawl waste.",
                S_TECH))
        return findings  # don't apply content rules to garbage URLs

    # ------- Missing critical on-page elements (only if we have crawl data) -------
    if has_crawl:
        if not title:
            p = "P1" if pt in ("Money Page", "Homepage", "Authority Page") else "P2"
            findings.append(("Content", "Rewrite title/meta", p,
                f"Missing page title. Add an SEO-optimized title (50-60 chars) targeting the page's primary keyword.",
                S_CONTENT))
        if title and not meta:
            p = "P2" if pt in ("Money Page", "Authority Page", "Homepage") else "P3"
            findings.append(("Content", "Rewrite title/meta", p,
                f"Missing meta description. Add a 140-160 char description with target keyword + CTA.",
                S_CONTENT))
        if title and not h1:
            p = "P2" if pt in ("Money Page", "Authority Page", "Homepage") else "P3"
            findings.append(("Content", "Update onpage", p,
                f"Missing H1. Add a clear H1 that matches user search intent and uses the primary keyword.",
                S_CONTENT))

    # ------- Thin content -------
    if has_crawl and 0 < wc < 618:
        if pt == "Money Page":
            findings.append(("Content", "Rewrite", "P1",
                f"Thin money page ({wc}w). Expand to 1,000w+ with conversion content, social proof, FAQs.",
                S_CONTENT))
        elif pt == "Authority Page":
            findings.append(("Content", "Rewrite", "P2",
                f"Thin authority page ({wc}w). Expand with credentials, expertise, social proof, schema.",
                S_CONTENT))
        else:
            findings.append(("Content", "Expand content", "P3",
                f"Thin content ({wc}w). Expand to 1,000w+ OR consolidate/noindex if low-value.",
                S_CONTENT))

    # ------- Depth gap: high-volume target KW but content under-depth -------
    if has_crawl and 618 <= wc < 1500 and top_vol >= 500 \
       and pt in ("Money Page", "Blog Post", "Content Hub", "Authority Page"):
        findings.append(("Content", "Expand content", "P2",
            f"Targets {top_vol:,}-vol KW '{top_kw}' but only {wc}w. Top-ranking competitors typically run 1,800-2,500w. Expand with depth + supporting topics.",
            S_CONTENT))

    # ------- Low CTR with high impressions -------
    if imp > 1000 and ctr < 0.02 and pos > 0:
        findings.append(("Content", "Rewrite title/meta", "P2",
            f"High impressions ({imp:,}) but low CTR ({ctr*100:.1f}%) at pos {pos:.1f}. Rewrite title/meta to match intent + add a hook.",
            S_CONTENT))

    # ------- Striking distance refresh -------
    if 4 <= pos <= 15 and imp >= 500:
        findings.append(("Content", "Update onpage", "P2",
            f"Striking distance — pos {pos:.1f} with {imp:,} imp. Refresh content + add internal links to push into top 5.",
            S_CONTENT))

    # ------- Stale-blog heuristic (proxy for content age — without last_modified) -------
    if (pt == "Blog Post" and has_crawl and wc >= 1000
            and 11 <= pos <= 30 and imp >= 200):
        if not any(f[1] == "Update onpage" for f in findings):
            findings.append(("Content", "Refresh content", "P3",
                f"Substantial blog post ({wc}w) stuck at pos {pos:.1f} with {imp:,} imp. Likely stale — refresh with current data, new examples, and updated screenshots.",
                S_CONTENT))

    return findings


# Apply detection and derive Aggregator-level primary actions
for r in records:
    findings = detect_issues(r)
    r["all_findings"] = findings
    tech_findings = sorted([f for f in findings if f[0] == "Technical"],
                           key=lambda x: PRIO_RANK.get(x[2], 9))
    content_findings = sorted([f for f in findings if f[0] == "Content"],
                              key=lambda x: PRIO_RANK.get(x[2], 9))
    r["tech_action"] = tech_findings[0][1] if tech_findings else ""
    r["content_action"] = content_findings[0][1] if content_findings else ""
    if findings:
        best_prio_rank = min(PRIO_RANK.get(f[2], 9) for f in findings)
        r["priority"] = next((p for p, rnk in PRIO_RANK.items() if rnk == best_prio_rank), "")
        # Sprint from highest-priority finding
        best_finding = min(findings, key=lambda f: PRIO_RANK.get(f[2], 9))
        r["sprint"] = best_finding[4]
    else:
        r["priority"] = ""
        r["sprint"] = ""
    r["action_notes"] = " | ".join(f[3] for f in findings)

def problem_areas(r):
    a = []
    if r["status_code"] and 400 <= r["status_code"] < 500: a.append("error")
    if r["status_code"] and 300 <= r["status_code"] < 400: a.append("redirect")
    if r.get("word_count") is not None and r["word_count"] < 618 and r["status_type"] == "live":
        a.append("thin_content")
    if r["gsc_impressions"] > 1000 and r["gsc_ctr"] < 0.02: a.append("low_ctr")
    if 5 <= r["gsc_position"] <= 15 and r["gsc_impressions"] > 500: a.append("striking_distance")
    return ", ".join(a)

for r in records:
    r["problem_areas"] = problem_areas(r)


# ============ Styling ============

FONT_NAME = "Arial"
hdr_font  = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
hdr_fill  = PatternFill("solid", fgColor="2563EB")
body_font = Font(name=FONT_NAME, size=10)
hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
overflow_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
link_font = Font(name=FONT_NAME, size=10, color="2563EB", underline="single")

TYPE_FILLS = {
    "Homepage":       PatternFill("solid", fgColor="DDD6FE"),
    "Money Page":     PatternFill("solid", fgColor="EDE9FE"),
    "Authority Page": PatternFill("solid", fgColor="CFFAFE"),
    "Content Hub":    PatternFill("solid", fgColor="DBEAFE"),
    "Blog Post":      PatternFill("solid", fgColor="E0E7FF"),
    "Utility Page":   PatternFill("solid", fgColor="F1F5F9"),
}
PRIO_FILLS = {
    "P1": PatternFill("solid", fgColor="FECACA"),
    "P2": PatternFill("solid", fgColor="FED7AA"),
    "P3": PatternFill("solid", fgColor="FEF3C7"),
}
STATUS_FILLS = {
    "live":         PatternFill("solid", fgColor="DCFCE7"),
    "redirect":     PatternFill("solid", fgColor="FEF3C7"),
    "broken":       PatternFill("solid", fgColor="FEE2E2"),
    "rate_limited": PatternFill("solid", fgColor="E0E7FF"),
    "unknown":      PatternFill("solid", fgColor="F1F5F9"),
}
TECH_FILLS = {
    "Fix 404":               PatternFill("solid", fgColor="FCA5A5"),
    "301 redirect":          PatternFill("solid", fgColor="FED7AA"),
    "Evaluate redirect":     PatternFill("solid", fgColor="E0E7FF"),
    "Add canonical":         PatternFill("solid", fgColor="FEF3C7"),
    "Add schema":            PatternFill("solid", fgColor="FEF3C7"),
    "Noindex":               PatternFill("solid", fgColor="F1F5F9"),
    "Update robots":         PatternFill("solid", fgColor="F1F5F9"),
    "Indexability fix":      PatternFill("solid", fgColor="FED7AA"),
    "Sitemap fix":           PatternFill("solid", fgColor="F1F5F9"),
    "Monitor":               PatternFill("solid", fgColor="F1F5F9"),
    "Leave as is":           PatternFill("solid", fgColor="DCFCE7"),
}
CONTENT_FILLS = {
    "Rewrite":               PatternFill("solid", fgColor="FECACA"),
    "Rewrite title/meta":    PatternFill("solid", fgColor="FED7AA"),
    "Expand content":        PatternFill("solid", fgColor="FEF3C7"),
    "Refresh content":       PatternFill("solid", fgColor="FEF3C7"),
    "Update onpage":         PatternFill("solid", fgColor="FEF3C7"),
    "Consolidate":           PatternFill("solid", fgColor="E0E7FF"),
    "Evaluate":              PatternFill("solid", fgColor="E0E7FF"),
    "Leave as is":           PatternFill("solid", fgColor="DCFCE7"),
}
# Combined for legacy code that references ACTION_FILLS (recommendations tab)
ACTION_FILLS = {**TECH_FILLS, **CONTENT_FILLS}

TECH_DROPDOWN = '"Fix 404,301 redirect,Evaluate redirect,Add canonical,Add schema,Noindex,Update robots,Indexability fix,Sitemap fix,Monitor,Leave as is"'
CONTENT_DROPDOWN = '"Rewrite,Rewrite title/meta,Expand content,Refresh content,Update onpage,Consolidate,Evaluate,Leave as is"'


def write_headers(ws, headers, widths):
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def apply_body_style(ws, wrap_cols=None):
    wrap_cols = set(wrap_cols or [])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = link_font if cell.hyperlink else body_font
            cell.alignment = wrap_align if cell.column in wrap_cols else overflow_align


# ============ Build workbook ============

wb = Workbook()

# --- Aggregator ---
ws = wb.active
ws.title = "Aggregator"
agg_headers = [
    "Technical Action", "Content Action", "Priority", "Sprint",
    "Address", "Page Type",            # cols A-F, all frozen
    "Status", "Status Type", "Redirect Target",
    "Funnel Stage", "Indexable",
    "Title", "Title Len", "Meta Description", "Meta Len", "H1",
    "Word Count", "Crawl Depth", "Inlinks",
    # GA4 — 90d current + Δ vs prior 90d
    "GA4 Sessions (90d)", "Sessions Δ", "GA4 Users (90d)", "Users Δ",
    "GA4 Eng % (90d)", "GA4 Conv (90d)", "Conv Δ",
    # GSC — 90d current + Δ vs prior 90d
    "GSC Clicks (90d)", "Clicks Δ", "GSC Imp (90d)", "Imp Δ",
    "GSC CTR (90d)", "CTR Δ",
    "Total KWs Ranking", "Top Keyword", "Top KW Vol", "Top KW Pos",
    "Ref Domains", "URL Rating",
    "Problem Areas", "Action Notes"
]
agg_widths = [16, 16, 9, 12, 35, 16,             # A-F
              8, 13, 45, 14, 10,                   # G-K
              50, 10, 55, 10, 45,                  # L-P
              11, 12, 9,                           # Q-S
              13, 9, 11, 9, 10, 10, 9,             # GA4 block
              11, 9, 11, 9, 9, 9,                  # GSC block
              13, 32, 11, 11, 12, 11, 24, 70]
write_headers(ws, agg_headers, agg_widths)

ACTION_PRIO = {"Rewrite": 0, "Rewrite title/meta": 1, "Fix 404": 2, "Update onpage": 3,
               "Expand content": 3, "Evaluate redirect": 4, "Evaluate": 4,
               "Monitor": 5, "Leave as is": 6, "": 7}

def sort_key(r):
    primary = r["tech_action"] or r["content_action"] or ""
    return (ACTION_PRIO.get(primary, 9), r["priority"] or "P9", -r["gsc_impressions"])

records_sorted = sorted(records, key=sort_key)

def fmt_delta(d):
    """Format delta as ▲ +12.3% / ▼ -5.4% / blank if unavailable."""
    if d is None: return ""
    pct = d * 100
    if abs(pct) < 0.05: return "—"
    arrow = "▲" if pct > 0 else "▼"
    return f"{arrow} {pct:+.1f}%"

for r in records_sorted:
    ws.append([
        r["tech_action"], r["content_action"], r["priority"], r["sprint"],
        r["address"], r["page_type"],          # cols A-F (all frozen)
        r["status_code"] or "", r["status_type"], r["redirect_target"] or "",
        r["funnel"],
        "Yes" if r["indexable"] else ("No" if r["indexable"] is False else ""),
        r["title"], len(r["title"] or ""),
        r["meta_description"], len(r["meta_description"] or ""), r["h1"],
        r["word_count"] if r["word_count"] is not None else "",
        r["crawl_depth"] if r["crawl_depth"] is not None else "",
        r["inlinks"] if r["inlinks"] is not None else "",
        # GA4 — current + Δ
        r["ga4_sessions"], fmt_delta(r["d_ga4_sessions"]),
        r["ga4_users"],    fmt_delta(r["d_ga4_users"]),
        round(r["ga4_engagement_rate"], 4),
        r["ga4_conversions"], fmt_delta(r["d_ga4_conv"]),
        # GSC — current + Δ
        r["gsc_clicks"],      fmt_delta(r["d_gsc_clicks"]),
        r["gsc_impressions"], fmt_delta(r["d_gsc_imp"]),
        round(r["gsc_ctr"], 4),       fmt_delta(r["d_gsc_ctr"]),
        r["ahrefs_keywords"],          # Total KWs Ranking (was AH Keywords)
        r["top_keyword"], r["top_kw_volume"],
        r["top_kw_pos"] if r["top_kw_pos"] is not None else "",
        r["referring_domains"],
        r["ur"] if r["ur"] is not None else "",
        r["problem_areas"], r["action_notes"]
    ])

for ridx, r in enumerate(records_sorted, 2):
    addr_cell = ws.cell(row=ridx, column=5)
    addr_cell.hyperlink = r["url"]
    if r["tech_action"] in TECH_FILLS:
        ws.cell(row=ridx, column=1).fill = TECH_FILLS[r["tech_action"]]
    if r["content_action"] in CONTENT_FILLS:
        ws.cell(row=ridx, column=2).fill = CONTENT_FILLS[r["content_action"]]
    if r["priority"] in PRIO_FILLS:
        ws.cell(row=ridx, column=3).fill = PRIO_FILLS[r["priority"]]
    # New column layout: Page Type=6, Status=7, Status Type=8
    ws.cell(row=ridx, column=6).fill = TYPE_FILLS.get(r["page_type"], PatternFill())
    ws.cell(row=ridx, column=8).fill = STATUS_FILLS.get(r["status_type"], PatternFill())
    # % formatting: GA4 Eng % col 24, GSC CTR col 31
    ws.cell(row=ridx, column=24).number_format = "0.00%"
    ws.cell(row=ridx, column=31).number_format = "0.00%"
    # Trend Δ conditional fills (1-indexed): Sessions Δ=21, Users Δ=23, Conv Δ=26, Clicks Δ=28, Imp Δ=30, CTR Δ=32
    POS_FILL = PatternFill("solid", fgColor="DCFCE7")
    NEG_FILL = PatternFill("solid", fgColor="FEE2E2")
    for col, delta_key, inverse in [
        (21, "d_ga4_sessions", False),
        (23, "d_ga4_users", False),
        (26, "d_ga4_conv", False),
        (28, "d_gsc_clicks", False),
        (30, "d_gsc_imp", False),
        (32, "d_gsc_ctr", False),
    ]:
        d = r.get(delta_key)
        if d is None: continue
        if abs(d) < 0.0005: continue
        is_positive = d > 0
        good = (not is_positive) if inverse else is_positive
        ws.cell(row=ridx, column=col).fill = POS_FILL if good else NEG_FILL

# Dropdowns
for col, formula in [
    ("A", TECH_DROPDOWN),
    ("B", CONTENT_DROPDOWN),
    ("C", '"P1,P2,P3"'),
    ("D", '"Sprint 1 (Planning),Sprint 2 (Technical),Sprint 3 (Local),Sprint 4 (Content),Sprint 5 (Links),Backlog,Done"'),
]:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col}2:{col}{ws.max_row}")
    ws.add_data_validation(dv)

ws.freeze_panes = "G2"
ws.auto_filter.ref = ws.dimensions


# --- Recommendations ---
# One row per (page × action). A page with both a tech action and a content action gets two rows.
ws_rec = wb.create_sheet("Recommendations")
rec_headers = ["#", "Approval", "Priority", "Category", "Sprint", "Action Type", "Page Address",
               "Specific Next Step"]
rec_widths = [4, 12, 9, 10, 16, 18, 35, 75]
write_headers(ws_rec, rec_headers, rec_widths)

# One Recommendation row per (page × finding). A page with three findings gets three rows.
recs = []
for r in records:
    for category, action, priority, notes, sprint in r.get("all_findings", []):
        if action in ("Leave as is", "Monitor"):
            continue
        recs.append({
            "category": category,
            "action_type": action,
            "priority": priority,
            "sprint": sprint,
            "url": r["url"],
            "next_step": notes,
            "current": f"Status {r['status_code']} | {r['word_count'] or '–'}w | {r['gsc_impressions']:,} imp · {r['gsc_clicks']:,} clicks · pos {r['gsc_position']:.1f}" if r["status_code"] else f"Status — | {r['word_count'] or '–'}w",
            "owner": "TBD",
            "status": "Not started",
        })

PRIO_ORDER = {"P1": 0, "P2": 1, "P3": 2, "": 9}
recs.sort(key=lambda x: (PRIO_ORDER.get(x["priority"], 9), x["category"], x["url"]))

for i, rec in enumerate(recs, 1):
    ws_rec.append([
        i, "", rec["priority"], rec["category"], rec["sprint"], rec["action_type"],
        strip_root(rec["url"]),
        rec["next_step"],
    ])

CAT_FILLS = {"Technical": PatternFill("solid", fgColor="DBEAFE"),
             "Content":   PatternFill("solid", fgColor="FCE7F3")}

for ridx, rec in enumerate(recs, 2):
    if rec["priority"] in PRIO_FILLS:
        ws_rec.cell(row=ridx, column=3).fill = PRIO_FILLS[rec["priority"]]
    if rec["category"] in CAT_FILLS:
        ws_rec.cell(row=ridx, column=4).fill = CAT_FILLS[rec["category"]]
    if rec["action_type"] in ACTION_FILLS:
        ws_rec.cell(row=ridx, column=6).fill = ACTION_FILLS[rec["action_type"]]
    ws_rec.cell(row=ridx, column=7).hyperlink = rec["url"]

for col, formula in [
    ("B", '"Approved,Edit,Rejected,Deferred"'),
    ("C", '"P1,P2,P3"'),
    ("D", '"Technical,Content"'),
    ("E", '"Sprint 1 (Planning),Sprint 2 (Technical),Sprint 3 (Local),Sprint 4 (Content),Sprint 5 (Links),Backlog,Done"'),
]:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col}2:{col}{ws_rec.max_row}")
    ws_rec.add_data_validation(dv)

# Freeze row 1 only — no column freeze (user prefers full horizontal scroll)
ws_rec.freeze_panes = "A2"
# Set auto_filter explicitly to the 8 columns we use — prevents Excel/Sheets from
# extending the table into empty columns I-Z when the file is opened.
ws_rec.auto_filter.ref = f"A1:H{ws_rec.max_row}"
wb.move_sheet("Recommendations", offset=-1)  # 2nd position


# --- Keywords ---
ws_kw = wb.create_sheet("Keywords")
kw_headers = ["Keyword", "Address", "Position", "Volume", "Est. Traffic", "Actual Clicks",
              "CPC ($)", "KW Difficulty", "Local", "Branded", "Commercial",
              "Informational", "Transactional", "Navigational",
              "SERP Features", "Intent Mix"]
kw_widths = [40, 35, 9, 10, 11, 12, 9, 11, 8, 9, 10, 13, 12, 12, 35, 22]
write_headers(ws_kw, kw_headers, kw_widths)

kws_sorted = sorted(keywords_data, key=lambda k: -(k.get("sum_traffic") or 0))
for k in kws_sorted:
    page_url = k.get("best_position_url") or ""
    intents = [n for n, key in [("Branded", "is_branded"), ("Commercial", "is_commercial"),
               ("Informational", "is_informational"), ("Transactional", "is_transactional"),
               ("Navigational", "is_navigational")] if k.get(key)]
    ws_kw.append([
        k.get("keyword", ""),
        strip_root(page_url),
        k.get("best_position", ""),
        k.get("volume", 0),
        k.get("sum_traffic", 0),
        "",
        round((k.get("cpc") or 0) / 100, 2) if k.get("cpc") else 0,
        k.get("keyword_difficulty", ""),
        *("Yes" if k.get(key) else "" for key in
          ("is_local", "is_branded", "is_commercial", "is_informational",
           "is_transactional", "is_navigational")),
        ", ".join(k.get("serp_features") or []),
        ", ".join(intents),
    ])

for ridx, k in enumerate(kws_sorted, 2):
    page_url = k.get("best_position_url") or ""
    if page_url:
        ws_kw.cell(row=ridx, column=2).hyperlink = page_url
    pos = k.get("best_position") or 999
    cell = ws_kw.cell(row=ridx, column=3)
    cell.fill = PatternFill("solid",
        fgColor="DCFCE7" if pos <= 3 else "FEF3C7" if pos <= 10 else "FED7AA" if pos <= 20 else "FEE2E2")

ws_kw.freeze_panes = "A2"
ws_kw.auto_filter.ref = ws_kw.dimensions


# Notes, Redirects, and Errors tabs removed per request — all redirect/error info
# is already in the Aggregator tab (filter by Status Type column).
# Keep these computed values for the summary print only.
redirects = [r for r in records if r["status_type"] == "redirect"]
errors = [r for r in records if r["status_type"] == "broken"]


# --- Target Pages (top 20 link-building roster) ---
PT_VALUE = {"Money Page": 6.0, "Authority Page": 1.8, "Content Hub": 1.3, "Blog Post": 1.0}

def pos_mult(p):
    if p is None or p == 0: return 0.4
    if p <= 3: return 0.7
    if p <= 10: return 1.5
    if p <= 30: return 1.2
    return 0.4

# High-value page types stay eligible for the roster even with a weak ranking
# signal — an under-ranking service/practice page is exactly where links help,
# so we don't want the "needs impressions" gate to drop them.
HIGH_VALUE_TYPES = ("Money Page", "Authority Page")

target_candidates = []
for r in records:
    if r["status_type"] != "live": continue
    if r["page_type"] not in PT_VALUE: continue
    if r["address"] == "/": continue
    is_high_value = r["page_type"] in HIGH_VALUE_TYPES
    imp = r["gsc_impressions"]
    if not is_high_value and not r["top_keyword"] and imp < 100: continue
    # Find KW match in keywords pull for this URL
    kw_match = next((k for k in keywords_data
                     if k.get("best_position_url") and norm(k["best_position_url"]) == norm(r["url"])), None)
    vol = (kw_match.get("volume") if kw_match else 0) or 0
    if not is_high_value and vol == 0 and imp < 200: continue
    pos = (kw_match.get("best_position") if kw_match else None) or r["gsc_position"]
    score = (vol if vol > 0 else 100) * pos_mult(pos) * PT_VALUE[r["page_type"]]
    target_candidates.append({
        "url": r["url"], "address": r["address"],
        "page_type": r["page_type"],
        "top_keyword": kw_match.get("keyword") if kw_match else r["top_keyword"],
        "kw_volume": vol,
        "current_position": pos if pos else 0,
        "ref_domains": r["referring_domains"],
        "word_count": r["word_count"] or 0,
        "gsc_impressions": imp,
        "score": round(score, 1),
    })

# Tier the roster by business value first, then by score within each tier.
# A link-building target list should lead with money/service pages — an
# under-ranking practice-area page is a better link target than a high-traffic
# informational post, so commercial intent outranks raw volume here.
ROSTER_TIER = {"Money Page": 0, "Authority Page": 1, "Content Hub": 2, "Blog Post": 3}
target_candidates.sort(key=lambda t: (ROSTER_TIER.get(t["page_type"], 9), -t["score"]))
target_top = target_candidates[:20]

ws_tp = wb.create_sheet("Target Pages")
tp_headers = ["#", "Approval", "Page Address", "Page Type", "Target Keyword",
              "KW Volume", "Current Pos", "Ref Domains", "Word Count",
              "GSC Impressions", "Score", "Approver Notes"]
tp_widths = [4, 12, 35, 16, 32, 11, 12, 12, 12, 14, 9, 40]
write_headers(ws_tp, tp_headers, tp_widths)

for i, t in enumerate(target_top, 1):
    ws_tp.append([
        i, "", t["address"], t["page_type"], t["top_keyword"] or "",
        t["kw_volume"], t["current_position"], t["ref_domains"], t["word_count"],
        t["gsc_impressions"], t["score"], ""
    ])

for ridx, t in enumerate(target_top, 2):
    ws_tp.cell(row=ridx, column=3).hyperlink = t["url"]
    ws_tp.cell(row=ridx, column=4).fill = TYPE_FILLS.get(t["page_type"], PatternFill())

if ws_tp.max_row > 1:
    dv_tp = DataValidation(type="list", formula1='"Approved,Edit,Rejected,Deferred"', allow_blank=True)
    dv_tp.add(f"B2:B{ws_tp.max_row}")
    ws_tp.add_data_validation(dv_tp)

ws_tp.freeze_panes = "C2"
ws_tp.auto_filter.ref = ws_tp.dimensions


# ============ Final styling pass ============

WRAP_COLS = {
    "Aggregator":      {14, 40},     # Meta Description · Action Notes
    "Recommendations": {8},          # Specific Next Step
    "Keywords":        set(),
    "Target Pages":    {12},
}
for sheet_name in wb.sheetnames:
    apply_body_style(wb[sheet_name], wrap_cols=WRAP_COLS.get(sheet_name, set()))
    wb[sheet_name].row_dimensions[1].height = 28


# ============ Save xlsx + snapshot ============

wb.save(OUT)
print(f"\nSaved: {OUT}")

# Snapshot the original Specific Next Step per row for later diff
snapshot = []
for i, rec in enumerate(recs, 1):
    snapshot.append({
        "row": i,
        "url": rec["url"],
        "address": strip_root(rec["url"]),
        "category": rec["category"],
        "action_type": rec["action_type"],
        "original_text": rec["next_step"],
    })
with open(SNAPSHOT, "w") as f:
    json.dump(snapshot, f, indent=2)
print(f"Snapshot: {SNAPSHOT}")

print(f"\nSummary:")
print(f"  Aggregator:      {len(records)} URLs")
print(f"  Recommendations: {len(recs)} actions ({sum(1 for r in recs if r['category']=='Technical')} tech / {sum(1 for r in recs if r['category']=='Content')} content)")
print(f"  Keywords:        {len(keywords_data)} keywords")
print(f"  Redirects:       {len(redirects)} URLs")
print(f"  Errors:          {len(errors)} URLs")
print(f"  Target Pages:    {len(target_top)} pages")
