#!/usr/bin/env python3
"""Parse the reviewed WQA audit xlsx and produce a structured approval summary.

Workflow:
  1. The strategist reviews the xlsx in Google Sheets.
  2. They mark Approval = Approved / Edit / Rejected / Deferred on each row.
  3. For Edit rows, they edit the "Specific Next Step" cell inline.
  4. They save the workbook back to .xlsx and overwrite the original.
  5. This script reads the workbook, buckets rows, and detects edits by diffing
     against {slug}-wqa-recommendations-original.json.

Outputs {slug}-approvals.json with shape:

{
  "audit_id":      "...",
  "totals":        {"Approved": n, "Edit": n, "Rejected": n, "Deferred": n, "Unmarked": n},
  "buckets": {
    "Approved":    [<row>, ...],
    "Edit":        [<row>, ...],
    "Rejected":    [<row>, ...],
    "Deferred":    [<row>, ...],
    "Unmarked":    [<row>, ...]
  },
  "edits_detected": [<row>, ...]   # rows where text changed vs snapshot (regardless of Approval value)
}

Each <row> has: row, approval, priority, effort, sprint, action_type, page_address, url,
                specific_next_step, original_text, edited (bool),
                current_state, target_state, why_it_matters, estimated_hours, owner,
                status, approver_notes
"""

import argparse
import json
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


parser = argparse.ArgumentParser()
parser.add_argument("--client-slug", required=True)
parser.add_argument("--audit-id", help="Audit UUID (folder under wqa/audits/)")
parser.add_argument("--audit-dir", help="Explicit audit folder path")
parser.add_argument("--workspace-root", default=os.environ.get("AGENCY_OS_ROOT", "."))
parser.add_argument("--root-domain", required=True,
                    help="Client root domain, e.g. https://theblueprint.training")
args = parser.parse_args()

SLUG = args.client_slug
ROOT = args.root_domain.rstrip("/")

if args.audit_dir:
    A_BASE = args.audit_dir
elif args.audit_id:
    A_BASE = os.path.join(args.workspace_root, "clients", SLUG, "wqa", "audits", args.audit_id)
else:
    sys.exit("Provide --audit-id or --audit-dir.")

XLSX     = os.path.join(A_BASE, f"{SLUG}-wqa-data.xlsx")
SNAPSHOT = os.path.join(A_BASE, f"{SLUG}-wqa-recommendations-original.json")
OUT      = os.path.join(A_BASE, f"{SLUG}-approvals.json")

assert os.path.exists(XLSX), f"xlsx not found: {XLSX}"


# Load snapshot keyed by row number
snapshot_by_row = {}
if os.path.exists(SNAPSHOT):
    for s in json.load(open(SNAPSHOT)):
        snapshot_by_row[s["row"]] = s
else:
    print(f"Warning: no snapshot at {SNAPSHOT}. Edit detection will be unavailable.", file=sys.stderr)


# Open workbook & locate Recommendations tab
wb = load_workbook(XLSX, data_only=True)
if "Recommendations" not in wb.sheetnames:
    sys.exit("Recommendations tab missing from workbook.")
ws = wb["Recommendations"]


# Header lookup (1-indexed)
headers = {cell.value: cell.column for cell in ws[1] if cell.value}
required = ["#", "Approval", "Priority", "Category", "Sprint", "Action Type",
            "Page Address", "Specific Next Step"]
missing = [c for c in required if c not in headers]
if missing:
    sys.exit(f"Recommendations tab missing required columns: {missing}")


# ---------- Aggregator overrides ----------
# If the strategist edited Technical Action / Content Action / Priority / Sprint on the
# Aggregator tab, those edits are authoritative — they override the Recommendations tab.
# We index Aggregator by normalized URL, then merge into each Recommendation row below.
#
# Disambiguation rule: the Aggregator has ONE action per category per URL, but the
# Recommendations tab can have MULTIPLE rows per (URL × category). To avoid
# false-positive overrides, we ONLY apply the Aggregator's action override when:
#   - The URL has exactly ONE Recommendation row in that category, OR
#   - The Aggregator's action matches one of the Rec rows (meaning the others are
#     untouched and don't get overridden)
# Priority and Sprint overrides apply per-row regardless.

aggregator_overrides = {}  # norm_url -> {"tech_action", "content_action", "priority", "sprint"}
if "Aggregator" in wb.sheetnames:
    ws_agg = wb["Aggregator"]
    agg_hdrs = {c.value: c.column for c in ws_agg[1] if c.value}
    needed = ["Technical Action", "Content Action", "Priority", "Sprint", "Address"]
    if all(n in agg_hdrs for n in needed):
        for ar in ws_agg.iter_rows(min_row=2, max_row=ws_agg.max_row, values_only=False):
            addr_cell = ar[agg_hdrs["Address"] - 1]
            if not addr_cell.value: continue
            url = addr_cell.hyperlink.target if addr_cell.hyperlink else None
            key = (url or "").rstrip("/").lower().split("?")[0].split("#")[0]
            if not key: continue
            aggregator_overrides[key] = {
                "tech_action": ar[agg_hdrs["Technical Action"] - 1].value,
                "content_action": ar[agg_hdrs["Content Action"] - 1].value,
                "priority": ar[agg_hdrs["Priority"] - 1].value,
                "sprint": ar[agg_hdrs["Sprint"] - 1].value,
            }

# First pass: build per-URL aggregates for override disambiguation.
# - url_cat_counts: how many rec rows of each category per URL
# - url_priorities, url_sprints: the SET of distinct values across rec rows per URL
url_cat_counts = {}      # (norm_url, category) -> count
url_priorities = {}      # norm_url -> set of distinct Priority values across rec rows
url_sprints = {}         # norm_url -> set of distinct Sprint values across rec rows
for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    n_val = r[headers["#"] - 1].value
    if n_val is None or n_val == "": continue
    addr_cell = r[headers["Page Address"] - 1]
    url_raw = addr_cell.hyperlink.target if addr_cell.hyperlink else (addr_cell.value or "")
    nk = url_raw.rstrip("/").lower().split("?")[0].split("#")[0]
    cat = r[headers["Category"] - 1].value or ""
    url_cat_counts[(nk, cat)] = url_cat_counts.get((nk, cat), 0) + 1
    url_priorities.setdefault(nk, set()).add(r[headers["Priority"] - 1].value)
    url_sprints.setdefault(nk, set()).add(r[headers["Sprint"] - 1].value)


VALID_APPROVALS = {"Approved", "Edit", "Rejected", "Deferred"}
buckets = {k: [] for k in (*VALID_APPROVALS, "Unmarked")}
edits_detected = []


def reconstruct_url(address):
    """Address column has the stripped path; rebuild full URL for downstream use."""
    if not address: return None
    if address.startswith("http"): return address
    if address == "/": return ROOT + "/"
    if address.startswith("/"): return ROOT + address
    return address  # leave whatever it is


for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    n_cell = r[headers["#"] - 1]
    if n_cell.value is None or n_cell.value == "":
        continue
    try:
        n = int(n_cell.value)
    except (TypeError, ValueError):
        continue

    def cell(name):
        c = r[headers[name] - 1]
        return c.value

    approval        = (cell("Approval") or "").strip()
    next_step       = cell("Specific Next Step") or ""
    address         = cell("Page Address") or ""
    addr_cell       = r[headers["Page Address"] - 1]
    url             = addr_cell.hyperlink.target if addr_cell.hyperlink else reconstruct_url(address)

    snap = snapshot_by_row.get(n, {})
    original_text = snap.get("original_text", "")
    edited = bool(original_text) and (next_step.strip() != original_text.strip())

    # Apply Aggregator overrides — but ONLY for action_type when the URL has exactly
    # one Recommendation row in this category (otherwise the Aggregator's single value
    # can't unambiguously target one of multiple rows; the rule "the strategist
    # explicitly changed it" can't be inferred safely).
    rec_category = cell("Category")
    rec_action_type = cell("Action Type")
    rec_priority = cell("Priority")
    rec_sprint = cell("Sprint")
    override_notes = []
    norm_url = (url or "").rstrip("/").lower().split("?")[0].split("#")[0]
    agg = aggregator_overrides.get(norm_url, {}) if norm_url else {}
    if agg:
        agg_action_for_cat = agg.get("tech_action") if rec_category == "Technical" else agg.get("content_action")
        n_rows_in_cat = url_cat_counts.get((norm_url, rec_category), 0)
        # action_type override: only when URL has exactly one rec row in this category
        if n_rows_in_cat == 1 and agg_action_for_cat and agg_action_for_cat != rec_action_type:
            override_notes.append(f"action_type: {rec_action_type} → {agg_action_for_cat}")
            rec_action_type = agg_action_for_cat
        # Priority override: only when ALL rec rows for this URL had the same priority
        # originally (i.e., the detector set them uniformly) AND the Aggregator differs.
        # If rec rows have varied priorities, can't safely infer strategist intent.
        prio_set = url_priorities.get(norm_url, set())
        if (len(prio_set) == 1 and agg.get("priority")
                and agg["priority"] not in prio_set):
            override_notes.append(f"priority: {rec_priority} → {agg['priority']}")
            rec_priority = agg["priority"]
        # Sprint override: same rule.
        sprint_set = url_sprints.get(norm_url, set())
        if (len(sprint_set) == 1 and agg.get("sprint")
                and agg["sprint"] not in sprint_set):
            override_notes.append(f"sprint: {rec_sprint} → {agg['sprint']}")
            rec_sprint = agg["sprint"]

    # Auto-derive Approval when the strategist left the cell empty. The
    # Aggregator tab already captures intent; the Approval column is only needed
    # to override it. Rules (empty cell only):
    #   action_type == "Leave as is" -> Rejected; row edited inline -> Edit;
    #   otherwise -> Approved. An explicit verdict always wins.
    auto_derived = False
    auto_derive_reason = None
    if not approval:
        if (rec_action_type or "").strip().lower() == "leave as is":
            approval = "Rejected"
            auto_derive_reason = "action_type=Leave as is -> Rejected"
        elif edited:
            approval = "Edit"
            auto_derive_reason = "inline text edit detected -> Edit"
        else:
            approval = "Approved"
            auto_derive_reason = "default-approve (no strategist override)"
        auto_derived = True

    row_data = {
        "row": n,
        "approval": approval or None,
        "priority": rec_priority,
        "category": rec_category,
        "sprint":   rec_sprint,
        "action_type": rec_action_type,
        "page_address": address,
        "url": url,
        "specific_next_step": next_step,
        "original_text": original_text,
        "edited": edited,
        "aggregator_overrides_applied": override_notes if override_notes else None,
        "auto_derived_approval": auto_derived,
        "auto_derive_reason": auto_derive_reason,
    }

    if edited:
        edits_detected.append(row_data)

    if approval in VALID_APPROVALS:
        buckets[approval].append(row_data)
    else:
        buckets["Unmarked"].append(row_data)


totals = {k: len(v) for k, v in buckets.items()}

# ---------- Target Pages tab (link-building roster) ----------
target_buckets = {k: [] for k in (*VALID_APPROVALS, "Unmarked")}
if "Target Pages" in wb.sheetnames:
    ws_tp = wb["Target Pages"]
    tp_headers = {c.value: c.column for c in ws_tp[1] if c.value}
    for r in ws_tp.iter_rows(min_row=2, max_row=ws_tp.max_row, values_only=False):
        if r[tp_headers["#"] - 1].value is None: continue
        approval = (r[tp_headers["Approval"] - 1].value or "").strip()
        addr_cell = r[tp_headers["Page Address"] - 1]
        url = addr_cell.hyperlink.target if addr_cell.hyperlink else reconstruct_url(addr_cell.value or "")
        # Target Pages auto-derive: the top-20 roster is auto-scored, so leaving
        # a row untouched implies acceptance → default to Approved.
        auto_derived_tp = False
        auto_derive_reason_tp = None
        if not approval:
            approval = "Approved"
            auto_derived_tp = True
            auto_derive_reason_tp = "default-approve (no strategist override)"
        target_data = {
            "rank": r[tp_headers["#"] - 1].value,
            "approval": approval or None,
            "page_address": addr_cell.value,
            "url": url,
            "page_type": r[tp_headers["Page Type"] - 1].value,
            "target_keyword": r[tp_headers["Target Keyword"] - 1].value,
            "kw_volume": r[tp_headers["KW Volume"] - 1].value,
            "current_position": r[tp_headers["Current Pos"] - 1].value,
            "ref_domains": r[tp_headers["Ref Domains"] - 1].value,
            "word_count": r[tp_headers["Word Count"] - 1].value,
            "gsc_impressions": r[tp_headers["GSC Impressions"] - 1].value,
            "score": r[tp_headers["Score"] - 1].value,
            "approver_notes": r[tp_headers["Approver Notes"] - 1].value,
            "auto_derived_approval": auto_derived_tp,
            "auto_derive_reason": auto_derive_reason_tp,
        }
        if approval in VALID_APPROVALS:
            target_buckets[approval].append(target_data)
        else:
            target_buckets["Unmarked"].append(target_data)

target_totals = {k: len(v) for k, v in target_buckets.items()}

result = {
    "audit_id": os.path.basename(A_BASE),
    "client_slug": SLUG,
    "xlsx_path": XLSX,
    "totals": totals,
    "buckets": buckets,
    "edits_detected": edits_detected,
    "target_pages": {
        "totals": target_totals,
        "buckets": target_buckets,
    },
}

with open(OUT, "w") as f:
    json.dump(result, f, indent=2)


# ============ Write final xlsx back ============
# Update the Recommendations tab in-place to reflect Aggregator overrides
# (Action Type / Priority / Sprint) and any edited Specific Next Step text.
# Overridden cells get a light blue background so the strategist can see what
# changed when they re-open the file. Saved back to the same path = single
# source of truth.

OVERRIDE_FILL = PatternFill("solid", fgColor="DBEAFE")  # light blue
EDIT_FILL     = PatternFill("solid", fgColor="FEF3C7")  # soft yellow for edited text

# Iterate same way we read but match to the in-memory bucketed row_data so we
# know what overrides got applied. Build a lookup: row_number -> row_data
rows_by_num = {}
for bucket_name in ("Approved", "Edit", "Rejected", "Deferred", "Unmarked"):
    for rd in buckets.get(bucket_name, []):
        rows_by_num[rd["row"]] = rd

# Re-open workbook (data_only=True doesn't preserve formulas, but xlsx has no formulas)
wb_write = load_workbook(XLSX)
ws_write = wb_write["Recommendations"]
hdrs_write = {cell.value: cell.column for cell in ws_write[1] if cell.value}

override_count = 0
edit_text_count = 0

for r in ws_write.iter_rows(min_row=2, max_row=ws_write.max_row, values_only=False):
    n_cell = r[hdrs_write["#"] - 1]
    if n_cell.value is None or n_cell.value == "":
        continue
    try:
        n = int(n_cell.value)
    except (TypeError, ValueError):
        continue
    rd = rows_by_num.get(n)
    if not rd: continue

    overrides = rd.get("aggregator_overrides_applied") or []
    if overrides:
        # Update + highlight cells based on which fields were overridden
        for override_str in overrides:
            field, _, _ = override_str.partition(":")
            field = field.strip()
            if field == "action_type":
                cell = r[hdrs_write["Action Type"] - 1]
                cell.value = rd["action_type"]
                cell.fill = OVERRIDE_FILL
            elif field == "priority":
                cell = r[hdrs_write["Priority"] - 1]
                cell.value = rd["priority"]
                cell.fill = OVERRIDE_FILL
            elif field == "sprint":
                cell = r[hdrs_write["Sprint"] - 1]
                cell.value = rd["sprint"]
                cell.fill = OVERRIDE_FILL
        override_count += 1

    # Highlight edited Specific Next Step cells
    if rd.get("edited"):
        cell = r[hdrs_write["Specific Next Step"] - 1]
        cell.fill = EDIT_FILL
        edit_text_count += 1

# Drop a parse-log row at the top of an Audit Summary tab (create if missing)
if "Audit Summary" in wb_write.sheetnames:
    ws_summary = wb_write["Audit Summary"]
    # Clear existing content
    for row in ws_summary.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
else:
    ws_summary = wb_write.create_sheet("Audit Summary", 0)  # first tab

ws_summary.column_dimensions["A"].width = 38
ws_summary.column_dimensions["B"].width = 40

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="0a0e1a")

def _row(ws, row_i, label, value, label_bold=True):
    ws.cell(row=row_i, column=1).value = label
    if label_bold:
        ws.cell(row=row_i, column=1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=row_i, column=2).value = value

ws_summary.cell(row=1, column=1).value = "Audit Summary"
ws_summary.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14)
_row(ws_summary, 3, "Client", SLUG)
_row(ws_summary, 4, "Audit ID", os.path.basename(A_BASE))
_row(ws_summary, 5, "Last parsed", datetime.now().isoformat(timespec="seconds"))
_row(ws_summary, 7, "Approved", totals.get("Approved", 0))
_row(ws_summary, 8, "Edit (inline text changes)", totals.get("Edit", 0))
_row(ws_summary, 9, "Rejected", totals.get("Rejected", 0))
_row(ws_summary, 10, "Deferred", totals.get("Deferred", 0))
_row(ws_summary, 11, "Unmarked (blocks next phase)", totals.get("Unmarked", 0))
_row(ws_summary, 13, "Aggregator overrides applied", override_count)
_row(ws_summary, 14, "Recommendations with edited text", edit_text_count)
_row(ws_summary, 16, "Target Pages — Approved", target_totals.get("Approved", 0))
_row(ws_summary, 17, "Target Pages — Edit", target_totals.get("Edit", 0))
_row(ws_summary, 18, "Target Pages — Rejected", target_totals.get("Rejected", 0))
_row(ws_summary, 19, "Target Pages — Deferred", target_totals.get("Deferred", 0))
_row(ws_summary, 20, "Target Pages — Unmarked", target_totals.get("Unmarked", 0))

# Legend for cell highlighting
_row(ws_summary, 22, "Legend", "", label_bold=True)
ws_summary.cell(row=23, column=1).value = "Light blue cells"
ws_summary.cell(row=23, column=2).value = "Aggregator override applied here"
ws_summary.cell(row=23, column=2).fill = OVERRIDE_FILL
ws_summary.cell(row=24, column=1).value = "Soft yellow cells"
ws_summary.cell(row=24, column=2).value = "Specific Next Step was edited inline"
ws_summary.cell(row=24, column=2).fill = EDIT_FILL

wb_write.save(XLSX)
print(f"\nFinal xlsx written back to: {XLSX}")
print(f"  · {override_count} Aggregator overrides applied to Recommendations rows")
print(f"  · {edit_text_count} Recommendations rows had Specific Next Step edited inline")
print(f"  · Added/refreshed 'Audit Summary' tab at position 1")


# Human-readable summary
def fmt(n): return f"{n:>4}"

print()
print(f"Approval Summary — {SLUG} / audit {os.path.basename(A_BASE)}")
print("=" * 60)
print(f"  Approved: {fmt(totals['Approved'])}  → become deliverables")
print(f"  Edit:     {fmt(totals['Edit'])}  → use edited text, become deliverables")
print(f"  Rejected: {fmt(totals['Rejected'])}  → dropped")
print(f"  Deferred: {fmt(totals['Deferred'])}  → backlog")
print(f"  Unmarked: {fmt(totals['Unmarked'])}  → blocks next phase, surface to user")
print(f"  ----")
print(f"  Total:    {fmt(sum(totals.values()))}")
print()

if totals["Unmarked"]:
    print("⚠  Unmarked rows (need approval before next phase):")
    for row in buckets["Unmarked"][:10]:
        print(f"     #{row['row']:>3}  {row['action_type'] or '–':<18}  {row['page_address']}")
    if totals["Unmarked"] > 10:
        print(f"     … and {totals['Unmarked'] - 10} more")
    print()

if edits_detected:
    print(f"✏  Edits detected on {len(edits_detected)} row(s):")
    for row in edits_detected[:10]:
        print(f"     #{row['row']:>3}  {row['page_address']}")
        print(f"           was: {(row['original_text'] or '')[:80]}…")
        print(f"           now: {(row['specific_next_step'] or '')[:80]}…")
    if len(edits_detected) > 10:
        print(f"     … and {len(edits_detected) - 10} more")
    print()

print(f"Saved: {OUT}")
