#!/usr/bin/env python3
"""Merge Aggregator-only rows (strategist supplemental audit) into Recommendations
tab so the project plan covers both the automated WQA output and the manual
Aggregator findings that never got promoted to a Recommendations row."""
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font

XLSX = sys.argv[1]

wb = openpyxl.load_workbook(XLSX)
ws = wb["Recommendations"]
ws_agg = wb["Aggregator"]

headers = {c.value: c.column for c in ws[1] if c.value}
agg_headers = {c.value: c.column for c in ws_agg[1] if c.value}

# Existing Recommendations pages + max row number
rec_pages = set()
max_n = 0
for r in range(2, ws.max_row + 1):
    n = ws.cell(row=r, column=headers["#"]).value
    if n is None or n == "":
        continue
    max_n = max(max_n, int(n))
    addr = ws.cell(row=r, column=headers["Page Address"]).value
    if addr:
        rec_pages.add(addr)

def g(row, name):
    col = agg_headers.get(name)
    if not col:
        return None
    return ws_agg.cell(row=row, column=col).value

def fmt_num(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v

TECH_TEMPLATES = {
    "Indexability fix": lambda a: (
        f"Marked non-indexable ({a.get('indexable') or 'no'}) but live and structurally part of the site"
        f"{' (' + str(fmt_num(a['word_count'])) + 'w)' if a.get('word_count') else ''}. "
        "Review robots/meta directives and internal linking; make indexable if it should rank, "
        "or consolidate/noindex deliberately if it's genuinely low-value."
    ),
    "301 redirect": lambda a: (
        f"Add a 301 redirect to the correct live equivalent"
        + (f" ({a['redirect_target']})" if a.get("redirect_target") else "")
        + " to preserve SEO value and avoid duplicate/thin-content weight."
    ),
    "Monitor": lambda a: (
        "Flagged for monitoring — no immediate fix required, but track status/performance next audit cycle."
    ),
    "Noindex": lambda a: (
        "Low-value or auto-generated page with no meaningful content. Add noindex to remove from crawl waste."
    ),
    "Fix 404": lambda a: (
        "Returns a 4xx error. Restore the page or 301 redirect to the closest live equivalent."
    ),
}

CONTENT_TEMPLATES = {
    "Update onpage": lambda a: (
        "Update on-page elements (title/H1/meta as needed)"
        + (f" — current title: “{a['title']}”." if a.get("title") else ".")
        + " Align with primary search intent and target keyword."
    ),
    "Refresh content": lambda a: (
        f"Refresh content"
        + (f" ({fmt_num(a['word_count'])}w currently)" if a.get("word_count") else "")
        + " — update with current information, examples, and internal links to lift relevance."
    ),
    "Rewrite title/meta": lambda a: (
        "Rewrite title/meta to better match search intent and improve CTR."
    ),
    "Expand content": lambda a: (
        f"Thin content"
        + (f" ({fmt_num(a['word_count'])}w)" if a.get("word_count") else "")
        + ". Expand to 1,000w+ with supporting depth, or consolidate/noindex if low-value."
    ),
    "Consolidate": lambda a: (
        "Overlaps with another page on the site (thin/duplicate topical coverage). "
        "Consolidate into the stronger page and 301 redirect."
    ),
    "Rewrite": lambda a: (
        f"Thin page"
        + (f" ({fmt_num(a['word_count'])}w)" if a.get("word_count") else "")
        + ". Rewrite with conversion content, social proof, FAQs."
    ),
}

EXCLUDE_ACTIONS = {"Leave as is"}
EXCLUDE_SPRINTS = {None, "Done"}

def norm(addr):
    return (addr or "").rstrip("/").lower()

rec_pages_norm = {norm(p) for p in rec_pages}

new_rows = []  # (category, action_type, priority, sprint, address, next_step)
seen_addrs = set()
for r in range(2, ws_agg.max_row + 1):
    addr = g(r, "Address")
    if not addr:
        continue
    if norm(addr) in rec_pages_norm:
        continue
    sprint = g(r, "Sprint")
    if sprint in EXCLUDE_SPRINTS:
        continue
    priority = g(r, "Priority") or "P3"
    tech = g(r, "Technical Action")
    cont = g(r, "Content Action")
    a = {
        "word_count": g(r, "Word Count"),
        "title": g(r, "Title"),
        "indexable": g(r, "Indexable"),
        "redirect_target": g(r, "Redirect Target"),
        "problem_areas": g(r, "Problem Areas"),
    }
    if tech and tech not in EXCLUDE_ACTIONS:
        fn = TECH_TEMPLATES.get(tech)
        text = fn(a) if fn else f"{tech} — see strategist Aggregator audit for details."
        new_rows.append(("Technical", tech, priority, sprint, addr, text))
    if cont and cont not in EXCLUDE_ACTIONS:
        fn = CONTENT_TEMPLATES.get(cont)
        text = fn(a) if fn else f"{cont} — see strategist Aggregator audit for details."
        new_rows.append(("Content", cont, priority, sprint, addr, text))

print(f"Existing Recommendations rows: {max_n}")
print(f"New rows to append from Aggregator (strategist supplemental): {len(new_rows)}")
from collections import Counter
print("  by category:", Counter(r[0] for r in new_rows))
print("  by action:", Counter(r[1] for r in new_rows))
print("  by sprint:", Counter(r[3] for r in new_rows))

NEW_FILL = PatternFill("solid", fgColor="E9D8FD")  # light purple = strategist supplemental

n = max_n
for category, action_type, priority, sprint, addr, text in new_rows:
    n += 1
    row_i = ws.max_row + 1
    ws.cell(row=row_i, column=headers["#"]).value = n
    ws.cell(row=row_i, column=headers["Approval"]).value = "Approved"
    ws.cell(row=row_i, column=headers["Priority"]).value = priority
    ws.cell(row=row_i, column=headers["Category"]).value = category
    ws.cell(row=row_i, column=headers["Sprint"]).value = sprint
    ws.cell(row=row_i, column=headers["Action Type"]).value = action_type
    ws.cell(row=row_i, column=headers["Page Address"]).value = addr
    ws.cell(row=row_i, column=headers["Specific Next Step"]).value = text
    for col in headers.values():
        ws.cell(row=row_i, column=col).fill = NEW_FILL

wb.save(XLSX)
print(f"\nSaved. Recommendations tab now has {n} total rows ({max_n} original + {len(new_rows)} from Aggregator).")
